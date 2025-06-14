# --- 0. IMPORTAÇÕES DE BIBLIOTECAS ---
import pdfplumber
import google.generativeai as genai
import json
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, PhotoImage
import re
import os
import sys
import subprocess
import openpyxl
from dotenv import load_dotenv
import time
import logging
from tenacity import retry, stop_after_attempt, wait_exponential, RetryError
from typing import Any, Dict, List, Optional, Tuple, Union
from PIL import Image, ImageTk 

# --- HELPER PARA PYINSTALLER ---
def resource_path(relative_path: str) -> Path:
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = Path(sys._MEIPASS)
    except AttributeError:
        base_path = Path(__file__).resolve().parent
    return base_path / relative_path

# --- CONSTANTES ---
CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS = False
ARQUIVO_MAPEAMENTO_CONFIG = "mapeamento_config.json"
ARQUIVO_SCHEMA_EXTRACAO = "extraction_schema.json"
LOG_FILE_PATH = resource_path("processamento_pdf.log")
MAX_TEXT_LENGTH_IA = 60
MARCADOR_INICIO_TEXTO_PDF_PROMPT = "[INICIO_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
MARCADOR_FIM_TEXTO_PDF_PROMPT = "[FIM_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES = True
GEMINI_TEMPERATURE = 0.1
GEMINI_MAX_OUTPUT_TOKENS = 8190
GEMINI_MODEL_NAME = 'gemini-1.5-flash-latest'

# --- CORES E FONTES PARA A GUI ---
COR_FUNDO_JANELA = "#F0F0F0"
COR_FUNDO_FRAMES_INTERNOS = "#F8F8F8"
COR_TEXTO_PADRAO = "#333333"
COR_TEXTO_TITULO_GUI = "#333333"
COR_TEXTO_LOG = "#1A1A1A"
COR_FUNDO_LOG = "#FFFFFF"
COR_BOTAO_PRIMARIO_BG = "#0078D4"
COR_BOTAO_PRIMARIO_FG = "#FFFFFF"
COR_BOTAO_SECUNDARIO_BG = "#FFC107"
COR_BOTAO_SECUNDARIO_FG = "#000000"
COR_BOTAO_HOVER_PRIMARIO = "#005FA3"
COR_BOTAO_HOVER_SECUNDARIO = "#FFD54F"
COR_STATUS_LABEL_FG = "#005A9E"
FONTE_TITULO_APP = ("Segoe UI", 18, "bold")
FONTE_SUBTITULO = ("Segoe UI", 10)
FONTE_BOTAO_PRINCIPAL = ("Segoe UI", 10, "bold")
FONTE_BOTAO_SECUNDARIO = ("Segoe UI", 10, "bold")
FONTE_STATUS = ("Segoe UI", 10, "italic")
FONTE_LOG = ("Consolas", 9)
FONTE_MENU = ("Segoe UI", 9)
FONTE_LABELFRAME_TITULO = ("Segoe UI", 10, "bold")
PADDING_X_BOTAO_STYLE = 15
PADDING_Y_BOTAO_STYLE = 5

# --- CONFIGURAÇÃO DO LOGGING ---
# Mova esta configuração para o mais cedo possível, após os imports e resource_path
logging.basicConfig(
    level=logging.DEBUG, # Use DEBUG para mais detalhes durante o desenvolvimento
    format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(lineno)d - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, encoding='utf-8', mode='w'), # 'w' para log limpo a cada execução (mudar para 'a' em produção)
        logging.StreamHandler(sys.stdout)
    ]
)
logging.info("!!! CONFIGURAÇÃO DE LOGGING EXECUTADA NO TOPO DO SCRIPT !!!")


# --- 4. FUNÇÕES DE FEEDBACK DA INTERFACE GRÁFICA (log, progresso) ---
_root_ref_for_log = None
_log_text_widget_ref = None

def setup_gui_logging_refs(root_window, log_widget):
    global _root_ref_for_log, _log_text_widget_ref
    _root_ref_for_log = root_window
    _log_text_widget_ref = log_widget

def log_to_gui(mensagem: str, level: str = "INFO") -> None:
    formatted_message = f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {level} - {mensagem}"
    if _root_ref_for_log and _root_ref_for_log.winfo_exists() and _log_text_widget_ref:
        _log_text_widget_ref.configure(state=tk.NORMAL)
        _log_text_widget_ref.insert(tk.END, formatted_message + "\n")
        _log_text_widget_ref.see(tk.END)
        _log_text_widget_ref.configure(state=tk.DISABLED)
        _root_ref_for_log.update_idletasks()
    else:
        if level == "INFO": logging.info(mensagem)
        elif level == "WARNING": logging.warning(mensagem)
        elif level == "ERROR": logging.error(mensagem)
        elif level == "DEBUG": logging.debug(mensagem)
        elif level == "CRITICAL": logging.critical(mensagem)
        else: logging.info(f"({level}) {mensagem}")

    if not (_root_ref_for_log and _root_ref_for_log.winfo_exists() and _log_text_widget_ref):
        pass
    elif level == "INFO": logging.info(mensagem)
    elif level == "WARNING": logging.warning(mensagem)
    elif level == "ERROR": logging.error(mensagem)
    elif level == "DEBUG": logging.debug(mensagem)
    elif level == "CRITICAL": logging.critical(mensagem)


def iniciar_progresso() -> None:
    if _root_ref_for_log and _root_ref_for_log.winfo_exists() and 'progress' in globals() and 'status_label' in globals():
        progress.start(20)
        status_label.config(text="Processando, por favor aguarde...")
        _root_ref_for_log.update_idletasks()

def parar_progresso(final_status: str = "") -> None:
    if _root_ref_for_log and _root_ref_for_log.winfo_exists() and 'progress' in globals() and 'status_label' in globals():
        progress.stop()
        if final_status: status_label.config(text=final_status)
        else: status_label.config(text="Pronto.")
        _root_ref_for_log.update_idletasks()

def limpar_log_gui() -> None:
    if _root_ref_for_log and _root_ref_for_log.winfo_exists() and _log_text_widget_ref:
        confirmed = messagebox.askyesno(
            "Limpar Log",
            "Tem certeza que deseja limpar todo o log da tela?\n(O arquivo de log em disco não será afetado)",
            parent=_root_ref_for_log
        )
        if confirmed:
            _log_text_widget_ref.configure(state=tk.NORMAL)
            _log_text_widget_ref.delete('1.0', tk.END)
            _log_text_widget_ref.configure(state=tk.DISABLED)
            log_to_gui("Log da tela limpo pelo usuário.", "INFO")
    else:
        logging.warning("Tentativa de limpar log da GUI, mas widget não disponível.")

# --- 1. CARREGAR VARIÁVEIS DE AMBIENTE (.env) ---
dotenv_path_obj = resource_path(".env")
if dotenv_path_obj.exists():
    load_dotenv(dotenv_path=dotenv_path_obj)
    logging.info(f"Arquivo .env carregado de: {dotenv_path_obj}")
else:
    logging.warning(f"Arquivo .env não encontrado em: {dotenv_path_obj}. Usando variáveis de ambiente do sistema se disponíveis.")

# --- 2. CONFIGURAÇÃO DA API KEY DO GOOGLE ---
GOOGLE_API_KEY: Optional[str] = os.getenv("GOOGLE_API_KEY")

# --- BLOCO DE CONFIGURAÇÃO (CARREGADO DE ARQUIVO EXTERNO) ---
BLOCO_CONFIG: Dict[str, Any] = {}

def carregar_schema_extracao() -> bool:
    global BLOCO_CONFIG
    caminho_schema_abs = resource_path(ARQUIVO_SCHEMA_EXTRACAO)
    try:
        with open(caminho_schema_abs, "r", encoding="utf-8") as f:
            BLOCO_CONFIG_RAW = json.load(f)

        if not isinstance(BLOCO_CONFIG_RAW, dict):
            log_to_gui(f"ERRO CRÍTICO: Conteúdo do schema '{ARQUIVO_SCHEMA_EXTRACAO}' não é um dicionário JSON no nível raiz.", "CRITICAL")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                messagebox.showerror("Erro de Configuração", f"O arquivo '{ARQUIVO_SCHEMA_EXTRACAO}' não contém um dicionário JSON válido no nível raiz.", parent=_root_ref_for_log)
            return False

        BLOCO_CONFIG_VALIDADO: Dict[str, Any] = {}
        schema_geral_valido = True

        for nome_bloco, config_bloco in BLOCO_CONFIG_RAW.items():
            bloco_atual_valido = True
            if not isinstance(config_bloco, dict):
                log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}' em '{ARQUIVO_SCHEMA_EXTRACAO}' não é um dicionário. Bloco ignorado.", "ERROR")
                schema_geral_valido = False
                continue

            chaves_obrigatorias = {"json_chave": str, "particao": int}
            for chave, tipo_esperado in chaves_obrigatorias.items():
                if chave not in config_bloco:
                    log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}' não possui a chave obrigatória '{chave}'. Bloco ignorado.", "ERROR")
                    bloco_atual_valido = False; break
                if not isinstance(config_bloco[chave], tipo_esperado):
                    log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}', chave '{chave}', esperava tipo '{tipo_esperado.__name__}' mas obteve '{type(config_bloco[chave]).__name__}'. Bloco ignorado.", "ERROR")
                    bloco_atual_valido = False; break

            if not bloco_atual_valido:
                schema_geral_valido = False; continue

            if config_bloco.get("particao", 0) <= 0:
                log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}', chave 'particao' deve ser um inteiro maior que 0. Valor: {config_bloco.get('particao')}. Bloco ignorado.", "ERROR")
                schema_geral_valido = False; continue

            chaves_opcionais_com_tipo = {
                "titulo_padrao": str, "campos_esperados": list, "nome_lista_json": str,
                "sub_campos_lista": list, "campos_texto_longo_limitar": list, "sub_lista_aninhada": dict
            }
            for chave, tipo_esperado in chaves_opcionais_com_tipo.items():
                if chave in config_bloco and config_bloco[chave] is not None and not isinstance(config_bloco[chave], tipo_esperado):
                    log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', chave opcional '{chave}', esperava tipo '{tipo_esperado.__name__}' mas obteve '{type(config_bloco[chave]).__name__}'.", "WARNING")

            if "sub_lista_aninhada" in config_bloco and isinstance(config_bloco.get("sub_lista_aninhada"), dict):
                sub_lista_conf = config_bloco["sub_lista_aninhada"]
                sub_chaves_obrigatorias = {"nome_json": str, "campos": list}
                for sub_chave, sub_tipo_esperado in sub_chaves_obrigatorias.items():
                    if sub_chave not in sub_lista_conf:
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada' não possui a chave '{sub_chave}'.", "WARNING"); break
                    if not isinstance(sub_lista_conf[sub_chave], sub_tipo_esperado):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada', chave '{sub_chave}', esperava tipo '{sub_tipo_esperado.__name__}' mas obteve '{type(sub_lista_conf[sub_chave]).__name__}'.", "WARNING"); break
                    if sub_chave == "campos" and isinstance(sub_lista_conf[sub_chave], list) and \
                       not all(isinstance(item, str) for item in sub_lista_conf[sub_chave]):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada', chave 'campos' deve conter apenas strings.", "WARNING")

            for chave_lista_str in ["campos_esperados", "sub_campos_lista", "campos_texto_longo_limitar"]:
                if chave_lista_str in config_bloco and config_bloco[chave_lista_str] is not None and isinstance(config_bloco.get(chave_lista_str), list):
                    if not all(isinstance(item, str) for item in config_bloco[chave_lista_str]):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', chave '{chave_lista_str}' deve conter apenas strings.", "WARNING")

            if bloco_atual_valido:
                BLOCO_CONFIG_VALIDADO[nome_bloco] = config_bloco
            else:
                schema_geral_valido = False

        BLOCO_CONFIG = BLOCO_CONFIG_VALIDADO

        if not schema_geral_valido and BLOCO_CONFIG_RAW:
            log_to_gui(f"AVISO: Schema '{ARQUIVO_SCHEMA_EXTRACAO}' contém erros/avisos. Alguns blocos podem ter sido ignorados.", "WARNING")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                 messagebox.showwarning("Aviso de Schema", f"Schema '{ARQUIVO_SCHEMA_EXTRACAO}' com erros/avisos. Verifique logs.", parent=_root_ref_for_log)

        if not BLOCO_CONFIG and BLOCO_CONFIG_RAW:
            log_to_gui(f"ERRO FATAL: Nenhum bloco válido no schema '{ARQUIVO_SCHEMA_EXTRACAO}'.", "CRITICAL")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                messagebox.showerror("Erro Fatal de Schema", f"Nenhum bloco válido no schema '{ARQUIVO_SCHEMA_EXTRACAO}'.", parent=_root_ref_for_log)
            return False
        elif not BLOCO_CONFIG and not BLOCO_CONFIG_RAW:
             log_to_gui(f"INFO: Schema '{ARQUIVO_SCHEMA_EXTRACAO}' está vazio.", "INFO")
        else:
            log_to_gui(f"Schema '{ARQUIVO_SCHEMA_EXTRACAO}' carregado. Blocos válidos: {len(BLOCO_CONFIG)}", "INFO")
        return True
    except FileNotFoundError:
        log_to_gui(f"ERRO CRÍTICO: Schema '{ARQUIVO_SCHEMA_EXTRACAO}' não encontrado em '{caminho_schema_abs}'.", "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Configuração", f"Arquivo '{ARQUIVO_SCHEMA_EXTRACAO}' não encontrado.", parent=_root_ref_for_log)
        return False
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO CRÍTICO: Falha JSON no schema '{ARQUIVO_SCHEMA_EXTRACAO}': {e}", "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Configuração", f"Erro ao ler '{ARQUIVO_SCHEMA_EXTRACAO}'. JSON inválido: {e}", parent=_root_ref_for_log)
        return False
    except Exception as e:
        log_to_gui(f"ERRO CRÍTICO ao carregar schema '{ARQUIVO_SCHEMA_EXTRACAO}': {e}", "CRITICAL")
        logging.exception("Erro detalhado ao carregar schema:")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Configuração", f"Erro fatal ao carregar '{ARQUIVO_SCHEMA_EXTRACAO}': {e}", parent=_root_ref_for_log)
        return False

# --- DEFINIÇÃO DAS PARTIÇÕES DO BLOCO_CONFIG (DINÂMICO) ---
LISTA_DE_NOMES_BLOCOS_PARTICIONADA: List[List[str]] = []
def gerar_particoes_dinamicamente() -> bool:
    global LISTA_DE_NOMES_BLOCOS_PARTICIONADA
    if not BLOCO_CONFIG:
        log_to_gui("AVISO: Schema (BLOCO_CONFIG) vazio. Sem partições.", "WARNING")
        LISTA_DE_NOMES_BLOCOS_PARTICIONADA = []
        return True

    max_particao = 0
    for config_bloco in BLOCO_CONFIG.values():
        num_part = config_bloco.get("particao")
        if isinstance(num_part, int) and num_part > 0:
            if num_part > max_particao:
                max_particao = num_part

    if max_particao == 0:
        all_block_names = list(BLOCO_CONFIG.keys())
        if all_block_names:
            log_to_gui(f"AVISO: Nenhuma 'particao' > 0 no schema. Usando todos os {len(all_block_names)} blocos em uma partição.", "WARNING")
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [all_block_names]
        else:
            log_to_gui("AVISO: Nenhum bloco no schema para particionar.", "WARNING")
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA = []
        return True

    LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [[] for _ in range(max_particao)]
    blocos_sem_particao_valida = []

    for nome_bloco, config_bloco in BLOCO_CONFIG.items():
        num_part = config_bloco.get("particao")
        if isinstance(num_part, int) and 1 <= num_part <= max_particao:
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA[num_part - 1].append(nome_bloco)
        else:
            blocos_sem_particao_valida.append(nome_bloco)

    if blocos_sem_particao_valida:
        log_to_gui(f"AVISO: Blocos sem 'particao' válida e ignorados no particionamento: {', '.join(blocos_sem_particao_valida)}", "WARNING")

    LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [p for p in LISTA_DE_NOMES_BLOCOS_PARTICIONADA if p]

    if not LISTA_DE_NOMES_BLOCOS_PARTICIONADA and BLOCO_CONFIG:
         log_to_gui("ERRO: Nenhuma partição gerada, mas schema não vazio. Verifique 'particao' nos blocos.", "ERROR")
         return False

    log_to_gui(f"Partições geradas: {len(LISTA_DE_NOMES_BLOCOS_PARTICIONADA)} ativas.", "INFO")
    for i, part in enumerate(LISTA_DE_NOMES_BLOCOS_PARTICIONADA):
        log_to_gui(f"  Partição {i+1} com {len(part)} blocos.", "DEBUG")
    return True


# --- 3. CONFIGURAÇÃO INICIAL DA INTERFACE GRÁFICA (Widgets Globais) ---
root = tk.Tk()
_root_ref_for_log = root

root.title(f"Processador de Súmulas de Crédito - Gemini v1.5")
root.geometry("900x750")
root.configure(bg=COR_FUNDO_JANELA)

style = ttk.Style()
style.theme_use('clam')
style.configure("TButton", padding=6, relief="flat", borderwidth=0)
style.configure("TLabel", font=FONTE_SUBTITULO, background=COR_FUNDO_JANELA, foreground=COR_TEXTO_PADRAO)
style.configure("TProgressbar", thickness=15, background=COR_BOTAO_PRIMARIO_BG)
style.configure("Title.TLabel", font=FONTE_TITULO_APP, foreground=COR_TEXTO_TITULO_GUI, background=COR_FUNDO_JANELA)
style.configure("Header.TFrame", background=COR_FUNDO_JANELA)
style.configure("Controls.TFrame", background=COR_FUNDO_JANELA)
style.configure("Status.TFrame", background=COR_FUNDO_JANELA)
style.configure("TLabelframe", background=COR_FUNDO_FRAMES_INTERNOS, relief=tk.GROOVE, borderwidth=1)
style.configure("TLabelframe.Label", font=FONTE_LABELFRAME_TITULO, background=COR_FUNDO_FRAMES_INTERNOS, foreground=COR_TEXTO_PADRAO, padding=(5,2))

frame_topo = ttk.Frame(root, padding=(20, 10), style="Header.TFrame")
frame_topo.pack(pady=(10,0), fill=tk.X)

# --- CARREGAR E POSICIONAR LOGO E TÍTULO ---
logging.info("--- INICIANDO BLOCO DE CARREGAMENTO DO LOGO PRINCIPAL (GUI) ---")
frame_titulo_com_logo = ttk.Frame(frame_topo, style="Header.TFrame")
# Centraliza o frame que contém o logo e o título
frame_titulo_com_logo.pack(pady=(0, 5), anchor=tk.CENTER)


logo_app_image_tk: Optional[ImageTk.PhotoImage] = None
caminho_logo_relativo = "assets/logo_sicoob.png"
NOVO_LARGURA_LOGO = 45
NOVO_ALTURA_LOGO = 45

try:
    logging.info("Dentro do TRY para carregar logo (GUI).")
    caminho_logo_abs = resource_path(caminho_logo_relativo)
    logging.info(f"resource_path (GUI) retornou: {caminho_logo_abs} (Tentando carregar logo deste caminho)")

    if caminho_logo_abs.is_file():
        logging.info(f"Arquivo '{caminho_logo_abs}' (GUI) EXISTE.")
        img_pil = Image.open(caminho_logo_abs)
        logging.info("Image.open() (GUI) SUCESSO.")
        try:
            img_redimensionada_pil = img_pil.resize((NOVO_LARGURA_LOGO, NOVO_ALTURA_LOGO), Image.Resampling.LANCZOS)
        except AttributeError:
            img_redimensionada_pil = img_pil.resize((NOVO_LARGURA_LOGO, NOVO_ALTURA_LOGO), Image.ANTIALIAS)
        logging.info("Redimensionamento (GUI) SUCESSO.")

        logo_app_image_tk = ImageTk.PhotoImage(img_redimensionada_pil)
        logging.info("ImageTk.PhotoImage() (GUI) SUCESSO.")

        # CRIA O LABEL DO LOGO PRIMEIRO
        label_logo = ttk.Label(frame_titulo_com_logo, image=logo_app_image_tk) # Removido style="Header.TFrame" temporariamente para simplificar
        label_logo.image = logo_app_image_tk # Manter referência!
        # EMPACOTA O LOGO À ESQUERDA
        label_logo.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        logging.info(f"Logo '{caminho_logo_relativo}' (GUI) carregada e posicionada à ESQUERDA.")
    else:
        logging.warning(f"Logo '{caminho_logo_abs}' (GUI) NÃO ENCONTRADA.")
        log_to_gui(f"AVISO: Logo '{caminho_logo_abs}' NÃO ENCONTRADA.", "WARNING") # Log para GUI também
except Exception as e_logo:
    logging.error(f"Erro ao carregar/redimensionar logo '{caminho_logo_relativo}' (GUI): {e_logo}", exc_info=True)
    log_to_gui(f"AVISO: Erro ao carregar/redimensionar logo '{caminho_logo_relativo}': {e_logo}", "WARNING") # Log para GUI também

# CRIA O LABEL DO TÍTULO DEPOIS
label_titulo_app = ttk.Label(frame_titulo_com_logo, text="Processador de Súmulas de Crédito", style="Title.TLabel")
# EMPACOTA O TÍTULO À ESQUERDA (FICARÁ À DIREITA DO LOGO)
label_titulo_app.pack(side=tk.LEFT, pady=5)
logging.info("--- FIM DO BLOCO DE CARREGAMENTO DO LOGO PRINCIPAL (GUI) ---")


# Label de instrução abaixo do conjunto logo+título
label_instrucao_app = ttk.Label(frame_topo, text="Selecione um PDF para análise. O log aparecerá abaixo.", justify=tk.CENTER)
label_instrucao_app.pack(pady=(5,10))
# --- FIM LOGO E TÍTULO ---


frame_botoes_principais = ttk.Frame(frame_topo, style="Controls.TFrame")
frame_botoes_principais.pack(pady=10, fill=tk.X, padx=20)
frame_status_progresso = ttk.Frame(root, padding=(20,5), style="Status.TFrame")
frame_status_progresso.pack(fill=tk.X, padx=10)
progress = ttk.Progressbar(frame_status_progresso, orient="horizontal", length=500, mode="indeterminate")
progress.pack(pady=5, fill=tk.X, expand=True)
status_label = ttk.Label(frame_status_progresso, text="Pronto para iniciar.", foreground=COR_STATUS_LABEL_FG, font=FONTE_STATUS, anchor=tk.CENTER)
status_label.pack(pady=5, fill=tk.X)
log_labelframe = ttk.LabelFrame(root, text=" Log de Processamento ", padding=(10,5))
log_labelframe.pack(pady=(5,10), padx=10, fill=tk.BOTH, expand=True)
log_text_widget = tk.Text(log_labelframe, height=15, width=90, wrap=tk.WORD, font=FONTE_LOG, bg=COR_FUNDO_LOG, fg=COR_TEXTO_LOG, relief=tk.SOLID, bd=1, state=tk.DISABLED)
log_scrollbar_y = ttk.Scrollbar(log_labelframe, orient="vertical", command=log_text_widget.yview)
log_text_widget.config(yscrollcommand=log_scrollbar_y.set)
log_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=5, padx=(0,5))
log_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

setup_gui_logging_refs(root, log_text_widget)


# --- 5. CONFIGURAÇÃO DA API GOOGLE GEMINI ---
genai_config_ok = False
if not GOOGLE_API_KEY:
    log_to_gui("ERRO CRÍTICO: Variável de ambiente 'GOOGLE_API_KEY' não encontrada.", "CRITICAL")
elif GOOGLE_API_KEY == "SUA_CHAVE_DE_API_AQUI":
    log_to_gui("ERRO CRÍTICO: GOOGLE_API_KEY no .env é placeholder. Configure-a.", "CRITICAL")
else:
    try:
        genai.configure(api_key=GOOGLE_API_KEY)
        log_to_gui("API Key do Google Gemini configurada com sucesso.", "INFO")
        genai_config_ok = True
    except Exception as e:
        log_to_gui(f"ERRO ao configurar API Key Google Gemini: {e}.", "ERROR")
        genai_config_ok = False

# --- FUNÇÕES PARA CARREGAR E SALVAR MAPEAMENTO DE CHAVES ---
def carregar_mapeamento_de_arquivo(caminho_arquivo_str: str) -> Optional[Dict[str, Any]]:
    caminho_arquivo_abs = resource_path(caminho_arquivo_str)
    if not caminho_arquivo_abs.is_file():
        log_to_gui(f"Arquivo de mapeamento '{caminho_arquivo_abs.name}' não encontrado.", "WARNING")
        return None
    try:
        with open(caminho_arquivo_abs, "r", encoding="utf-8") as f:
            mapeamento = json.load(f)
        log_to_gui(f"Mapeamento carregado de '{caminho_arquivo_abs.name}'.", "INFO")
        if not isinstance(mapeamento, dict):
            log_to_gui(f"ERRO: Mapeamento '{caminho_arquivo_abs.name}' não é dicionário.", "ERROR")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Mapeamento", f"'{caminho_arquivo_abs.name}' não é dicionário JSON.", parent=_root_ref_for_log)
            return None
        return mapeamento
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO JSON no mapeamento '{caminho_arquivo_abs.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Mapeamento", f"Erro JSON em '{caminho_arquivo_abs.name}': {e}", parent=_root_ref_for_log)
        return None
    except Exception as e:
        log_to_gui(f"ERRO ao carregar mapeamento '{caminho_arquivo_abs.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Mapeamento", f"Erro ao carregar '{caminho_arquivo_abs.name}': {e}", parent=_root_ref_for_log)
        return None

def salvar_mapeamento_em_arquivo(mapeamento: Dict[str, Any], caminho_arquivo_str: str) -> bool:
    caminho_arquivo_abs = resource_path(caminho_arquivo_str)
    try:
        caminho_arquivo_abs.parent.mkdir(parents=True, exist_ok=True)
        with open(caminho_arquivo_abs, "w", encoding="utf-8") as f:
            json.dump(mapeamento, f, indent=2, ensure_ascii=False)
        log_to_gui(f"Mapeamento salvo em '{caminho_arquivo_abs.name}'.", "INFO")
        return True
    except Exception as e:
        log_to_gui(f"ERRO ao salvar mapeamento '{caminho_arquivo_abs.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar mapeamento '{caminho_arquivo_abs.name}': {e}", parent=_root_ref_for_log)
        return False

# --- 6. FUNÇÕES AUXILIARES DE PROCESSAMENTO ---
def extrair_texto_do_pdf(caminho_pdf: Path) -> Optional[str]:
    texto_completo = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            log_to_gui(f"Extraindo texto de '{caminho_pdf.name}' ({len(pdf.pages)} páginas)...", "INFO")
            for i, pagina in enumerate(pdf.pages):
                if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                    status_label.config(text=f"Extraindo texto da página {i+1}/{len(pdf.pages)}...")
                    _root_ref_for_log.update_idletasks()

                texto_pagina_stream = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=False)
                texto_pagina_layout = None

                if not texto_pagina_stream or len(texto_pagina_stream.split()) < 5:
                    texto_pagina_layout = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=True)
                
                texto_pagina = texto_pagina_stream
                if texto_pagina_layout and (not texto_pagina_stream or len(texto_pagina_layout.split()) > len(texto_pagina_stream.split())):
                    log_to_gui(f"Página {i+1}: Usando extração com layout=True.", "DEBUG")
                    texto_pagina = texto_pagina_layout
                
                if texto_pagina:
                    texto_completo += texto_pagina if texto_pagina.endswith("\n") else texto_pagina + "\n"
                else:
                    log_to_gui(f"Página {i+1} de '{caminho_pdf.name}' não retornou texto.", "DEBUG")

            log_to_gui("Extração de texto PDF concluída.", "INFO")
            texto_limpo_para_ia = texto_completo
            texto_limpo_para_ia = re.sub(r"^\s*Súmula de Crédito\s*$", "", texto_limpo_para_ia, flags=re.MULTILINE | re.IGNORECASE)
            texto_limpo_para_ia = re.sub(r"^\s*SICOOB\s*\n(?!Data:)", "", texto_limpo_para_ia, count=1, flags=re.IGNORECASE | re.MULTILINE)
            texto_limpo_para_ia = re.sub(r'Página:\s*\d+\s*\/\s*\d+\s*$', '', texto_limpo_para_ia, flags=re.MULTILINE | re.IGNORECASE)
            texto_limpo_para_ia = texto_limpo_para_ia.strip()
            return texto_limpo_para_ia
    except pdfplumber.exceptions.PDFSyntaxError as e_syntax:
        log_to_gui(f"Erro sintaxe no PDF '{caminho_pdf.name}': {e_syntax}.", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro PDF", f"Erro PDF '{caminho_pdf.name}': {e_syntax}", parent=_root_ref_for_log)
        return None
    except Exception as e:
        log_to_gui(f"Erro extrair texto PDF '{caminho_pdf.name}': {e}", "ERROR")
        logging.error(f"Erro detalhado extrair PDF {caminho_pdf.name}:", exc_info=True)
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Extração PDF", f"Erro ao extrair texto PDF '{caminho_pdf.name}': {e}", parent=_root_ref_for_log)
        return None

def preencher_excel_novo_com_placeholders(
    caminho_json_dados: Path,
    caminho_excel_modelo: Path,
    caminho_excel_saida: Path,
    nome_planilha_alvo: Optional[str] = None
) -> bool:
    log_to_gui(f"Preenchendo Excel: Mod='{caminho_excel_modelo.name}', Dados='{caminho_json_dados.name}', Saída='{caminho_excel_saida.name}'", "INFO")
    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): iniciar_progresso()
    try:
        with open(caminho_json_dados, 'r', encoding='utf-8') as f:
            dados_para_preencher: Dict[str, Any] = json.load(f)
        
        if not dados_para_preencher:
            log_to_gui("AVISO: JSON de dados vazio. Excel de saída será cópia do modelo.", "WARNING")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showwarning("Dados Vazios", "JSON de dados vazio. Excel copiado.", parent=_root_ref_for_log)
            workbook = openpyxl.load_workbook(caminho_excel_modelo)
            workbook.save(caminho_excel_saida)
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso(f"Excel copiado (dados vazios): {caminho_excel_saida.name}")
            return True

        workbook = openpyxl.load_workbook(caminho_excel_modelo)
        sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

        if nome_planilha_alvo:
            if nome_planilha_alvo in workbook.sheetnames:
                sheet = workbook[nome_planilha_alvo]
            else:
                log_to_gui(f"AVISO: Aba '{nome_planilha_alvo}' não encontrada. Usando ativa: '{workbook.active.title}'.", "WARNING")
                if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showwarning("Aba Não Encontrada", f"Aba '{nome_planilha_alvo}' não encontrada. Usando ativa.", parent=_root_ref_for_log)
                sheet = workbook.active
        else:
            sheet = workbook.active
        
        if sheet is None:
            log_to_gui("ERRO CRÍTICO: Nenhuma planilha selecionada no workbook.", "ERROR")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Planilha", "Nenhuma planilha selecionada.", parent=_root_ref_for_log)
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Erro no Excel")
            return False

        log_to_gui(f"Varrendo planilha '{sheet.title}' para placeholders...", "INFO")
        substituicoes_feitas = 0
        placeholder_regex = re.compile(r'\{\{([A-Z0-9_]+?)\}\}')

        for r_idx, row in enumerate(sheet.iter_rows()):
            if _root_ref_for_log and _root_ref_for_log.winfo_exists() and r_idx % 50 == 0:
                status_label.config(text=f"Processando linha {r_idx+1} do Excel...")
                _root_ref_for_log.update_idletasks()

            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_cell_value = str(cell.value)
                    current_cell_string = original_cell_value
                    modified_in_loop = False
                    matches_in_cell = list(placeholder_regex.finditer(original_cell_value))

                    if not matches_in_cell: continue

                    is_single_full_match = len(matches_in_cell) == 1 and matches_in_cell[0].group(0) == original_cell_value
                    
                    if is_single_full_match:
                        placeholder_com_chaves = matches_in_cell[0].group(0)
                        if placeholder_com_chaves in dados_para_preencher:
                            cell.value = dados_para_preencher[placeholder_com_chaves]
                            substituicoes_feitas += 1
                        elif INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES:
                            cell.value = "N/A"
                            substituicoes_feitas += 1
                    else:
                        for match_obj in reversed(matches_in_cell):
                            placeholder_com_chaves = match_obj.group(0)
                            span_inicio, span_fim = match_obj.span(0)
                            
                            if placeholder_com_chaves in dados_para_preencher:
                                valor_substituto = dados_para_preencher[placeholder_com_chaves]
                                valor_substituto_str = str(valor_substituto) if valor_substituto is not None else ""
                                current_cell_string = current_cell_string[:span_inicio] + valor_substituto_str + current_cell_string[span_fim:]
                                modified_in_loop = True; substituicoes_feitas += 1
                            elif INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES:
                                current_cell_string = current_cell_string[:span_inicio] + "N/A" + current_cell_string[span_fim:]
                                modified_in_loop = True; substituicoes_feitas += 1
                        
                        if modified_in_loop: cell.value = current_cell_string
        
        if substituicoes_feitas == 0:
            log_to_gui(f"AVISO: Nenhuma substituição placeholder realizada. Verifique formato e chaves.", "WARNING")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showwarning("Nenhuma Substituição", "Nenhum placeholder substituído.", parent=_root_ref_for_log)
        else:
            log_to_gui(f"INFO: {substituicoes_feitas} substituições realizadas na planilha '{sheet.title}'.", "INFO")

        workbook.save(caminho_excel_saida)
        log_to_gui(f"Novo Excel salvo em '{caminho_excel_saida.name}'.", "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso(f"Excel Gerado: {caminho_excel_saida.name}")
        return True

    except FileNotFoundError as e:
        log_to_gui(f"ERRO Arquivo não encontrado Excel: {e.filename}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Arquivo", f"Arquivo não encontrado: {e.filename}", parent=_root_ref_for_log)
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO JSON Dados Excel '{caminho_json_dados.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro JSON", f"Erro JSON dados Excel: {e}", parent=_root_ref_for_log)
    except Exception as e:
        log_to_gui(f"ERRO geral preencher Excel: {e}", "ERROR")
        logging.error("Erro geral Excel:", exc_info=True)
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Excel", f"Erro preenchimento Excel: {e}", parent=_root_ref_for_log)
    
    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Erro no Excel")
    return False

def salvar_texto_em_arquivo(texto: str, nome_arquivo_path: Path) -> None:
    try:
        nome_arquivo_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arquivo_path, "w", encoding="utf-8") as f: f.write(texto)
        log_to_gui(f"Texto salvo em '{nome_arquivo_path.name}'.", "INFO")
    except Exception as e: log_to_gui(f"Erro salvar texto '{nome_arquivo_path.name}': {e}", "ERROR")

def salvar_json_em_arquivo(dados: Dict[str, Any], nome_arquivo_path: Path) -> None:
    try:
        nome_arquivo_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arquivo_path, "w", encoding="utf-8") as f: json.dump(dados, f, ensure_ascii=False, indent=4)
        log_to_gui(f"JSON salvo em '{nome_arquivo_path.name}'.", "INFO")
    except Exception as e:
        log_to_gui(f"Erro salvar JSON '{nome_arquivo_path.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Salvar JSON", f"Erro salvar JSON '{nome_arquivo_path.name}': {e}", parent=_root_ref_for_log)

@retry(wait=wait_exponential(multiplier=1, min=2, max=30), stop=stop_after_attempt(3), reraise=True)
def gerar_conteudo_gemini_com_retry(model: genai.GenerativeModel, prompt_usuario: str, generation_config: genai.types.GenerationConfig) -> str:
    log_to_gui("Enviando requisição para API Gemini...", "DEBUG")
    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): _root_ref_for_log.update_idletasks()
    
    response = model.generate_content(prompt_usuario, generation_config=generation_config)
    
    log_to_gui("Resposta recebida da API Gemini.", "DEBUG")

    if hasattr(response, 'prompt_feedback') and response.prompt_feedback:
        log_to_gui(f"API Prompt Feedback: Block Reason: {response.prompt_feedback.block_reason}, Safety Ratings: {response.prompt_feedback.safety_ratings}", "DEBUG")

    if not response.parts:
        log_to_gui("AVISO: Resposta da API Gemini não contém 'parts'.", "WARNING")
        if hasattr(response, 'prompt_feedback') and response.prompt_feedback.block_reason:
            block_reason_value = response.prompt_feedback.block_reason
            block_reason_str = block_reason_value.name if hasattr(block_reason_value, 'name') else str(block_reason_value)
            
            block_reason_msg = f"Prompt bloqueado pela API Gemini. Razão: {block_reason_str}."
            safety_ratings_info = ""
            if hasattr(response.prompt_feedback, 'safety_ratings') and response.prompt_feedback.safety_ratings:
                ratings_str_list = []
                for rating in response.prompt_feedback.safety_ratings:
                    category_name = getattr(getattr(rating, 'category', None), 'name', 'N/A')
                    probability_name = getattr(getattr(rating, 'probability', None), 'name', 'N/A')
                    if category_name != 'N/A' and probability_name != 'N/A':
                         ratings_str_list.append(f"  - {category_name}: {probability_name}")
                
                if ratings_str_list:
                    safety_ratings_info = "\n\nDetalhes de Segurança do Prompt:\n" + "\n".join(ratings_str_list)
            
            log_to_gui(f"ERRO API: {block_reason_msg}{safety_ratings_info}", "ERROR")

            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                messagebox.showwarning(
                    "Bloqueio pela API",
                    f"A solicitação para a API Gemini foi bloqueada.\n\n"
                    f"Razão: {block_reason_str}"
                    f"{safety_ratings_info}\n\n"
                    f"O processamento desta parte pode ter falhado. Verifique os logs.",
                    parent=_root_ref_for_log
                )
            raise Exception(block_reason_msg)
        return ""
    return response.text

def enviar_texto_completo_para_gemini_todos_blocos(
    texto_completo_do_pdf: str,
    blocos_config_map: Dict[str, Any],
    pdf_path_para_logs: Path
) -> Optional[Dict[str, Any]]:
    if not texto_completo_do_pdf:
        log_to_gui("Nenhum texto completo para enviar à API.", "WARNING"); return None
    if not genai_config_ok:
        log_to_gui("API Key Google Gemini não configurada. Abortando API.", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro API", "API Key Google Gemini não configurada.", parent=_root_ref_for_log)
        return None

    resposta_texto_bruto_api = ""
    try:
        log_to_gui("Construindo prompt para API Gemini...", "DEBUG")
        model = genai.GenerativeModel(GEMINI_MODEL_NAME)
        prompt_instrucoes_blocos = []

        for nome_bloco, config_bloco in blocos_config_map.items():
            json_chave = config_bloco["json_chave"]
            titulo_bloco_regex_hint = config_bloco.get("titulo_padrao", nome_bloco)
            campos_esperados = config_bloco.get("campos_esperados", [])
            nome_lista_json = config_bloco.get("nome_lista_json")
            sub_campos_lista = config_bloco.get("sub_campos_lista", [])
            sub_lista_aninhada_config = config_bloco.get("sub_lista_aninhada")
            campos_texto_longo_limitar = config_bloco.get("campos_texto_longo_limitar", [])

            instrucao_especifica = (
                f"Para o bloco '{nome_bloco}' (identificado por títulos como '{titulo_bloco_regex_hint}'), "
                f"mapeie para a chave JSON '{json_chave}':\n")

            if nome_bloco == "0 Informacoes do Documento":
                instrucao_especifica += (
                    "  - Para este bloco, extraia os seguintes dados geralmente encontrados no topo do documento:\n"
                    "    - O valor após 'Cooperativa:' para 'cooperativa'.\n"
                    "    - O valor após 'PA:' para 'pa'.\n"
                    "    - O valor após 'Data:' para 'data'.\n"
                    "    - O valor após 'Hora:' para 'hora'.\n"
                    "    - O valor após 'Data/hora referência:' para 'data_ref_doc'.\n"
                    "    - Ignore 'SICOOB' ou 'Súmula de Crédito' como títulos soltos, foque nos valores rotulados.\n"
                )
                campos_ja_tratados_info_doc = ["cooperativa", "pa", "data", "hora", "data_ref_doc"]
                campos_esperados_restantes = [c for c in (campos_esperados or []) if c not in campos_ja_tratados_info_doc]
            else:
                campos_esperados_restantes = campos_esperados or []

            if campos_esperados_restantes:
                instrucao_especifica += f"  - Extraia campos diretos: {', '.join(campos_esperados_restantes)}.\n"
                for campo_limitar in (campos_texto_longo_limitar or []):
                    if campo_limitar in campos_esperados_restantes:
                        instrucao_especifica += (
                            f"    - Para o campo '{campo_limitar}', limite o texto aos primeiros {MAX_TEXT_LENGTH_IA} caracteres OU um resumo conciso DENTRO desse limite. "
                            f"Se mais longo, TRUNQUE. Garanta JSON válido. Se impossível serializar ou excede limite, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'. NÃO exceda {MAX_TEXT_LENGTH_IA} caracteres.\n"
                        )
            
            if nome_lista_json and (sub_campos_lista or []):
                instrucao_especifica += (
                    f"  - Extraia LISTA DE OBJETOS '{nome_lista_json}' com campos: {', '.join(sub_campos_lista)}.\n")
                for campo_limitar in (campos_texto_longo_limitar or []):
                    if campo_limitar in (sub_campos_lista or []):
                         instrucao_especifica += (
                            f"    - Em cada objeto de '{nome_lista_json}', para '{campo_limitar}', limite o texto a {MAX_TEXT_LENGTH_IA} chars OU resumo. "
                            f"Se mais longo, TRUNQUE. Garanta JSON válido. Se impossível ou excede, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'. NÃO exceda {MAX_TEXT_LENGTH_IA} chars.\n"
                        )
                if sub_lista_aninhada_config and isinstance(sub_lista_aninhada_config, dict):
                    nome_sub_lista = sub_lista_aninhada_config.get("nome_json")
                    campos_sub_lista_aninhada = sub_lista_aninhada_config.get("campos", [])
                    if nome_sub_lista and (campos_sub_lista_aninhada or []):
                        instrucao_especifica += (
                            f"    - Dentro de CADA objeto de '{nome_lista_json}', extraia SUB-LISTA '{nome_sub_lista}' com campos: {', '.join(campos_sub_lista_aninhada)}.\n")
                        for campo_limitar_sub in (campos_texto_longo_limitar or []):
                            if campo_limitar_sub in (campos_sub_lista_aninhada or []):
                                instrucao_especifica += (
                                    f"      - Em '{nome_sub_lista}', para '{campo_limitar_sub}', limite a {MAX_TEXT_LENGTH_IA} chars OU resumo. "
                                    f"Se mais longo, TRUNQUE. Garanta JSON válido. Se impossível ou excede, 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'. NÃO exceda {MAX_TEXT_LENGTH_IA} chars.\n"
                                )
                instrucao_especifica += f"  - Se não houver itens para '{nome_lista_json}', retorne lista vazia [].\n"
            prompt_instrucoes_blocos.append(instrucao_especifica)

        prompt_usuario = (
            "Você é um especialista em análise de Súmulas de Crédito. Analise o texto do documento entre "
            f"'{MARCADOR_INICIO_TEXTO_PDF_PROMPT}' e '{MARCADOR_FIM_TEXTO_PDF_PROMPT}' E EXTRAIR INFORMAÇÕES **EXCLUSIVAMENTE PARA OS BLOCOS ESPECIFICADOS ABAIXO**. "
            "Ignore outros blocos não listados. A saída deve ser um ÚNICO objeto JSON.\n\n"
            "REGRAS JSON: Válido, chaves/strings com aspas duplas, escape \\\" e \\n. Se valor não encontrado, omita chave ou use null/[], lista vazia como [].\n"
            "INSTRUÇÕES GERAIS: Use 'json_chave' para cada bloco. Extraia literal mas limpo. Preserve formato de datas, números, códigos. "
            f"TEXTO LONGO (pareceres, etc) em 'campos_texto_longo_limitar': limite a {MAX_TEXT_LENGTH_IA} chars OU resumo conciso. Se impossível serializar ou excede, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'.\n\n"
            f"BLOCOS ESPECÍFICOS PARA EXTRAIR (IGNORE TODOS OS OUTROS):\n"
            + "\n\n".join(prompt_instrucoes_blocos) + "\n\n"
            f"TEXTO COMPLETO DO DOCUMENTO (foque apenas nos blocos listados):\n{MARCADOR_INICIO_TEXTO_PDF_PROMPT}\n"
            f"{texto_completo_do_pdf}\n"
            f"{MARCADOR_FIM_TEXTO_PDF_PROMPT}\n\n"
            "CRÍTICO: RESPOSTA DEVE SER ÚNICO OBJETO JSON. Nível raiz DEVE conter APENAS chaves de primeiro nível: " +
            ", ".join([f'"{blocos_config_map[bn]["json_chave"]}"' for bn in blocos_config_map.keys()]) + ". " +
            "Cada chave conterá dados do bloco. Verifique DUAS VEZES sintaxe JSON e limites de texto."
        )

        log_to_gui(f"Enviando {len(texto_completo_do_pdf)} chars para API (Modelo: {GEMINI_MODEL_NAME})...", "INFO")
        if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
            prompt_debug_path = pdf_path_para_logs.parent / f"{pdf_path_para_logs.stem}_prompt_gemini_part.txt"
            salvar_texto_em_arquivo(prompt_usuario, prompt_debug_path)
        
        generation_config = genai.types.GenerationConfig(
            temperature=GEMINI_TEMPERATURE,
            max_output_tokens=GEMINI_MAX_OUTPUT_TOKENS,
            response_mime_type="application/json"
        )
        
        resposta_texto_bruto_api = gerar_conteudo_gemini_com_retry(model, prompt_usuario, generation_config).strip()
        
        if resposta_texto_bruto_api.startswith("```json"): resposta_texto_bruto_api = resposta_texto_bruto_api[7:]
        if resposta_texto_bruto_api.endswith("```"): resposta_texto_bruto_api = resposta_texto_bruto_api[:-3]
        resposta_texto_bruto_api = resposta_texto_bruto_api.strip()

        if not (resposta_texto_bruto_api.startswith("{") and resposta_texto_bruto_api.endswith("}")):
            log_to_gui(f"AVISO: Resposta API não é JSON completo. Início: {resposta_texto_bruto_api[:200]}...", "WARNING")

        dados_json_combinados: Dict[str, Any] = json.loads(resposta_texto_bruto_api)
        log_to_gui("JSON da API decodificado.", "INFO")
        return dados_json_combinados

    except RetryError as e_retry:
        log_to_gui(f"ERRO FATAL API: Falha conexão Gemini: {e_retry}", "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro API", f"Falha conexão Gemini: {e_retry}", parent=_root_ref_for_log)
    except json.JSONDecodeError as e_json:
        log_to_gui(f"ERRO JSONDecodeError API: {e_json.msg} pos {e_json.pos}", "ERROR")
        if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
            nome_arq_erro = pdf_path_para_logs.parent / f"gemini_resposta_erro_json_{pdf_path_para_logs.stem}.txt"
            salvar_texto_em_arquivo(resposta_texto_bruto_api or "Nenhuma resposta API.", nome_arq_erro)
            log_to_gui(f"Resposta API com erro JSON salva em '{nome_arq_erro.name}'.", "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro API", f"JSON inválido da API: {e_json.msg}", parent=_root_ref_for_log)
    except Exception as e_api_general:
        log_to_gui(f"ERRO GERAL API: {e_api_general}", "ERROR")
        logging.exception("Erro geral API:")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro API", f"Erro API Gemini: {e_api_general}", parent=_root_ref_for_log)
    return None

def achatar_json(objeto_json: Union[Dict[str, Any], List[Any]], prefixo_pai: str = '', separador: str = '_') -> Dict[str, Any]:
    items_achatados: Dict[str, Any] = {}
    if isinstance(objeto_json, dict):
        for chave, valor in objeto_json.items():
            chave_str = str(chave)
            nova_chave_prefixada = f"{prefixo_pai}{separador}{chave_str}" if prefixo_pai else chave_str
            items_achatados.update(achatar_json(valor, nova_chave_prefixada, separador=separador))
    elif isinstance(objeto_json, list):
        if not objeto_json:
            pass
        else:
            if not any(isinstance(item, (dict, list)) for item in objeto_json):
                try:
                    items_achatados[prefixo_pai if prefixo_pai else "lista_raiz_simples"] = ', '.join(map(str, objeto_json))
                except TypeError:
                     items_achatados[prefixo_pai if prefixo_pai else "lista_raiz_simples_erro_str"] = str(objeto_json)
            else:
                for i, item_lista in enumerate(objeto_json):
                    chave_item_lista_indexada = f"{prefixo_pai}{separador}{i+1}"
                    items_achatados.update(achatar_json(item_lista, chave_item_lista_indexada, separador=separador))
    else:
        if prefixo_pai:
            items_achatados[prefixo_pai] = objeto_json
    return items_achatados


def gerenciar_chave_nao_mapeada_interativamente(
    chave_original_nao_mapeada: str,
    caminho_arquivo_config_str: str,
    mapa_atual_em_memoria: Dict[str, str]
) -> Tuple[str, bool, bool, bool]:
    log_to_gui(f"INFO: Chave IA não mapeada: '{chave_original_nao_mapeada}'", "INFO")
    caminho_arquivo_config_abs = resource_path(caminho_arquivo_config_str)
    
    nomes_padronizados_existentes: List[str] = sorted(list(set(mapa_atual_em_memoria.values()))) if mapa_atual_em_memoria else []
    nome_original_formatado = chave_original_nao_mapeada.replace("_", " ").title()

    msg_prompt = (
        f"Chave IA '{chave_original_nao_mapeada}' não mapeada em '{caminho_arquivo_config_abs.name}'.\n\n"
        "Escolha uma opção:\n"
        f"1. Usar Nome Existente (salvo no mapeamento).\n"
        f"2. Criar Novo Nome (salvo no mapeamento).\n"
        f"3. Usar '{nome_original_formatado}' (derivado; salvo; automático para próximas *neste PDF*).\n"
        f"4. PULAR TODAS interações de mapeamento para *este PDF* (nomes originais IA).\n\n"
        "Digite 1-4 (Cancelar/ESC para PULAR SOMENTE ESTA, usando nome original IA)."
    )
    if nomes_padronizados_existentes:
        msg_prompt += "\n\nNomes Existentes (exemplos):\n - " + "\n - ".join(nomes_padronizados_existentes[:min(5, len(nomes_padronizados_existentes))])
        if len(nomes_padronizados_existentes) > 5: msg_prompt += "\n - ..."

    escolha_opcao_str: Optional[str] = simpledialog.askstring("Mapeamento Nova Chave", msg_prompt, parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None)

    nome_para_salvar_no_mapa: Optional[str] = None
    nome_padronizado_escolhido: str = chave_original_nao_mapeada
    mapa_atualizado_arquivo: bool = False
    pular_todas_proximas_pdf: bool = False
    mapear_todas_auto_pdf: bool = False

    if escolha_opcao_str is None:
        log_to_gui(f"PULADO (esta): Mapeamento '{chave_original_nao_mapeada}'. Usando original.", "INFO")
        return nome_padronizado_escolhido, False, False, False

    escolha_opcao_str = escolha_opcao_str.strip()

    if escolha_opcao_str == "1":
        nome_existente_input: Optional[str] = simpledialog.askstring("Nome Existente", "Digite Nome Padronizado Existente:", parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None)
        if nome_existente_input and nome_existente_input.strip():
            nome_para_salvar_no_mapa = nome_existente_input.strip()
            nome_padronizado_escolhido = nome_para_salvar_no_mapa
    elif escolha_opcao_str == "2":
        novo_nome_input: Optional[str] = simpledialog.askstring("Novo Nome", f"Novo nome para '{chave_original_nao_mapeada}':", parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None)
        if novo_nome_input and novo_nome_input.strip():
            nome_para_salvar_no_mapa = novo_nome_input.strip()
            nome_padronizado_escolhido = nome_para_salvar_no_mapa
    elif escolha_opcao_str == "3":
        nome_para_salvar_no_mapa = nome_original_formatado
        nome_padronizado_escolhido = nome_para_salvar_no_mapa
        mapear_todas_auto_pdf = True
        log_to_gui(f"Opção 3: '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}'. Próximas não mapeadas usarão nomes formatados.", "INFO")
    elif escolha_opcao_str == "4":
        pular_todas_proximas_pdf = True
        log_to_gui(f"Opção 4: PULAR TODAS interações mapeamento para *este PDF*.", "INFO")
    else:
        log_to_gui(f"Opção inválida/vazia '{escolha_opcao_str}'. Pulando '{chave_original_nao_mapeada}'.", "INFO")
        return chave_original_nao_mapeada, False, False, False

    if nome_para_salvar_no_mapa and not pular_todas_proximas_pdf:
        config_completo = carregar_mapeamento_de_arquivo(caminho_arquivo_config_str) or {"mapeamento_para_chaves_padronizadas": {}}
        if "mapeamento_para_chaves_padronizadas" not in config_completo or \
           not isinstance(config_completo["mapeamento_para_chaves_padronizadas"], dict):
            config_completo["mapeamento_para_chaves_padronizadas"] = {}
            log_to_gui(f"AVISO: Estrutura 'mapeamento_para_chaves_padronizadas' recriada em '{caminho_arquivo_config_abs.name}'.", "WARNING")

        current_mapping = config_completo["mapeamento_para_chaves_padronizadas"]
        
        if current_mapping.get(chave_original_nao_mapeada) != nome_para_salvar_no_mapa:
            current_mapping[chave_original_nao_mapeada] = nome_para_salvar_no_mapa
            if salvar_mapeamento_em_arquivo(config_completo, caminho_arquivo_config_str):
                mapa_atualizado_arquivo = True
                mapa_atual_em_memoria[chave_original_nao_mapeada] = nome_para_salvar_no_mapa
                log_to_gui(f"Mapeamento salvo: '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}' em '{caminho_arquivo_config_abs.name}'.", "INFO")
            else:
                log_to_gui(f"ERRO: Falha salvar mapeamento '{chave_original_nao_mapeada}'.", "ERROR")
        else:
            mapa_atualizado_arquivo = True
            log_to_gui(f"Mapeamento '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}' já existe.", "DEBUG")

    return nome_padronizado_escolhido, mapa_atualizado_arquivo, pular_todas_proximas_pdf, mapear_todas_auto_pdf


def normalizar_chaves_json(
    json_achatado_da_ia: Dict[str, Any],
    mapeamento_chaves_padronizadas: Optional[Dict[str, str]],
    pular_mapeamento_interativo_global_para_este_pdf: bool
) -> Tuple[Dict[str, Any], Dict[str, str], bool]:
    if not json_achatado_da_ia:
        log_to_gui("Normalizador: Nenhum JSON achatado válido.", "WARNING")
        return {}, dict(mapeamento_chaves_padronizadas or {}), pular_mapeamento_interativo_global_para_este_pdf

    mapeamento_local: Dict[str, str] = dict(mapeamento_chaves_padronizadas or {})
    json_normalizado_placeholders: Dict[str, Any] = {}
    
    _pular_interativo_nesta_exec = pular_mapeamento_interativo_global_para_este_pdf
    _mapear_auto_nesta_exec = False

    mapa_regex: Dict[str, re.Pattern] = {}
    if mapeamento_local:
        sorted_wildcard_keys = sorted(
            [k for k in mapeamento_local.keys() if '*' in k],
            key=lambda k: (k.count('_'), len(k)), reverse=True
        )
        for chave_mapa_wc in sorted_wildcard_keys:
            regex_str = '^' + re.escape(chave_mapa_wc).replace(r'\*', r'([0-9]+)') + '$'
            try: mapa_regex[chave_mapa_wc] = re.compile(regex_str)
            except re.error as e_re: log_to_gui(f"ERRO Regex mapeamento '{chave_mapa_wc}': {e_re}", "ERROR")

    for chave_ia, valor in json_achatado_da_ia.items():
        nome_chave_final: Optional[str] = None

        if _pular_interativo_nesta_exec:
            nome_chave_final = chave_ia
        elif chave_ia in mapeamento_local:
            nome_chave_final = str(mapeamento_local[chave_ia])
        else:
            for chave_mapa_wc_original, regex_compilado in mapa_regex.items():
                match = regex_compilado.match(chave_ia)
                if match:
                    nome_base_mapa_wc = str(mapeamento_local[chave_mapa_wc_original])
                    try:
                        if '*' in nome_base_mapa_wc and match.groups():
                            nome_chave_final = nome_base_mapa_wc.replace("*", match.group(1))
                        else:
                            nome_chave_final = nome_base_mapa_wc
                    except IndexError: nome_chave_final = nome_base_mapa_wc.replace("*", "X_IDX_ERR")
                    break
            
            if not nome_chave_final:
                if _mapear_auto_nesta_exec:
                    nome_chave_final = chave_ia.replace("_", " ").title()
                    mapeamento_local[chave_ia] = nome_chave_final
                    config_atual = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG) or {"mapeamento_para_chaves_padronizadas": {}}
                    if "mapeamento_para_chaves_padronizadas" not in config_atual: config_atual["mapeamento_para_chaves_padronizadas"] = {}
                    config_atual["mapeamento_para_chaves_padronizadas"][chave_ia] = nome_chave_final
                    salvar_mapeamento_em_arquivo(config_atual, ARQUIVO_MAPEAMENTO_CONFIG)
                    log_to_gui(f"MAPEAMENTO AUTO: '{chave_ia}' -> '{nome_chave_final}' (salvo).", "INFO")
                elif not _pular_interativo_nesta_exec:
                    nome_usr, _, _pular_usr, _mapear_auto_usr = \
                        gerenciar_chave_nao_mapeada_interativamente(
                            chave_ia, ARQUIVO_MAPEAMENTO_CONFIG, mapeamento_local)
                    
                    if _pular_usr: _pular_interativo_nesta_exec = True
                    if _mapear_auto_usr: _mapear_auto_nesta_exec = True
                    
                    nome_chave_final = nome_usr
                    if nome_usr != chave_ia and not _pular_interativo_nesta_exec:
                         mapeamento_local[chave_ia] = nome_usr

        if nome_chave_final is None:
            nome_chave_final = chave_ia
            log_to_gui(f"DEBUG: nome_chave_final fallback para '{chave_ia}'.", "DEBUG")

        placeholder_key = str(nome_chave_final).strip().upper()
        placeholder_key = re.sub(r'[^A-Z0-9_]+', '_', placeholder_key)
        while "__" in placeholder_key: placeholder_key = placeholder_key.replace("__", "_")
        placeholder_key = placeholder_key.strip('_')

        if not placeholder_key:
            placeholder_key = f"CHAVE_INVALIDA_{len(json_normalizado_placeholders)}"
            log_to_gui(f"AVISO: Normalização de '{nome_chave_final}' (origem: '{chave_ia}') resultou em placeholder vazio. Usando '{placeholder_key}'.", "WARNING")

        if placeholder_key not in json_normalizado_placeholders:
            json_normalizado_placeholders[placeholder_key] = valor
        else:
            ct = 1; chave_conflito = f"{placeholder_key}_DUPLICADO_{ct}"
            while chave_conflito in json_normalizado_placeholders: ct += 1; chave_conflito = f"{placeholder_key}_DUPLICADO_{ct}"
            json_normalizado_placeholders[chave_conflito] = valor
            log_to_gui(f"AVISO: Conflito placeholder para '{placeholder_key}' (de '{nome_chave_final}'). Salvo como '{chave_conflito}'.", "WARNING")
            
    return json_normalizado_placeholders, mapeamento_local, _pular_interativo_nesta_exec


def gerar_mapeamento_sugestao(json_achatado_ia: Dict[str, Any], nome_arquivo_origem: str = "Desconhecido") -> Dict[str, str]:
    if not json_achatado_ia: log_to_gui("Mapeamento Sugestão: JSON IA vazio.", "WARNING"); return {}

    caminho_map_config_abs = resource_path(ARQUIVO_MAPEAMENTO_CONFIG)
    log_to_gui(f"\n--- SUGESTÃO MAPEAMENTO para '{caminho_map_config_abs.name}' (Base: {nome_arquivo_origem}) ---", "INFO")
    log_to_gui("{\n  \"mapeamento_para_chaves_padronizadas\": {", "INFO")
    
    sugestoes: Dict[str, str] = {}
    
    def nome_sugerido(chave_ia: str) -> str:
        s = str(chave_ia).replace("_", " ").title()
        s = s.replace("Cpf Cnpj", "CPF/CNPJ").replace("Src", "SCR")
        return s.strip()

    list_item_pattern = re.compile(r"^(.*?)_(\d+)(_.*)?$")
    potential_wildcards: Dict[str, List[str]] = {}

    for chave_ia in json_achatado_ia.keys():
        match = list_item_pattern.match(chave_ia)
        if match:
            prefixo, _, sufixo_opcional = match.groups()
            sufixo = sufixo_opcional if sufixo_opcional else ""
            wildcard_pattern_base = f"{prefixo}_*{sufixo}"
            if wildcard_pattern_base not in potential_wildcards:
                potential_wildcards[wildcard_pattern_base] = []
            potential_wildcards[wildcard_pattern_base].append(chave_ia)

    for wc_pattern, ia_keys_list in potential_wildcards.items():
        if len(ia_keys_list) > 1:
            match_exemplo = list_item_pattern.match(ia_keys_list[0])
            if match_exemplo:
                prefixo_ex, _, sufixo_ex_op = match_exemplo.groups()
                sufixo_ex = sufixo_ex_op.lstrip('_') if sufixo_ex_op else ""

                nome_sug_wc = nome_sugerido(prefixo_ex)
                if sufixo_ex: nome_sug_wc += f" - {nome_sugerido(sufixo_ex)}"
                nome_sug_wc += " (Item *)"
                sugestoes[wc_pattern] = nome_sug_wc

    for chave_ia in json_achatado_ia.keys():
        coberta_por_wildcard = False
        for wc_sug in sugestoes.keys():
            if '*' in wc_sug:
                regex_wc_match_str = '^' + re.escape(wc_sug).replace(r'\*', r'[0-9]+') + '$'
                if re.fullmatch(regex_wc_match_str, chave_ia):
                    coberta_por_wildcard = True
                    break
        if not coberta_por_wildcard and chave_ia not in sugestoes:
            sugestoes[chave_ia] = nome_sugerido(chave_ia)
    
    chaves_ordenadas_log = sorted(sugestoes.keys(), key=lambda k: ('*' not in k, k))
    for i, chave_ia_sug in enumerate(chaves_ordenadas_log):
        nome_pad_sug = sugestoes[chave_ia_sug]
        log_to_gui(f'    "{chave_ia_sug}": "{nome_pad_sug}"{"," if i < len(chaves_ordenadas_log) - 1 else ""}', "INFO")
    
    log_to_gui("  }\n}", "INFO")
    log_to_gui(f"--- Copie e cole em '{caminho_map_config_abs.name}'. --- \n", "INFO")
    return sugestoes


def gerar_json_com_chaves_placeholder(json_dados_normalizados: Dict[str, Any], nome_arq_saida_path: Path) -> bool:
    if not json_dados_normalizados: log_to_gui("JSON Placeholders: Dados normalizados vazios.", "WARNING"); return False
        
    json_para_excel: Dict[str, Any] = { f"{{{{{key}}}}}" : val for key, val in json_dados_normalizados.items() }
        
    try:
        nome_arq_saida_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arq_saida_path, "w", encoding="utf-8") as f:
            json.dump(json_para_excel, f, indent=2, ensure_ascii=False)
        log_to_gui(f"JSON para Excel salvo em: '{nome_arq_saida_path.name}'", "INFO"); return True
    except Exception as e:
        log_to_gui(f"Erro salvar JSON para Excel '{nome_arq_saida_path.name}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Salvar JSON Excel", f"Erro salvar JSON para Excel: {e}", parent=_root_ref_for_log)
        return False

# --- FUNÇÃO PRINCIPAL DE PROCESSAMENTO ---
def processar_pdf_e_gerar_saidas(caminho_pdf_path_obj: Path) -> bool:
    log_to_gui(f"--- Iniciando processamento: {caminho_pdf_path_obj.name} ---", "INFO")
    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): iniciar_progresso()

    if not BLOCO_CONFIG:
        log_to_gui("ERRO CRÍTICO: Schema (BLOCO_CONFIG) vazio ou inválido. Interrompido.", "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            parar_progresso("Erro de configuração do schema.")
            messagebox.showerror("Erro Configuração", "Schema vazio ou inválido. Verifique logs.", parent=_root_ref_for_log)
        return False
    
    texto_completo_extraido = extrair_texto_do_pdf(caminho_pdf_path_obj)
    if texto_completo_extraido is None:
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso(f"Erro extrair texto de {caminho_pdf_path_obj.name}")
        return False
    
    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
        nome_arquivo_txt_completo = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_texto_completo_para_ia.txt"
        salvar_texto_em_arquivo(texto_completo_extraido, nome_arquivo_txt_completo)

    resultado_final_combinado_ia: Dict[str, Any] = {}
    sucesso_geral_api = True

    if not LISTA_DE_NOMES_BLOCOS_PARTICIONADA and BLOCO_CONFIG:
        log_to_gui("AVISO: Nenhuma partição válida definida. Processando todos os blocos do schema juntos.", "WARNING")
        lista_particoes_a_usar = [list(BLOCO_CONFIG.keys())]
    elif not LISTA_DE_NOMES_BLOCOS_PARTICIONADA and not BLOCO_CONFIG:
        log_to_gui("INFO: Schema vazio, nenhuma partição para processar.", "INFO")
        lista_particoes_a_usar = []
    else:
        lista_particoes_a_usar = LISTA_DE_NOMES_BLOCOS_PARTICIONADA


    for i, nomes_blocos_nesta_parte in enumerate(lista_particoes_a_usar):
        if not nomes_blocos_nesta_parte:
            log_to_gui(f"Partição {i+1} está vazia. Pulando.", "DEBUG")
            continue

        blocos_config_parcial = {nome_bloco: BLOCO_CONFIG[nome_bloco] for nome_bloco in nomes_blocos_nesta_parte if nome_bloco in BLOCO_CONFIG}
        
        if not blocos_config_parcial:
            log_to_gui(f"Partição {i+1}: Nenhum bloco do schema encontrado para {nomes_blocos_nesta_parte}. Pulando.", "WARNING")
            continue
            
        log_to_gui(f"Enviando Partição {i+1}/{len(lista_particoes_a_usar)} ({len(blocos_config_parcial)} blocos) para API...", "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            status_label.config(text=f"Processando Partição {i+1}/{len(lista_particoes_a_usar)} com IA...")
            _root_ref_for_log.update_idletasks()
            
        resultado_parcial_api = enviar_texto_completo_para_gemini_todos_blocos(
            texto_completo_do_pdf=texto_completo_extraido, blocos_config_map=blocos_config_parcial, pdf_path_para_logs=caminho_pdf_path_obj)
        
        if resultado_parcial_api and isinstance(resultado_parcial_api, dict):
            log_to_gui(f"Partição {i+1} recebida e decodificada.", "INFO")
            if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
                nome_json_parcial = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_parcial_{i+1}.json"
                salvar_json_em_arquivo(resultado_parcial_api, nome_json_parcial)
            
            for key, value in resultado_parcial_api.items():
                if key in resultado_final_combinado_ia:
                    log_to_gui(f"AVISO: Chave JSON '{key}' Partição {i+1} já existe e será sobrescrita.", "WARNING")
                resultado_final_combinado_ia[key] = value
        else:
            log_to_gui(f"ERRO: Falha processar Partição {i+1} com API ou resultado inválido.", "ERROR")
            sucesso_geral_api = False
            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                if not messagebox.askyesno("Erro API", f"Falha API Partição {i+1}. Continuar?", parent=_root_ref_for_log):
                    parar_progresso(f"Erro API Partição {i+1}. Interrompido.")
                    return False
            else: return False
                
    if not sucesso_geral_api and not resultado_final_combinado_ia:
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Erro: Nenhuma parte processada pela API.")
        log_to_gui("ERRO: API falhou em todas as partições ou não retornou dados.", "ERROR")
        return False
    if not resultado_final_combinado_ia and lista_particoes_a_usar:
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Erro: Nenhum dado da API.")
        log_to_gui("ERRO: Resultado combinado da API vazio apesar de partições processadas.", "ERROR")
        return False
    elif not resultado_final_combinado_ia and not lista_particoes_a_usar:
        log_to_gui("INFO: Nenhuma partição foi processada (schema provavelmente vazio ou sem blocos particionáveis).", "INFO")


    log_to_gui("API: Processamento das partes concluído.", "INFO")
    if resultado_final_combinado_ia and CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
        nome_json_bruto = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_bruto_combinado.json"
        salvar_json_em_arquivo(resultado_final_combinado_ia, nome_json_bruto)

    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Achatando JSON IA..."); _root_ref_for_log.update_idletasks()
    json_achatado_da_ia = achatar_json(resultado_final_combinado_ia) if resultado_final_combinado_ia else {}
    
    if not json_achatado_da_ia and resultado_final_combinado_ia:
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso(f"Erro: JSON achatado IA vazio para {caminho_pdf_path_obj.name}")
        log_to_gui("ERRO: Falha achatar JSON API. Resultado vazio.", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Processamento", "JSON IA achatado vazio.", parent=_root_ref_for_log)
        return False
    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS and json_achatado_da_ia :
        nome_json_achatado = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_achatado_debug.json"
        salvar_json_em_arquivo(json_achatado_da_ia, nome_json_achatado)

    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Normalizando chaves..."); _root_ref_for_log.update_idletasks()
    pular_interativo_pdf = False
    map_config_completo = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG)
    
    map_para_chaves_pad: Optional[Dict[str, str]] = None
    if map_config_completo and "mapeamento_para_chaves_padronizadas" in map_config_completo and \
       isinstance(map_config_completo["mapeamento_para_chaves_padronizadas"], dict):
        map_para_chaves_pad = map_config_completo["mapeamento_para_chaves_padronizadas"]
        log_to_gui(f"Usando mapeamento '{resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name}'. Chaves: {len(map_para_chaves_pad)}", "INFO")
    elif json_achatado_da_ia:
        caminho_map_log = resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name
        log_to_gui(f"'{caminho_map_log}' não encontrado/inválido. Gerando sugestão...", "WARNING")
        sugestoes = gerar_mapeamento_sugestao(json_achatado_da_ia, caminho_pdf_path_obj.name)
        map_para_chaves_pad = sugestoes
        
        if salvar_mapeamento_em_arquivo({"mapeamento_para_chaves_padronizadas": sugestoes}, ARQUIVO_MAPEAMENTO_CONFIG):
            log_to_gui(f"Mapeamento com sugestões salvo em '{caminho_map_log}'.", "INFO")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                messagebox.showinfo("Mapeamento Criado", f"'{caminho_map_log}' criado com sugestões. Edite para refinar.", parent=_root_ref_for_log)
        else:
             log_to_gui(f"ERRO: Não foi possível salvar mapeamento sugerido em '{caminho_map_log}'.", "ERROR")
    else:
        log_to_gui("INFO: Nenhum dado da IA para normalizar ou gerar sugestões de mapeamento.", "INFO")


    json_final_norm: Dict[str, Any] = {}
    if json_achatado_da_ia:
        json_final_norm, _, _ = normalizar_chaves_json(
            json_achatado_da_ia, map_para_chaves_pad, pular_interativo_pdf )
        if not json_final_norm and json_achatado_da_ia:
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Erro normalização JSON.")
            log_to_gui("ERRO: Falha normalizar chaves JSON.", "ERROR")
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro Normalização", "Normalização de chaves falhou.", parent=_root_ref_for_log)
            return False
    
    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS and json_final_norm:
        nome_json_norm_debug = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_final_normalizado_debug.json"
        salvar_json_em_arquivo(json_final_norm, nome_json_norm_debug)

    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Gerando JSON para Excel..."); _root_ref_for_log.update_idletasks()
    nome_json_saida_excel = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_dados_para_excel.json"
    
    if gerar_json_com_chaves_placeholder(json_final_norm, nome_json_saida_excel):
        log_to_gui("Selecione ARQUIVO EXCEL MODELO...", "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Aguardando Excel modelo..."); _root_ref_for_log.update_idletasks()

        caminho_template_str: Optional[str] = filedialog.askopenfilename(parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None, title="Selecione EXCEL MODELO", filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos", "*.*")))
        
        if caminho_template_str:
            caminho_template = Path(caminho_template_str)
            nome_saida_sug = f"{caminho_pdf_path_obj.stem}_PREENCHIDO.xlsx"
            caminho_saida_str: Optional[str] = filedialog.asksaveasfilename(parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None, title="Salvar Excel Preenchido Como...", initialdir=str(caminho_pdf_path_obj.parent), initialfile=nome_saida_sug, defaultextension=".xlsx", filetypes=(("Arquivos Excel", "*.xlsx"),))
            
            if caminho_saida_str:
                caminho_saida = Path(caminho_saida_str)
                nome_aba_input: Optional[str] = simpledialog.askstring("Nome da Aba", "Nome da ABA no Excel (deixe em branco para ativa):", parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None)
                nome_plan_alvo = nome_aba_input.strip() if nome_aba_input and nome_aba_input.strip() else None
                
                if not preencher_excel_novo_com_placeholders(nome_json_saida_excel, caminho_template, caminho_saida, nome_plan_alvo):
                    log_to_gui(f"Falha preencher Excel '{caminho_saida.name}'.", "ERROR")
            else:
                if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Salvamento Excel cancelado.")
        else:
            if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso("Seleção modelo Excel cancelada.")
    else:
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): parar_progresso(f"Falha gerar JSON para Excel ({caminho_pdf_path_obj.name})")
        log_to_gui(f"ERRO: Não foi possível gerar JSON para Excel ({nome_json_saida_excel.name}).", "ERROR")

    log_to_gui(f"--- Fim processamento: {caminho_pdf_path_obj.name} ---", "INFO")
    if _root_ref_for_log and _root_ref_for_log.winfo_exists() and not status_label.cget("text").startswith(("Pronto", "Excel Gerado", "Erro")):
        parar_progresso(f"Processamento de {caminho_pdf_path_obj.name} finalizado.")
    return True

# --- FUNÇÕES DA INTERFACE E MAINLOOP ---
def iniciar_fluxo_analise_pdf() -> None:
    log_to_gui("\n--- Novo Ciclo de Análise ---", "INFO")

    if not genai_config_ok:
        log_to_gui("AVISO: API Key Google Gemini não configurada/inválida.", "WARNING")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            messagebox.showerror("Erro API", "API Key Google Gemini não configurada. Verifique .env e logs.", parent=_root_ref_for_log)
            return
    
    if not BLOCO_CONFIG:
        log_to_gui("ERRO CRÍTICO: Schema de extração vazio ou inválido. Processamento não pode continuar.", "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            messagebox.showerror("Erro Configuração", f"Schema '{resource_path(ARQUIVO_SCHEMA_EXTRACAO).name}' vazio ou inválido. Verifique logs.", parent=_root_ref_for_log)
            parar_progresso("Erro configuração schema.")
        return
    
    if BLOCO_CONFIG and not LISTA_DE_NOMES_BLOCOS_PARTICIONADA:
        log_to_gui("AVISO: Nenhuma partição de blocos foi definida no schema. Todos os blocos válidos serão processados juntos.", "WARNING")

    if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Aguardando PDF..."); _root_ref_for_log.update_idletasks()
    caminho_pdf_str: Optional[str] = filedialog.askopenfilename(parent=_root_ref_for_log if _root_ref_for_log and _root_ref_for_log.winfo_exists() else None, title="Selecione PDF de ENTRADA", filetypes=(("Arquivos PDF", "*.pdf"), ("Todos", "*.*")))

    if not caminho_pdf_str:
        log_to_gui("Nenhum PDF selecionado.", "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): status_label.config(text="Nenhum PDF selecionado.")
        return

    caminho_pdf = Path(caminho_pdf_str)
    try:
        sucesso = processar_pdf_e_gerar_saidas(caminho_pdf)
        final_msg = f"Processamento '{caminho_pdf.name}' "
        final_msg += "concluído." if sucesso else "encontrou falhas (ver logs)."
        log_to_gui(final_msg, "INFO")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            messagebox.showinfo("Análise Finalizada", final_msg, parent=_root_ref_for_log)
            status_label.config(text="Pronto.")

    except Exception as e_main:
        log_to_gui(f"ERRO INESPERADO fluxo principal '{caminho_pdf.name}': {e_main}", "CRITICAL")
        logging.exception(f"Exceção não tratada: {caminho_pdf.name}:")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            parar_progresso(f"Erro Crítico Processando {caminho_pdf.name}")
            messagebox.showerror("Erro Crítico", f"Erro inesperado '{caminho_pdf.name}'.\nConsulte '{LOG_FILE_PATH.name}'.\nErro: {e_main}", parent=_root_ref_for_log)
            status_label.config(text="Erro crítico. Pronto.")

def mostrar_sobre() -> None:
    if _root_ref_for_log and _root_ref_for_log.winfo_exists():
        messagebox.showinfo("Sobre",
                            f"Processador de Súmulas v1.0\n"
                            f"Criado por: Claudeir de Souza Alves\n"
                            f"Analista de TI\n"
                            f"Schema: {resource_path(ARQUIVO_SCHEMA_EXTRACAO).name}\n"
                            f"Mapeamento: {resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name}\n"
                            f"Log: {LOG_FILE_PATH.name}\n"
                            f"Modelo Gemini: {GEMINI_MODEL_NAME}",
                            parent=_root_ref_for_log)

def sair_aplicacao() -> None:
    if _root_ref_for_log and _root_ref_for_log.winfo_exists():
        if messagebox.askokcancel("Sair", "Tem certeza que deseja sair?", parent=_root_ref_for_log):
            log_to_gui("Aplicação encerrando...", "INFO")
            _root_ref_for_log.quit()
            _root_ref_for_log.destroy()
    else: sys.exit()

def abrir_arquivo_para_edicao(nome_arq_relativo: str, desc_arq: str) -> None:
    caminho_abs = resource_path(nome_arq_relativo)
    if not caminho_abs.is_file():
        log_to_gui(f"Arquivo {desc_arq} '{caminho_abs.name}' não encontrado.", "WARNING")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showinfo("Arquivo Não Existe", f"Arquivo '{caminho_abs.name}' não existe em '{caminho_abs}'.", parent=_root_ref_for_log)
        return
    try:
        log_to_gui(f"Tentando abrir '{caminho_abs}' para edição...", "INFO")
        if sys.platform == "win32": os.startfile(str(caminho_abs))
        elif sys.platform == "darwin": subprocess.run(["open", str(caminho_abs)], check=True)
        else: subprocess.run(["xdg-open", str(caminho_abs)], check=True)
    except Exception as e:
        log_to_gui(f"ERRO ao tentar abrir '{caminho_abs}': {e}", "ERROR")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists(): messagebox.showerror("Erro ao Abrir", f"Não foi possível abrir '{caminho_abs.name}': {e}", parent=_root_ref_for_log)

def abrir_mapeamento_para_edicao(): abrir_arquivo_para_edicao(ARQUIVO_MAPEAMENTO_CONFIG, "mapeamento")
def abrir_schema_para_edicao(): abrir_arquivo_para_edicao(ARQUIVO_SCHEMA_EXTRACAO, "schema de extração")

# --- 8. CONFIGURAÇÃO FINAL DA GUI (Botões, Menu, mainloop) ---
style.configure("Primary.TButton", font=FONTE_BOTAO_PRINCIPAL, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE), foreground=COR_BOTAO_PRIMARIO_FG, background=COR_BOTAO_PRIMARIO_BG)
style.configure("Secondary.TButton", font=FONTE_BOTAO_SECUNDARIO, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE), foreground=COR_BOTAO_SECUNDARIO_FG, background=COR_BOTAO_SECUNDARIO_BG)
style.map("Primary.TButton", background=[('active', COR_BOTAO_HOVER_PRIMARIO), ('pressed', COR_BOTAO_HOVER_PRIMARIO), ('!disabled', COR_BOTAO_PRIMARIO_BG)], foreground=[('!disabled', COR_BOTAO_PRIMARIO_FG)])
style.map("Secondary.TButton", background=[('active', COR_BOTAO_HOVER_SECUNDARIO), ('pressed', COR_BOTAO_HOVER_SECUNDARIO), ('!disabled', COR_BOTAO_SECUNDARIO_BG)], foreground=[('!disabled', COR_BOTAO_SECUNDARIO_FG)])

frame_botoes_principais.columnconfigure(0, weight=1)
frame_botoes_principais.columnconfigure(1, weight=1)
frame_botoes_principais.columnconfigure(2, weight=1)

botao_analisar = ttk.Button(frame_botoes_principais, text="Analisar PDF e Gerar Saídas", command=iniciar_fluxo_analise_pdf, style="Primary.TButton")
botao_analisar.grid(row=0, column=0, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")

botao_editar_mapeamento = ttk.Button(frame_botoes_principais, text="Editar Mapeamento", command=abrir_mapeamento_para_edicao, style="Secondary.TButton")
botao_editar_mapeamento.grid(row=0, column=1, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")

botao_editar_schema = ttk.Button(frame_botoes_principais, text="Editar Schema IA", command=abrir_schema_para_edicao, style="Secondary.TButton")
botao_editar_schema.grid(row=0, column=2, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")


menu_bar = tk.Menu(root, font=FONTE_MENU)
menu_arquivo = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
icone_menu_pdf_tk: Optional[ImageTk.PhotoImage] = None
caminho_icone_menu_relativo = "assets/icone_pdf.ico"

try:
    caminho_icone_menu_abs = resource_path(caminho_icone_menu_relativo)
    log_to_gui(f"Tentando carregar ícone de menu de: {caminho_icone_menu_abs}", "DEBUG")
    if caminho_icone_menu_abs.is_file():
        img_pil_menu = Image.open(caminho_icone_menu_abs)
        icone_menu_pdf_tk = ImageTk.PhotoImage(img_pil_menu)
        log_to_gui(f"Ícone de menu '{caminho_icone_menu_relativo}' carregado com ImageTk.", "DEBUG")
    else:
        log_to_gui(f"AVISO: Ícone de menu '{caminho_icone_menu_abs}' NÃO ENCONTRADO.", "WARNING")
except Exception as e_icon_menu:
    log_to_gui(f"ERRO ao carregar ícone de menu '{caminho_icone_menu_relativo}': {e_icon_menu}", "ERROR")
    logging.exception(f"Erro detalhado ao carregar ícone de menu '{caminho_icone_menu_relativo}':")


if icone_menu_pdf_tk:
    menu_arquivo.add_command(label="Analisar PDF...", image=icone_menu_pdf_tk, compound="left", command=iniciar_fluxo_analise_pdf)
else:
    menu_arquivo.add_command(label="Analisar PDF...", command=iniciar_fluxo_analise_pdf)

menu_arquivo.add_command(label=f"Editar Mapeamento ({Path(ARQUIVO_MAPEAMENTO_CONFIG).name})", command=abrir_mapeamento_para_edicao)
menu_arquivo.add_command(label=f"Editar Schema ({Path(ARQUIVO_SCHEMA_EXTRACAO).name})", command=abrir_schema_para_edicao)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="Limpar Log da Tela", command=limpar_log_gui)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="Sair", command=sair_aplicacao)
menu_bar.add_cascade(label="Arquivo", menu=menu_arquivo)
menu_ajuda = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
menu_ajuda.add_command(label="Sobre", command=mostrar_sobre)
menu_bar.add_cascade(label="Ajuda", menu=menu_ajuda)
root.config(menu=menu_bar)
root.protocol("WM_DELETE_WINDOW", sair_aplicacao)

if __name__ == "__main__":
    log_to_gui(f"Aplicação iniciada. PID: {os.getpid()}", "INFO")
    log_to_gui(f"Log principal salvo em: {LOG_FILE_PATH.resolve()}", "INFO")
    log_to_gui("Por favor, verifique o arquivo de log para mensagens de erro iniciais que podem não aparecer aqui.", "INFO")

    schema_ok = carregar_schema_extracao()
    particoes_ok = False
    
    if schema_ok:
        if BLOCO_CONFIG:
            particoes_ok = gerar_particoes_dinamicamente()
            if not particoes_ok:
                 log_to_gui("ERRO CRÍTICO: Falha ao gerar partições do schema. Processamento pode falhar.", "CRITICAL")
                 if _root_ref_for_log and _root_ref_for_log.winfo_exists():
                     messagebox.showerror("Erro de Partição", "Falha ao gerar partições do schema. Verifique os logs.", parent=_root_ref_for_log)
        else:
            particoes_ok = True
            log_to_gui("INFO: Schema de extração está vazio. Nenhuma partição será gerada.", "INFO")

    if not schema_ok or (BLOCO_CONFIG and not particoes_ok):
        msg_erro_config = "ERRO FATAL: Falha na configuração inicial (schema/partições). Funcionalidade comprometida."
        log_to_gui(msg_erro_config, "CRITICAL")
        if _root_ref_for_log and _root_ref_for_log.winfo_exists():
            messagebox.showerror("Erro Configuração Inicial", f"{msg_erro_config} Verifique logs.", parent=_root_ref_for_log)

    log_to_gui("Interface Gráfica Pronta.", "INFO")
    if not genai_config_ok and (_root_ref_for_log and _root_ref_for_log.winfo_exists()):
         messagebox.showwarning("Configuração API", "API Key Google Gemini não configurada ou inválida. Extração de PDFs falhará.", parent=_root_ref_for_log)

    root.mainloop()