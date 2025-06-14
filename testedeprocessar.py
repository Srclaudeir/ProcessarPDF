# --- 0. IMPORTAÇÕES DE BIBLIOTECAS ---
import pdfplumber
import google.generativeai as genai
import json
from pathlib import Path
import tkinter as tk
from tkinter import ttk, font as tkFont, filedialog, messagebox, simpledialog, PhotoImage
import re
import os
import sys
import subprocess
import openpyxl
from dotenv import load_dotenv
import time
import logging
from tenacity import retry, stop_after_attempt, wait_exponential, RetryError

# --- CONSTANTES ---
ARQUIVO_MAPEAMENTO_CONFIG = "mapeamento_config.json"
LOG_FILE_NAME = "processamento_pdf.log"
MAX_TEXT_LENGTH_IA = 200
MARCADOR_INICIO_TEXTO_PDF_PROMPT = "[INICIO_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
MARCADOR_FIM_TEXTO_PDF_PROMPT = "[FIM_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES = True # Flag para controlar se "N/A" é inserido

# --- CORES E FONTES PARA A GUI ---
COR_FUNDO_JANELA = "#F0F0F0"
COR_FUNDO_FRAMES_INTERNOS = "#F8F8F8"
COR_TEXTO_PADRAO = "#333333"
COR_TEXTO_TITULO_GUI = "#333333"
COR_TEXTO_LOG = "#1A1A1A"
COR_FUNDO_LOG = "#FFFFFF"
COR_BOTAO_PRIMARIO_BG = "#0078D4"
COR_BOTAO_PRIMARIO_FG = "#000000"
COR_BOTAO_SECUNDARIO_BG = "#FFC107"
COR_BOTAO_SECUNDARIO_FG = "#000000"
COR_STATUS_LABEL_FG = "#005A9E"
FONTE_TITULO_APP = ("Segoe UI", 18, "bold")
FONTE_SUBTITULO = ("Segoe UI", 10)
FONTE_BOTAO_PRINCIPAL = ("Segoe UI", 10, "bold")
FONTE_BOTAO_SECUNDARIO = ("Segoe UI", 10, "bold")
FONTE_BOTAO = ("Segoe UI", 10, "bold")
FONTE_STATUS = ("Segoe UI", 10, "italic")
FONTE_LOG = ("Consolas", 9)
FONTE_MENU = ("Segoe UI", 9)
FONTE_LABELFRAME_TITULO = ("Segoe UI", 10, "bold")
PADDING_X_BOTAO_STYLE = 15
PADDING_Y_BOTAO_STYLE = 5

# --- CONFIGURAÇÃO DO LOGGING ---
logging.basicConfig(
    level=logging.INFO, # Mude para DEBUG para logs mais detalhados
    format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_NAME, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

# --- 1. CARREGAR VARIÁVEIS DE AMBIENTE (.env) ---
load_dotenv()

# --- 2. CONFIGURAÇÃO DA API KEY DO GOOGLE ---
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")

# --- BLOCO DE CONFIGURAÇÃO (COLE SEU BLOCO_CONFIG COMPLETO E REFINADO AQUI) ---
BLOCO_CONFIG = {
    "1 Dados do Associado": {
        "titulo_padrao": r"Dados do Associado",
        "json_chave": "dados_associado",
        "campos_esperados": ["nome_associado", "cpf_cnpj_associado", "c_c_associado" , "endereco_associado", "risco_associado", "pd_associado", "tipo_pessoa_associado", "idade_constituicao_associado", "limite_disponivel_associado", "vigencia_limite_associado", "situacao_limite_associado", "associado_desde_data"]
    },
    "2 Linha de Credito": {
        "titulo_padrao": r"Linha de crédito",
        "json_chave": "linha_credito",
        "campos_esperados": ["linha_credito_nome", "origem_principal", "ind_calculo", "finalidade_principal", "produto_sisbr_legado_principal", "modalidade_sisbr", "produto_sisbr_detalhado", "origem_sisbr_detalhada", "finalidade_sisbr_detalhada"]
    },
    "3 Dados da Proposta": {
        "titulo_padrao": r"Dados da Proposta",
        "json_chave": "dados_proposta",
        "campos_esperados": ["nr_proposta", "tx_juros_proposta", "menor_parc_rs", "financia_seguro_proposta", "data_proposta", "tx_mora_proposta", "maior_parc_rs", "total_despesas_proposta", "data_operacao_proposta", "tx_juros_inad_proposta", "tipo_venc_proposta", "despesas_adic_proposta", "valor_proposta", "tx_multa_proposta", "dia_venc_proposta", "valor_adicional_proposta", "ind_pos_proposta", "tipo_seguro_proposta", "perc_perda_esperada_proposta", "valor_total_financiado_proposta", "perc_ind_pos_proposta", "contr_seguro_proposta", "valor_perda_proposta", "vencimento_proposta", "ind_atraso_proposta", "valor_seguro_proposta", "ativo_problematico_proposta", "cet_mensal_proposta", "perc_ind_atraso_proposta", "iof_adc_proposta", "carteira_proposta", "cet_anual_proposta", "prazo_proposta", "financia_iof_proposta", "estagio_proposta", "pre_autorizado_proposta", "qtd_parcelas_proposta", "tarifa_proposta", "pd_operacao_proposta", "periodicidade_proposta", "primeiro_venc_proposta", "financia_tac_proposta"]
    },
    "4 Reciprocidade Negócios Cooperativa": {
        "titulo_padrao": r"Reciprocidade\s*–\s*Portfólio\s*de\s*Negócios\s*\(\s*Cooperativa\s*\)",
        "json_chave": "reciprocidade_coop",
        "campos_esperados": ["coop_cap_subscrito", "coop_cap_integralizado", "coop_cap_bloqueado", "coop_cap_a_integralizar", "coop_cc_desde", "coop_cc_saldo", "coop_cc_dep_bloqueado", "coop_cc_saldo_bloq_jud", "coop_cc_limite", "coop_cc_saldo_medio_mes_ano_1", "coop_cc_saldo_medio_mes_ano_2", "coop_cc_saldo_medio_mes_ano_3", "coop_cc_trimestre", "coop_cobr_simples", "coop_cobr_caucionada", "coop_cobr_vinculada", "coop_aplic_disp_poupanca", "coop_aplic_disp_dap", "coop_aplic_disp_rdc", "coop_aplic_disp_lca", "coop_aplic_disp_lci", "coop_aplic_bloq_poupanca", "coop_aplic_bloq_dap", "coop_aplic_bloq_rdc", "coop_aplic_bloq_lca", "coop_aplic_bloq_lci"]
    },
    "5 Reciprocidade Negócios Bancoob": {
        "titulo_padrao": r"Reciprocidade\s*–\s*Portfólio\s*de\s*Negócios\s*\(\s*Bancoob\s*\)",
        "json_chave": "reciprocidade_bancoob",
        "campos_esperados": ["bco_cap_subscrito", "bco_cap_integralizado", "bco_cap_bloqueado", "bco_cap_a_integralizar", "bco_cc_desde", "bco_cc_saldo", "bco_cc_dep_bloqueado", "bco_cc_saldo_bloq_jud", "bco_cc_limite", "bco_cc_saldo_medio_mes_ano_1", "bco_cc_saldo_medio_mes_ano_2", "bco_cc_saldo_medio_mes_ano_3", "bco_cc_trimestre", "bco_cobr_simples", "bco_cobr_caucionada", "bco_cobr_vinculada", "bco_aplic_disp_poupanca", "bco_aplic_disp_dap", "bco_aplic_disp_rdc", "bco_aplic_disp_lca", "bco_aplic_disp_lci", "bco_aplic_bloq_poupanca", "bco_aplic_bloq_dap", "bco_aplic_bloq_rdc", "bco_aplic_bloq_lca", "bco_aplic_bloq_lci"]
    },
    "6 Responsabilidade Conta Corrente": {
        "titulo_padrao": r"Responsabilidade\s*–\s*Conta\s*Corrente",
        "json_chave": "responsabilidade_cc",
        "campos_esperados": ["utilizacao_limite_cc", "qtd_dias_cc", "adiantamento_cc", "risco_sacado_cc"]
    },
    "7 Responsabilidade Direta Op Credito Associado Cooperativa": {
        "titulo_padrao": r"Responsabilidade\s*Direta\s*-\s*Operações\s*de\s*Crédito\s*do\s*Associado\s*COOPERATIVA",
        "json_chave": "resp_direta_assoc_coop",
        "nome_lista_json": "operacoes",
        "sub_campos_lista": ["produto", "vencidos", "curto_prazo", "medio_prazo", "longo_prazo", "saldo_devedor"]
    },
    "8 Responsabilidade Direta Op Credito Associado Bancoob": {
        "titulo_padrao": r"Responsabilidade\s*Direta\s*-\s*Operações\s*de\s*Crédito\s*do\s*Associado[\s\S]*?BANCOOB",
        "json_chave": "resp_direta_assoc_bancoob",
        "nome_lista_json": "operacoes",
        "sub_campos_lista": ["produto", "vencidos", "curto_prazo", "medio_prazo", "longo_prazo", "saldo_devedor"]
    },
    "19 Responsabilidade Limites Contratado e Não Utilizado": {
        "titulo_padrao": r"Responsabilidade\s*-\s*Limites\s*Contratados\s*e\s*Não\s*Utilizados",
        "json_chave": "limites_nao_utilizados",
        "campos_esperados": ["titulos_descontados_limite", "credito_rotativo_limite", "cheque_especial_limite","total_limites_nao_utilizados"]
    },
    "20 Central Risco SCR Associado": {
        "titulo_padrao": r"Central\s*de\s*Risco\s*–\s*SCR\s*Associado:",
        "json_chave": "central_risco_src_associado",
        "campos_esperados": ["nome_associado_scr", "data_base_scr", "perc_doc_processados_scr", "qtd_instituicoes_scr", "qtd_operacoes_scr"],
        "nome_lista_json": "modalidades_scr",
        "sub_campos_lista": ["modalidade", "curto_prazo", "medio_prazo", "longo_prazo", "vencidos", "prejuizo", "creditos_a", "total_modalidade"]
    },
    "22 Central Risco SCR Avalistas": {
        "titulo_padrao": r"Central\s*de\s*Risco\s*–\s*SCR[\s\S]*?Avalista:",
        "json_chave": "central_risco_src_avalistas",
        "nome_lista_json": "lista_avalistas_scr",
        "sub_campos_lista": ["nome_avalista_scr", "data_base_avalista_scr", "perc_doc_proc_avalista_scr", "qtd_inst_avalista_scr", "qtd_op_avalista_scr"],
        "sub_lista_aninhada": {
            "nome_json": "modalidades_do_avalista",
            "campos": ["modalidade", "curto_prazo", "medio_prazo", "longo_prazo", "vencidos", "prejuizo", "creditos_a", "total_modalidade"]
        }
    },
    "23 Central Risco SCR Socios": {
        "titulo_padrao": r"Central\s*de\s*Risco\s*–\s*SCR[\s\S]*?Socio:",
        "json_chave": "central_risco_src_socios",
        "nome_lista_json": "lista_socios_scr",
        "sub_campos_lista": ["nome_socio_scr", "data_base_socio_scr", "perc_doc_proc_socio_scr", "qtd_inst_socio_scr", "qtd_op_socio_scr"],
        "sub_lista_aninhada": {
            "nome_json": "modalidades_do_socio",
            "campos": ["modalidade", "curto_prazo", "medio_prazo", "longo_prazo", "vencidos", "prejuizo", "creditos_a", "total_modalidade"]
        }
    },
    "Anotacao Cadastral Informativa": {
        "titulo_padrao": r"Anotação cadastral\s*INFORMATIVA",
        "json_chave": "anotacao_cadastral_informativa",
        "nome_lista_json": "anotacoes_informativas",
        "sub_campos_lista": ["tipo_anotacao", "quantidade", "valor", "origem_info", "dt_anotacao", "dt_ocorrencia"]
    },
    "Anotacao Cadastral Impeditiva Tabela": {
        "titulo_padrao": r"IMPEDITIVA RELATIVA\nTipo\s*Quantid",
        "json_chave": "anotacao_cadastral_impeditiva_tabela",
        "nome_lista_json": "anotacoes_impeditivas_tabela",
        "sub_campos_lista": ["tipo_anotacao_imp", "quantidade_imp", "valor_imp", "origem_info_imp", "dt_anotacao_imp", "dt_ocorrencia_imp"]
    },
    "24 Grupo Econômico": {
        "titulo_padrao": r"Grupo\s*Econômico",
        "json_chave": "grupo_economico",
        "campos_esperados": ["status_participacao_ge"],
        "nome_lista_json": "membros_ge",
        "sub_campos_lista": ["nome_razao_social_membro_ge", "cpf_cnpj_membro_ge", "risco_membro_ge", "valor_exposicao_membro_ge"],
        "sub_lista_aninhada": {
            "nome_json": "responsabilidades_do_membro_ge",
            "campos": ["tipo_responsabilidade_ge", "valor_responsabilidade_ge", "percentual_exposicao_ge"]
        }
    },
    "25 Detalhamento Risco Total": {
        "titulo_padrao": r"Detalhamento\s*Risco\s*Total",
        "json_chave": "detalhamento_risco",
        "campos_esperados": ["risco_total_valor_det", "valor_proposta_det", "responsabilidades_ge_referencia_det", "texto_tabela_responsabilidades_detalhamento"],
        "campos_texto_longo_limitar": ["texto_tabela_responsabilidades_detalhamento"]
    },
    "26 Análises diversas": {
        "titulo_padrao": r"Análises\s*diversas",
        "json_chave": "analises_diversas",
        "campos_esperados": [
            "receita_bruta_anual_valor", "dividas_entidade_ate_360d_valor", "dividas_scr_ate_360d_valor",
            "id_consulta_analise", "patrimonio_total_analise",
            "margem_financeira_1_valor", "margem_financeira_2_valor", "margem_financeira_3_valor",
            "endividamento_patrimonio_total_valor", "endividamento_divida_entidade_valor", "endividamento_dividas_scr_valor",
            "hist_pagamento_texto"
        ],
        "campos_texto_longo_limitar": ["hist_pagamento_texto"]
    },
    "27 Garantias Geral Sumario": {
        "titulo_padrao": r"Garantias\b(?! Pessoal| Real)",
        "json_chave": "garantias_geral_sumario",
        "campos_esperados": ["real_exigido_geral", "real_alcancado_geral", "pessoal_exigido_geral", "pessoal_alcancado_geral"]
    },
    "28 Garantia Pessoal Detalhada": {
        "titulo_padrao": r"Garantia\s*Pessoal",
        "json_chave": "garantia_pessoal_detalhada",
        "nome_lista_json": "garantidores_pessoais",
        "sub_campos_lista": ["cpf_cnpj_gar_pes", "nome_razao_social_gar_pes", "risco_gar_pes", "responsabilidade_gar_pes", "renda_fixa_gar_pes", "renda_variavel_gar_pes", "qtd_op_direta_gar_pes", "valor_op_direta_gar_pes", "qtd_op_indireta_gar_pes", "valor_op_indireta_gar_pes"]
    },
    "29 Garantia Real Detalhada": {
        "titulo_padrao": r"Garantia\s*Real",
        "json_chave": "garantias_real_detalhada",
        "campos_esperados": ["grupo_garantia_real", "enquadramento_sicoob_gar_real", "data_inclusao_gar_real", "valor_garantia_real_total", "ultima_avaliacao_gar_real", "descricao_gar_real"],
        "campos_texto_longo_limitar": ["descricao_gar_real"],
        "nome_lista_json": "titulares_garantia_real",
        "sub_campos_lista": ["tipo_responsabilidade_titular_real", "cpf_cnpj_titular_real", "nome_razao_social_titular_real"]
    },
    "30 Estudo Credito": {
        "titulo_padrao": r"Estudo\nData Início:",
        "json_chave": "estudo_credito",
        "campos_esperados": ["data_inicio_estudo", "data_termino_estudo", "usuario_estudo", "estado_estudo", "parecer_negocial_estudo"],
        "campos_texto_longo_limitar": ["parecer_negocial_estudo"]
    },
    "31 Analise Técnica Credito": {
        "titulo_padrao": r"Análise\s*Técnica\nDt\. Ini:",
        "json_chave": "analise_tecnica_credito",
        "campos_esperados": ["data_inicio_analise_tec", "data_termino_analise_tec", "usuario_analise_tec", "parecer_tecnico_analise"],
        "campos_texto_longo_limitar": ["parecer_tecnico_analise"]
    },
    "Anotacao de Credito Texto Livre": {
        "titulo_padrao": r"Anotação de Crédito\n(?:IMPEDITIVA RELATIVA|ALERTA|INFORMATIVA)",
        "json_chave": "anotacao_credito_texto_livre",
        "nome_lista_json": "lista_anotacoes_texto",
        "sub_campos_lista": ["codigo_tipo_anotacao_txt", "descricao_anotacao_txt", "alcada_anotacao_txt", "situacao_anotacao_txt", "usuario_anotacao_txt", "data_anotacao_txt", "fase_anotacao_txt"],
        "campos_texto_longo_limitar": ["descricao_anotacao_txt"]
    },
    "32 Fluxo Aprovação Parâmetros": {
        "titulo_padrao": r"Fluxo Aprovação\s*Parâmetros de Enquadramento de Alçadas",
        "json_chave": "fluxo_aprovacao_parametros",
        "campos_esperados": ["nome_fluxo_param", "usuario_fluxo_param", "data_hora_fluxo_param", "risco_total_fluxo_param", "media_liberacao_inst_param", "patrimonio_ref_param", "prov_risco_param", "liberacao_perc_param", "comprometimento_perc_param", "anotacao_credito_fluxo_param"],
        "campos_texto_longo_limitar": ["anotacao_credito_fluxo_param"]
    },
    "33 Fluxo Aprovação Alçadas Votacao": {
        "titulo_padrao": r"Resultado\s*Votação\s*Alçadas",
        "json_chave": "fluxo_aprovacao_alcadas_votacao",
        "nome_lista_json": "niveis_aprovacao_votacao",
        "sub_campos_lista": ["nivel_alcada_vot", "usuario_alcada_vot", "data_hora_alcada_vot", "procedimento_alcada_vot", "parecer_alcada_vot"],
        "campos_texto_longo_limitar": ["parecer_alcada_vot"]
    },
    "37 Informações não disponibilizadas": {
        "titulo_padrao": r"Informações não disponibilizadas",
        "json_chave": "info_nao_disponibilizadas",
        "campos_esperados": ["texto_info_nao_disponibilizadas"],
        "campos_texto_longo_limitar": ["texto_info_nao_disponibilizadas"]
    }
}

# --- DEFINIÇÃO DAS PARTIÇÕES DO BLOCO_CONFIG ---
NOMES_BLOCOS_PARTE_1 = [
    "1 Dados do Associado", "2 Linha de Credito", "3 Dados da Proposta",
    "4 Reciprocidade Negócios Cooperativa", "5 Reciprocidade Negócios Bancoob",
    "6 Responsabilidade Conta Corrente"
]
NOMES_BLOCOS_PARTE_2 = [
    "7 Responsabilidade Direta Op Credito Associado Cooperativa",
    "8 Responsabilidade Direta Op Credito Associado Bancoob",
    "19 Responsabilidade Limites Contratado e Não Utilizado"
]
NOMES_BLOCOS_PARTE_3 = [
    "20 Central Risco SCR Associado", "22 Central Risco SCR Avalistas", "23 Central Risco SCR Socios",
    "Anotacao Cadastral Informativa", "Anotacao Cadastral Impeditiva Tabela",
]
NOMES_BLOCOS_PARTE_4 = [
    "24 Grupo Econômico", "25 Detalhamento Risco Total", "26 Análises diversas",
]
NOMES_BLOCOS_PARTE_5 = [
    "27 Garantias Geral Sumario", "28 Garantia Pessoal Detalhada", "29 Garantia Real Detalhada",
    "30 Estudo Credito", "31 Analise Técnica Credito", "Anotacao de Credito Texto Livre",
    "32 Fluxo Aprovação Parâmetros", "33 Fluxo Aprovação Alçadas Votacao",
    "37 Informações não disponibilizadas"
]
LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [
    NOMES_BLOCOS_PARTE_1, NOMES_BLOCOS_PARTE_2, NOMES_BLOCOS_PARTE_3,
    NOMES_BLOCOS_PARTE_4, NOMES_BLOCOS_PARTE_5
]

# --- 3. CONFIGURAÇÃO INICIAL DA INTERFACE GRÁFICA (Widgets Globais) ---
root = tk.Tk()
root.title("Sistema de Processamento PDF com Gemini v2.8 (Final)")
root.geometry("800x700")
root.configure(bg=COR_FUNDO_JANELA)

style = ttk.Style()
style.configure("TButton", padding=6)
style.configure("TLabel", font=FONTE_SUBTITULO, background=COR_FUNDO_JANELA, foreground=COR_TEXTO_PADRAO)
style.configure("TProgressbar", thickness=15, background='#0078D4')
style.configure("Title.TLabel", font=FONTE_TITULO_APP, foreground=COR_TEXTO_PADRAO, background=COR_FUNDO_JANELA)
style.configure("Header.TFrame", background=COR_FUNDO_JANELA)
style.configure("Controls.TFrame", background=COR_FUNDO_JANELA)
style.configure("Status.TFrame", background=COR_FUNDO_JANELA)
style.configure("TLabelframe", background=COR_FUNDO_JANELA, relief=tk.GROOVE, borderwidth=2)
style.configure("TLabelframe.Label", font=FONTE_LABELFRAME_TITULO, background=COR_FUNDO_JANELA, foreground=COR_TEXTO_PADRAO, padding=(5,2))

frame_topo = ttk.Frame(root, padding=(20, 10), style="Header.TFrame")
frame_topo.pack(pady=10, fill=tk.X)
label_titulo_app = ttk.Label(frame_topo, text="Processador de Súmulas de Crédito", style="Title.TLabel")
label_titulo_app.pack(pady=(0, 5))
label_instrucao_app = ttk.Label(frame_topo, text="Selecione um PDF para análise. O log aparecerá abaixo.", justify=tk.CENTER)
label_instrucao_app.pack(pady=(0,10))
frame_botoes_principais = ttk.Frame(frame_topo, style="Controls.TFrame")
frame_botoes_principais.pack(pady=10)
frame_status_progresso = ttk.Frame(root, padding=(20,5), style="Status.TFrame")
frame_status_progresso.pack(fill=tk.X, padx=10)
progress = ttk.Progressbar(frame_status_progresso, orient="horizontal", length=500, mode="indeterminate")
progress.pack(pady=5, fill=tk.X, expand=True)
status_label = ttk.Label(frame_status_progresso, text="Pronto para iniciar.", foreground=COR_STATUS_LABEL_FG, font=FONTE_STATUS, anchor=tk.CENTER)
status_label.pack(pady=5, fill=tk.X)
log_labelframe = ttk.LabelFrame(root, text=" Log de Processamento ", padding=(10,5))
log_labelframe.pack(pady=(5,10), padx=10, fill=tk.BOTH, expand=True)
log_text_widget = tk.Text(log_labelframe, height=15, width=90, wrap=tk.WORD, font=FONTE_LOG, bg=COR_FUNDO_LOG, fg=COR_TEXTO_LOG, relief=tk.SOLID, bd=1, state=tk.DISABLED)
log_scrollbar_y = ttk.Scrollbar(log_labelframe, orient="vertical", command=log_text_widget.yview)
log_text_widget.config(yscrollcommand=log_scrollbar_y.set)
log_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=5, padx=(0,5))
log_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

# --- 4. FUNÇÕES DE FEEDBACK DA INTERFACE GRÁFICA (log, progresso) ---
def log_to_gui(mensagem, level="INFO"):
    formatted_message = f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {level} - {mensagem}"
    log_text_widget.configure(state=tk.NORMAL)
    log_text_widget.insert(tk.END, formatted_message + "\n")
    log_text_widget.see(tk.END)
    log_text_widget.configure(state=tk.DISABLED)
    if root.winfo_exists(): root.update_idletasks()
    if level == "INFO": logging.info(mensagem)
    elif level == "WARNING": logging.warning(mensagem)
    elif level == "ERROR": logging.error(mensagem)
    elif level == "DEBUG": logging.debug(mensagem)
    elif level == "CRITICAL": logging.critical(mensagem)

def iniciar_progresso():
    progress.start(20)
    status_label.config(text="Processando, por favor aguarde...")
    if root.winfo_exists(): root.update_idletasks()

def parar_progresso(final_status=""):
    progress.stop()
    if final_status: status_label.config(text=final_status)
    else: status_label.config(text="Pronto.")
    if root.winfo_exists(): root.update_idletasks()

# --- 5. CONFIGURAÇÃO DA API GOOGLE GEMINI ---
genai_config_ok = False
if not GOOGLE_API_KEY:
    log_to_gui("ERRO CRÍTICO: Variável de ambiente 'GOOGLE_API_KEY' não encontrada.", "CRITICAL")
elif GOOGLE_API_KEY == "SUA_CHAVE_DE_API_AQUI":
    log_to_gui("ERRO CRÍTICO: A GOOGLE_API_KEY no .env ainda é o placeholder. Configure-a.", "CRITICAL")
else:
    try:
        genai.configure(api_key=GOOGLE_API_KEY)
        log_to_gui("API Key do Google Gemini configurada com sucesso.", "INFO")
        genai_config_ok = True
    except Exception as e:
        log_to_gui(f"ERRO ao configurar API Key do Google Gemini: {e}", "ERROR")
        genai_config_ok = False

# --- FUNÇÕES PARA CARREGAR E SALVAR MAPEAMENTO DE CHAVES ---
def carregar_mapeamento_de_arquivo(caminho_arquivo_str: str):
    caminho_arquivo = Path(caminho_arquivo_str)
    if not caminho_arquivo.is_file():
        log_to_gui(f"Arquivo de mapeamento '{caminho_arquivo.name}' não encontrado.", "WARNING")
        return None
    try:
        with open(caminho_arquivo, "r", encoding="utf-8") as f:
            mapeamento = json.load(f)
        log_to_gui(f"Mapeamento de chaves carregado de '{caminho_arquivo.name}'.", "INFO")
        if not isinstance(mapeamento, dict):
            log_to_gui(f"ERRO: Conteúdo de '{caminho_arquivo.name}' não é um dicionário JSON válido.", "ERROR")
            if root.winfo_exists(): messagebox.showerror("Erro de Mapeamento", f"O arquivo '{caminho_arquivo.name}' não contém um mapeamento válido.")
            return None
        return mapeamento
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO ao decodificar JSON do arquivo de mapeamento '{caminho_arquivo.name}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro de Mapeamento", f"Erro ao ler o arquivo '{caminho_arquivo.name}'. Verifique a formatação JSON.\nDetalhes: {e}")
        return None
    except Exception as e:
        log_to_gui(f"ERRO desconhecido ao carregar mapeamento de '{caminho_arquivo.name}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro de Mapeamento", f"Erro ao carregar '{caminho_arquivo.name}'.\nDetalhes: {e}")
        return None

def salvar_mapeamento_em_arquivo(mapeamento: dict, caminho_arquivo_str: str) -> bool:
    caminho_arquivo = Path(caminho_arquivo_str)
    try:
        with open(caminho_arquivo, "w", encoding="utf-8") as f:
            json.dump(mapeamento, f, indent=2, ensure_ascii=False)
        log_to_gui(f"Mapeamento de chaves salvo em '{caminho_arquivo.name}'.", "INFO")
        return True
    except Exception as e:
        log_to_gui(f"ERRO ao salvar mapeamento em '{caminho_arquivo.name}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo de mapeamento '{caminho_arquivo.name}'.\nDetalhes: {e}")
        return False

# --- 6. FUNÇÕES AUXILIARES DE PROCESSAMENTO ---
def extrair_texto_do_pdf(caminho_pdf: Path) -> str | None:
    texto_completo = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            log_to_gui(f"Extraindo texto de '{caminho_pdf.name}'...", "INFO")
            for i, pagina in enumerate(pdf.pages):
                texto_pagina = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=False)
                if not texto_pagina or len(texto_pagina.split()) < 5 :
                    texto_pagina_layout = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=True)
                    if texto_pagina_layout and len(texto_pagina_layout.split()) > len(texto_pagina.split() if texto_pagina else []):
                        texto_pagina = texto_pagina_layout
                if texto_pagina:
                    texto_completo += texto_pagina if texto_pagina.endswith("\n") else texto_pagina + "\n"
            log_to_gui("Extração de texto do PDF concluída.", "INFO")
            texto_limpo_para_ia = re.sub(r'SICOOB\s*Data:\s*\d{2}\/\d{2}\/\d{4}\s*Súmula\s*de\s*Crédito\s*Hora\s*:\s*\d{2}:\d{2}', '', texto_completo, flags=re.IGNORECASE | re.DOTALL)
            texto_limpo_para_ia = re.sub(r'Página:\s*\d+\s*\/\s*\d+', '', texto_limpo_para_ia, flags=re.IGNORECASE | re.DOTALL)
            texto_limpo_para_ia = texto_limpo_para_ia.strip()
            return texto_limpo_para_ia
    except Exception as e:
        log_to_gui(f"Erro ao extrair texto do PDF '{caminho_pdf.name}': {e}", "ERROR")
        logging.error(f"Erro detalhado ao extrair PDF {caminho_pdf.name}:", exc_info=True)
        return None

def preencher_excel_novo_com_placeholders(caminho_json_dados: Path, caminho_excel_modelo: Path, caminho_excel_saida: Path, nome_planilha_alvo: str | None = None) -> bool:
    log_to_gui(f"Iniciando preenchimento do Excel: Modelo='{caminho_excel_modelo.name}', Dados='{caminho_json_dados.name}', Saída='{caminho_excel_saida.name}'", "INFO")
    iniciar_progresso()
    try:
        log_to_gui(f"Carregando dados JSON de: {caminho_json_dados}", "DEBUG")
        with open(caminho_json_dados, 'r', encoding='utf-8') as f:
            dados_para_preencher = json.load(f)
        log_to_gui(f"Dados JSON carregados. Total de chaves: {len(dados_para_preencher)}", "DEBUG")
        if dados_para_preencher:
            log_to_gui(f"DEBUG: Primeiras 5 chaves do JSON de dados: {list(dados_para_preencher.keys())[:5]}", "DEBUG")
        else:
            log_to_gui("DEBUG: JSON de dados está vazio.", "DEBUG")

        if not dados_para_preencher:
            log_to_gui("AVISO: JSON de dados para preencher está vazio. Nenhum dado será inserido no Excel.", "WARNING")
            if root.winfo_exists(): messagebox.showwarning("Dados Vazios", "O arquivo JSON de dados está vazio. O Excel de saída será uma cópia do modelo.")
            workbook = openpyxl.load_workbook(caminho_excel_modelo)
            workbook.save(caminho_excel_saida)
            parar_progresso(f"Excel copiado (vazio): {caminho_excel_saida.name}")
            return True

        log_to_gui(f"Carregando workbook Excel de: {caminho_excel_modelo}", "DEBUG")
        workbook = openpyxl.load_workbook(caminho_excel_modelo)
        sheet = None
        if nome_planilha_alvo:
            log_to_gui(f"Tentando selecionar aba: '{nome_planilha_alvo}'", "DEBUG")
            if nome_planilha_alvo in workbook.sheetnames:
                sheet = workbook[nome_planilha_alvo]
                log_to_gui(f"Aba '{nome_planilha_alvo}' selecionada.", "INFO")
            else:
                log_to_gui(f"AVISO: Aba '{nome_planilha_alvo}' não encontrada. Usando aba ativa: '{workbook.active.title}'.", "WARNING")
                sheet = workbook.active
                log_to_gui(f"DEBUG: Usando planilha ativa: '{sheet.title}'", "DEBUG")
        else: 
            sheet = workbook.active
            log_to_gui(f"Usando aba ativa: '{sheet.title}'", "INFO")
            log_to_gui(f"DEBUG: Usando planilha ativa por padrão: '{sheet.title}'", "DEBUG")
        
        if sheet is None:
            log_to_gui("ERRO CRÍTICO: Nenhuma planilha pôde ser selecionada no workbook.", "ERROR")
            if root.winfo_exists(): messagebox.showerror("Erro de Planilha", "Nenhuma planilha selecionada ou ativa no Excel modelo.")
            parar_progresso("Erro no Excel")
            return False

        log_to_gui(f"Iniciando varredura da planilha '{sheet.title}' para placeholders...", "INFO")
        substituicoes_feitas = 0

        for r_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if isinstance(cell.value, str):
                    valor_celula_log = cell.value if len(str(cell.value)) < 100 else str(cell.value)[:100] + "..."
                    # Descomente a linha abaixo para um log MUITO verboso de cada célula string
                    # log_to_gui(f"DEBUG: Célula {cell.coordinate} (Linha {r_idx+1}) string: '{valor_celula_log}'", "DEBUG")

                    original_cell_value_for_comparison = cell.value
                    current_cell_value_as_string = cell.value
                    
                    matches_internos = re.findall(r'\{\{([A-Z0-9_ .\/\-%()]+?)\}\}', cell.value)

                    if matches_internos:
                        # log_to_gui(f"DEBUG: Nomes internos de placeholders encontrados em {cell.coordinate}: {matches_internos}", "DEBUG")
                        
                        placeholder_unico_na_celula = False
                        if len(matches_internos) == 1:
                            placeholder_construido_para_teste = f"{{{{{matches_internos[0]}}}}}"
                            if cell.value == placeholder_construido_para_teste:
                                placeholder_unico_na_celula = True
                                # log_to_gui(f"DEBUG: Célula {cell.coordinate} contém placeholder único: '{placeholder_construido_para_teste}'", "DEBUG")

                        for nome_interno_placeholder in matches_internos:
                            placeholder_completo_com_chaves = f"{{{{{nome_interno_placeholder}}}}}"
                            # log_to_gui(f"DEBUG: -- Verificando placeholder JSON: '{placeholder_completo_com_chaves}' para célula {cell.coordinate}", "DEBUG")

                            if placeholder_completo_com_chaves in dados_para_preencher:
                                valor_substituto = dados_para_preencher[placeholder_completo_com_chaves]
                                # valor_substituto_log = str(valor_substituto) if len(str(valor_substituto)) < 50 else str(valor_substituto)[:50] + "..."
                                # log_to_gui(f"DEBUG: ---- Placeholder '{placeholder_completo_com_chaves}' ENCONTRADO no JSON. Valor: '{valor_substituto_log}', Tipo: {type(valor_substituto)}", "DEBUG")
                                
                                if placeholder_unico_na_celula:
                                    # log_to_gui(f"DEBUG: ------ Substituindo célula INTEIRA {cell.coordinate} com valor de tipo {type(valor_substituto)}.", "DEBUG")
                                    cell.value = valor_substituto 
                                    substituicoes_feitas +=1
                                    break 
                                else:
                                    valor_substituto_str = str(valor_substituto) if valor_substituto is not None else ""
                                    # log_to_gui(f"DEBUG: ------ Substituindo PARTE da string em {cell.coordinate}: '{placeholder_completo_com_chaves}' por '{valor_substituto_str[:50]}'", "DEBUG")
                                    current_cell_value_as_string = current_cell_value_as_string.replace(placeholder_completo_com_chaves, valor_substituto_str)
                                    substituicoes_feitas +=1
                            else:
                                log_to_gui(f"AVISO: Placeholder '{placeholder_completo_com_chaves}' (de '{nome_interno_placeholder}') NÃO encontrado no JSON para célula {cell.coordinate}.", "WARNING")
                                if INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES:
                                    if placeholder_unico_na_celula:
                                        cell.value = "N/A"
                                        substituicoes_feitas +=1
                                        break
                                    else:
                                        current_cell_value_as_string = current_cell_value_as_string.replace(placeholder_completo_com_chaves, "N/A")
                                        substituicoes_feitas +=1
                        
                        if not placeholder_unico_na_celula and cell.value == original_cell_value_for_comparison and current_cell_value_as_string != original_cell_value_for_comparison:
                            # log_to_gui(f"DEBUG: Atualizando célula {cell.coordinate} com string mista final: '{current_cell_value_as_string[:100]}'", "DEBUG")
                            cell.value = current_cell_value_as_string
                        elif placeholder_unico_na_celula and cell.value == placeholder_construido_para_teste and not (placeholder_completo_com_chaves in dados_para_preencher):
                             if not INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES: # Se não foi substituído por N/A e não foi encontrado
                                log_to_gui(f"DEBUG: Placeholder único '{cell.value}' em {cell.coordinate} não encontrado no JSON e não substituído por N/A, mantendo original.", "DEBUG")
                            
        if substituicoes_feitas > 0:
            log_to_gui(f"INFO: {substituicoes_feitas} substituições de placeholders realizadas/tentadas.", "INFO")
        else:
            log_to_gui("AVISO: Nenhuma substituição de placeholder foi realizada. Verifique os placeholders no Excel e as chaves no JSON de dados.", "WARNING")

        workbook.save(caminho_excel_saida)
        log_to_gui(f"Novo arquivo Excel salvo em '{caminho_excel_saida.name}'.", "INFO")
        parar_progresso(f"Excel Gerado: {caminho_excel_saida.name}")
        return True
    except FileNotFoundError as e:
        log_to_gui(f"ERRO CRÍTICO: Arquivo não encontrado ao preencher Excel: {e}", "ERROR")
        logging.error("Detalhes do FileNotFoundError no Excel:", exc_info=True)
        if root.winfo_exists(): messagebox.showerror("Erro de Arquivo", f"Arquivo não encontrado: {e.filename}")
        parar_progresso("Erro no Excel")
        return False
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO CRÍTICO ao ler JSON de dados '{caminho_json_dados.name}': {e}", "ERROR")
        logging.error("Detalhes do JSONDecodeError no Excel:", exc_info=True)
        if root.winfo_exists(): messagebox.showerror("Erro de JSON", f"Erro ao ler o arquivo JSON.\nDetalhes: {e}")
        parar_progresso("Erro no JSON")
        return False
    except Exception as e:
        log_to_gui(f"ERRO CRÍTICO geral ao preencher o Excel: {e}", "ERROR")
        logging.error("Detalhes do erro geral no Excel:", exc_info=True)
        if root.winfo_exists(): messagebox.showerror("Erro de Preenchimento Excel", f"Ocorreu um erro inesperado.\nDetalhes: {e}")
        parar_progresso("Erro no Excel")
        return False

def salvar_texto_em_arquivo(texto: str, nome_arquivo_path: Path):
    try:
        with open(nome_arquivo_path, "w", encoding="utf-8") as f: f.write(texto)
        log_to_gui(f"Texto salvo em '{nome_arquivo_path.name}'.", "INFO")
    except IOError as e: log_to_gui(f"Erro de I/O ao salvar texto '{nome_arquivo_path.name}': {e}", "ERROR")
    except Exception as e: log_to_gui(f"Erro desconhecido ao salvar texto '{nome_arquivo_path.name}': {e}", "ERROR")

def salvar_json_em_arquivo(dados: dict, nome_arquivo_path: Path):
    try:
        with open(nome_arquivo_path, "w", encoding="utf-8") as f: json.dump(dados, f, ensure_ascii=False, indent=4)
        log_to_gui(f"JSON salvo com sucesso em '{nome_arquivo_path.name}'.", "INFO")
    except TypeError as e: log_to_gui(f"Erro de tipo ao serializar JSON para '{nome_arquivo_path.name}': {e}.", "ERROR")
    except IOError as e: log_to_gui(f"Erro de I/O ao salvar JSON '{nome_arquivo_path.name}': {e}", "ERROR")
    except Exception as e:
        log_to_gui(f"Erro desconhecido ao salvar JSON '{nome_arquivo_path.name}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro ao Salvar JSON", f"Erro ao salvar JSON.\nDetalhes: {e}")

# --- FUNÇÃO DE CHAMADA À API GEMINI COM RETENTATIVAS ---
@retry(wait=wait_exponential(multiplier=1, min=2, max=30), stop=stop_after_attempt(3), reraise=True)
def gerar_conteudo_gemini_com_retry(model, prompt_usuario, generation_config):
    log_to_gui("Enviando requisição para API Gemini...", "DEBUG")
    response = model.generate_content(prompt_usuario, generation_config=generation_config)
    log_to_gui("Resposta recebida da API Gemini.", "DEBUG")
    return response.text

# --- FUNÇÃO DE ENVIO PARA GEMINI (COM PROMPT REFINADO) ---
def enviar_texto_completo_para_gemini_todos_blocos(texto_completo_do_pdf: str, blocos_config_map: dict, pdf_path_para_logs: Path) -> dict | None:
    if not texto_completo_do_pdf:
        log_to_gui("Nenhum texto completo para enviar à API.", "WARNING")
        return None
    if not genai_config_ok:
        log_to_gui("API Key do Google Gemini não configurada.", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro de API", "API Key do Google Gemini não está configurada.")
        return None
    resposta_texto_bruto_api = ""
    try:
        log_to_gui("Construindo prompt para API Gemini...", "DEBUG")
        model = genai.GenerativeModel('gemini-1.5-flash-latest')
        prompt_instrucoes_blocos = []
        for nome_bloco, config_bloco in blocos_config_map.items():
            json_chave = config_bloco["json_chave"]
            titulo_bloco_regex_hint = config_bloco.get("titulo_padrao", nome_bloco)
            campos_esperados = config_bloco.get("campos_esperados", [])
            nome_lista_json = config_bloco.get("nome_lista_json")
            sub_campos_lista = config_bloco.get("sub_campos_lista", [])
            sub_lista_aninhada_config = config_bloco.get("sub_lista_aninhada")
            campos_texto_longo_limitar = config_bloco.get("campos_texto_longo_limitar", [])
            instrucao_especifica = (
                f"Para o bloco '{nome_bloco}' (identificado por títulos como '{titulo_bloco_regex_hint}'), "
                f"mapeie para a chave JSON '{json_chave}':\n")
            if campos_esperados:
                instrucao_especifica += f"  - Extraia campos diretos: {', '.join(campos_esperados)}.\n"
                for campo_limitar in campos_texto_longo_limitar:
                    if campo_limitar in campos_esperados:
                        instrucao_especifica += (
                            f"    - Para '{campo_limitar}', se extenso, resuma ou use os primeiros {MAX_TEXT_LENGTH_IA} chars. "
                            f"CRUCIAL: String JSON válida (escape \\\" e \\n). Se complexo, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'.\n"
                        )
            if nome_lista_json and sub_campos_lista:
                instrucao_especifica += (
                    f"  - Extraia LISTA DE OBJETOS '{nome_lista_json}' com campos: {', '.join(sub_campos_lista)}.\n")
                for campo_limitar in campos_texto_longo_limitar:
                    if campo_limitar in sub_campos_lista:
                         instrucao_especifica += (
                            f"    - Em cada objeto de '{nome_lista_json}', para '{campo_limitar}', se extenso, resuma ou use os primeiros {MAX_TEXT_LENGTH_IA} chars. "
                            f"CRUCIAL: String JSON válida. Se complexo, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'.\n"
                        )
                if sub_lista_aninhada_config:
                    nome_sub_lista = sub_lista_aninhada_config.get("nome_json")
                    campos_sub_lista = sub_lista_aninhada_config.get("campos", [])
                    if nome_sub_lista and campos_sub_lista:
                        instrucao_especifica += (
                            f"    - Dentro de CADA objeto de '{nome_lista_json}', extraia SUB-LISTA '{nome_sub_lista}' com campos: {', '.join(campos_sub_lista)}.\n")
                        for campo_limitar_sub in campos_texto_longo_limitar:
                            if campo_limitar_sub in campos_sub_lista:
                                instrucao_especifica += (
                                    f"      - Em '{nome_sub_lista}', para '{campo_limitar_sub}', se extenso, resuma ou use os primeiros {MAX_TEXT_LENGTH_IA} chars. "
                                    f"CRUCIAL: String JSON válida. Se complexo, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'.\n"
                                )
                instrucao_especifica += f"  - Se não houver itens para '{nome_lista_json}', retorne [].\n"
            prompt_instrucoes_blocos.append(instrucao_especifica)

        prompt_usuario = (
            "Você é um especialista em análise de Súmulas de Crédito. Sua tarefa é analisar o texto do documento fornecido entre os marcadores "
            f"'{MARCADOR_INICIO_TEXTO_PDF_PROMPT}' e '{MARCADOR_FIM_TEXTO_PDF_PROMPT}' e extrair informações estruturadas em um ÚNICO objeto JSON.\n\n"
            "REGRAS CRÍTICAS PARA A SAÍDA JSON:\n"
            "1.  A SAÍDA DEVE SER UM ÚNICO OBJETO JSON VÁLIDO E COMPLETO. Sem nenhum texto ou formatação adicional antes ou depois do objeto JSON.\n"
            "2.  Todas as chaves JSON e todos os valores do tipo string DEVEM estar entre aspas duplas (ex: \"chave\": \"valor\").\n"
            "3.  VALORES STRING: Aspas duplas LITERAIS dentro de um valor string DEVEM ser escapadas com uma barra invertida (ex: \"descrição com \\\"aspas\\\" internas\").\n"
            "4.  VALORES STRING: Quebras de linha LITERAIS dentro de um valor string DEVEM ser representadas como '\\n'. NÃO inclua quebras de linha reais dentro de uma string JSON.\n"
            "5.  Caracteres especiais como barras invertidas dentro de strings também devem ser escapados (ex: \"caminho\\\\para\\\\arquivo\").\n\n"
            "INSTRUÇÕES GERAIS DE EXTRAÇÃO:\n"
            "- Use a 'json_chave' especificada para cada bloco como chave de nível superior.\n"
            "- Se um campo/lista não for encontrado no texto do documento fornecido, omita-o do JSON, ou use `null` para campos, ou uma lista vazia `[]` para listas.\n"
            "- EXTRAIA DADOS ESTRITAMENTE DO DOCUMENTO FORNECIDO. NÃO inclua nomes de arquivos, marcadores do prompt, ou qualquer informação externa ao documento.\n"
            "- VALORES NUMÉRICOS: Extraia como números (int ou float), sem 'R$', '.' como decimal. Datas: DD/MM/AAAA. Percentuais: números (ex: 10.5 para 10,5%).\n"
            f"- CAMPOS DE TEXTO LONGO (pareceres, descrições): Quando instruído, limite a extração a um resumo conciso ou aos primeiros {MAX_TEXT_LENGTH_IA} caracteres. "
            "   Priorize a extração correta e a validade da string JSON. Se um texto longo for muito complexo para serializar corretamente dentro do limite (devido a muitos caracteres especiais que precisam de escape), "
            "   como ÚLTIMO RECURSO para esse campo específico, retorne a string literal 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL' em vez do conteúdo truncado/resumido, para evitar quebrar o JSON.\n"
            "- 'NADA CONSTA' no PDF deve ser extraído como a string \"NADA CONSTA\". Campos visivelmente vazios no PDF como `null`.\n\n"
            "CONFIGURAÇÃO DOS BLOCOS A SEREM EXTRAÍDOS:\n" + "\n\n".join(prompt_instrucoes_blocos) + "\n\n"
            f"TEXTO COMPLETO DO DOCUMENTO PARA ANÁLISE:\n{MARCADOR_INICIO_TEXTO_PDF_PROMPT}\n"
            f"{texto_completo_do_pdf}\n"
            f"{MARCADOR_FIM_TEXTO_PDF_PROMPT}\n\n"
            "Retorne APENAS o objeto JSON resultante. Verifique DUAS VEZES a sintaxe do JSON, especialmente a formatação de todas as strings, antes de finalizar a resposta."
        )
        log_to_gui("Enviando documento para API Gemini...", "INFO")
        prompt_debug_path = pdf_path_para_logs.parent / f"{pdf_path_para_logs.stem}_prompt_gemini.txt"
        salvar_texto_em_arquivo(prompt_usuario, prompt_debug_path)
        generation_config = genai.types.GenerationConfig(temperature=0.1, max_output_tokens=8190)
        resposta_texto_bruto_api = gerar_conteudo_gemini_com_retry(model, prompt_usuario, generation_config).strip()
        if resposta_texto_bruto_api.startswith("```json"): resposta_texto_bruto_api = resposta_texto_bruto_api[7:]
        if resposta_texto_bruto_api.endswith("```"): resposta_texto_bruto_api = resposta_texto_bruto_api[:-3]
        resposta_texto_bruto_api = resposta_texto_bruto_api.strip()
        log_to_gui("Resposta recebida. Decodificando JSON...", "INFO")
        dados_json_combinados = json.loads(resposta_texto_bruto_api)
        log_to_gui("JSON da API decodificado com sucesso.", "INFO")
        return dados_json_combinados
    except RetryError as e:
        log_to_gui(f"ERRO FATAL API: {e}", "CRITICAL")
        if root.winfo_exists(): messagebox.showerror("Erro de API", f"Falha ao conectar à API Gemini.\nDetalhes: {e}")
        return None
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO JSONDecodeError: {e}", "ERROR")
        log_to_gui(f"Contexto do erro no JSON: '{resposta_texto_bruto_api[max(0, e.pos-40):e.pos+40]}'", "DEBUG")
        nome_arquivo_erro = pdf_path_para_logs.parent / f"gemini_resposta_erro_json_{pdf_path_para_logs.stem}.txt"
        salvar_texto_em_arquivo(resposta_texto_bruto_api if resposta_texto_bruto_api else "Nenhuma resposta de texto.", nome_arquivo_erro)
        log_to_gui(f"Resposta da API com erro salva em '{nome_arquivo_erro.name}'.", "INFO")
        if root.winfo_exists(): messagebox.showerror("Erro de API", f"JSON inválido da API.\nVerifique '{nome_arquivo_erro.name}'.\nDetalhes: {e}")
        return None
    except Exception as e:
        log_to_gui(f"ERRO GERAL API/Processamento: {e}", "ERROR")
        logging.error("Detalhes do erro geral API:", exc_info=True)
        if root.winfo_exists(): messagebox.showerror("Erro de API", f"Erro com API Gemini.\nDetalhes: {e}")
        return None

# --- FUNÇÕES DE ACHATAMENTO, NORMALIZAÇÃO, GERAÇÃO DE MAPEAMENTO (ADAPTADAS) ---
def achatar_json(objeto_json, prefixo_pai='', separador='_'):
    items_achatados = {}
    if isinstance(objeto_json, dict):
        for chave, valor in objeto_json.items():
            nova_chave_prefixada = f"{prefixo_pai}{separador}{chave}" if prefixo_pai else chave
            items_achatados.update(achatar_json(valor, nova_chave_prefixada, separador=separador))
    elif isinstance(objeto_json, list):
        if not objeto_json:
            items_achatados[prefixo_pai if prefixo_pai else "lista_raiz_vazia"] = "[]"
        else:
            if any(isinstance(item, (dict, list)) for item in objeto_json):
                for i, item_lista in enumerate(objeto_json):
                    chave_item_lista_indexada = f"{prefixo_pai}{separador}{i}"
                    items_achatados.update(achatar_json(item_lista, chave_item_lista_indexada, separador=separador))
            else:
                items_achatados[prefixo_pai if prefixo_pai else "lista_raiz_simples"] = ', '.join(map(str, objeto_json))
    else:
        if prefixo_pai: items_achatados[prefixo_pai] = objeto_json
    return items_achatados

def gerenciar_chave_nao_mapeada_interativamente(chave_original_nao_mapeada, caminho_arquivo_config_str, mapa_atual_em_memoria):
    log_to_gui(f"INFO: Chave nova não mapeada: '{chave_original_nao_mapeada}'", "INFO")
    caminho_arquivo_config = Path(caminho_arquivo_config_str)
    sinal_para_pular_todas_as_proximas = False
    sinal_mapear_todas_automaticamente_nesta_chave = False
    nomes_padronizados_existentes = []
    if mapa_atual_em_memoria and isinstance(mapa_atual_em_memoria, dict):
        nomes_padronizados_existentes = sorted(list(set(mapa_atual_em_memoria.values())))
    msg_prompt = (
        f"Chave '{chave_original_nao_mapeada}' não mapeada em '{caminho_arquivo_config.name}'.\n\n"
        "1. Usar Existente (salva).\n2. Criar Novo (salva).\n"
        "3. Usar Original '{chave_original_nao_mapeada.replace('_', ' ').title()}' (salva ESTA e AUTOMATICAMENTE as PRÓXIMAS não mapeadas deste PDF com seus nomes originais formatados).\n"
        "4. PULAR TODAS para este PDF.\n\nDigite 1-4 (Cancelar/ESC para PULAR SÓ ESTA)."
    )
    if nomes_padronizados_existentes:
        msg_prompt += "\n\nExistentes (exemplos):\n - " + "\n - ".join(nomes_padronizados_existentes[:5])
        if len(nomes_padronizados_existentes) > 5: msg_prompt += "\n - ..."
    escolha_opcao_str = simpledialog.askstring("Mapeamento de Nova Chave", msg_prompt, parent=root)
    if escolha_opcao_str is None:
        log_to_gui(f"PULADO (só esta): Mapeamento para '{chave_original_nao_mapeada}'.", "INFO")
        return chave_original_nao_mapeada, False, False, False
    escolha_opcao_str = escolha_opcao_str.strip()
    nome_para_salvar_no_mapa = None
    nome_padronizado_escolhido_para_uso_atual = chave_original_nao_mapeada
    mapa_foi_atualizado_no_arquivo = False
    if escolha_opcao_str == "1":
        nome_existente_input = simpledialog.askstring("Nome Existente", "Digite Nome Padronizado Existente:", parent=root)
        if nome_existente_input and nome_existente_input.strip():
            nome_para_salvar_no_mapa = nome_existente_input.strip()
            nome_padronizado_escolhido_para_uso_atual = nome_para_salvar_no_mapa
        else: return chave_original_nao_mapeada, False, False, False
    elif escolha_opcao_str == "2":
        novo_nome_input = simpledialog.askstring("Novo Nome Padronizado", f"Novo para '{chave_original_nao_mapeada}':", parent=root)
        if novo_nome_input and novo_nome_input.strip():
            nome_para_salvar_no_mapa = novo_nome_input.strip()
            nome_padronizado_escolhido_para_uso_atual = nome_para_salvar_no_mapa
        else: return chave_original_nao_mapeada, False, False, False
    elif not escolha_opcao_str or escolha_opcao_str == "3":
        nome_para_salvar_no_mapa = chave_original_nao_mapeada.replace("_", " ").title()
        nome_padronizado_escolhido_para_uso_atual = nome_para_salvar_no_mapa
        sinal_mapear_todas_automaticamente_nesta_chave = True
        log_to_gui(f"Opção 3: '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}'. Próximas não mapeadas serão automáticas.", "INFO")
    elif escolha_opcao_str == "4":
        sinal_para_pular_todas_as_proximas = True
        log_to_gui(f"Opção 4: PULAR TODO o mapeamento para as próximas chaves deste PDF.", "INFO")
    else:
        log_to_gui(f"Opção inválida '{escolha_opcao_str}'. Pulando mapeamento para '{chave_original_nao_mapeada}'.", "INFO")
        return chave_original_nao_mapeada, False, False, False
    if nome_para_salvar_no_mapa and not sinal_para_pular_todas_as_proximas:
        mapeamento_completo = carregar_mapeamento_de_arquivo(caminho_arquivo_config_str) or {"mapeamento_para_chaves_padronizadas": {}}
        if "mapeamento_para_chaves_padronizadas" not in mapeamento_completo: mapeamento_completo["mapeamento_para_chaves_padronizadas"] = {}
        if chave_original_nao_mapeada not in mapeamento_completo["mapeamento_para_chaves_padronizadas"] or \
           mapeamento_completo["mapeamento_para_chaves_padronizadas"][chave_original_nao_mapeada] != nome_para_salvar_no_mapa:
            mapeamento_completo["mapeamento_para_chaves_padronizadas"][chave_original_nao_mapeada] = nome_para_salvar_no_mapa
            if salvar_mapeamento_em_arquivo(mapeamento_completo, caminho_arquivo_config_str):
                mapa_foi_atualizado_no_arquivo = True
                log_to_gui(f"Mapeamento salvo: '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}'", "DEBUG")
        else:
            mapa_foi_atualizado_no_arquivo = True
    return nome_padronizado_escolhido_para_uso_atual, mapa_foi_atualizado_no_arquivo, sinal_para_pular_todas_as_proximas, sinal_mapear_todas_automaticamente_nesta_chave

def normalizar_chaves_json(json_achatado_da_ia, mapeamento_chaves_a_usar_param, pular_mapeamento_interativo_ja_ativo):
    if not json_achatado_da_ia or not isinstance(json_achatado_da_ia, dict):
        log_to_gui("Normalizador: Nenhum JSON achatado válido.", "WARNING")
        return {}, dict(mapeamento_chaves_a_usar_param or {}), pular_mapeamento_interativo_ja_ativo
    mapeamento_local_atualizado = dict(mapeamento_chaves_a_usar_param or {})
    json_normalizado_para_placeholders = {}
    manter_sinal_pular_todas_proximas = pular_mapeamento_interativo_ja_ativo
    mapear_todas_automaticamente_global = False
    mapa_regex_mapeamento = {}
    if mapeamento_local_atualizado:
        for chave_mapa in sorted(mapeamento_local_atualizado.keys(), key=lambda k: '*' in k, reverse=True):
            if '*' in chave_mapa:
                regex_str = '^' + re.escape(chave_mapa).replace(r'\*', r'([0-9]+)') + '$'
                try: mapa_regex_mapeamento[chave_mapa] = re.compile(regex_str)
                except re.error as e: log_to_gui(f"ERRO Regex: '{chave_mapa}': {e}", "ERROR")
    for chave_achatada_original_str, valor_dado in json_achatado_da_ia.items():
        nome_chave_padrao_base = None
        if chave_achatada_original_str in mapeamento_local_atualizado and not mapear_todas_automaticamente_global:
            nome_chave_padrao_base = str(mapeamento_local_atualizado[chave_achatada_original_str])
        else:
            match_regex = False
            if mapa_regex_mapeamento and not mapear_todas_automaticamente_global:
                for chave_mapa_padrao_regex, regex_compilado in mapa_regex_mapeamento.items():
                    match = regex_compilado.match(chave_achatada_original_str)
                    if match:
                        match_regex = True
                        nome_base_mapa_regex = str(mapeamento_local_atualizado[chave_mapa_padrao_regex])
                        try:
                            indice_num_str = match.group(1)
                            nome_chave_padrao_base = nome_base_mapa_regex.replace("*", indice_num_str)
                        except Exception as e:
                            log_to_gui(f"AVISO Regex match: '{chave_mapa_padrao_regex}' com '{chave_achatada_original_str}', erro índice: {e}.", "WARNING")
                            match_regex = False
                        break
            if not nome_chave_padrao_base:
                if mapear_todas_automaticamente_global and chave_achatada_original_str not in mapeamento_local_atualizado:
                    nome_chave_padrao_base = chave_achatada_original_str.replace("_", " ").title()
                    mapeamento_local_atualizado[chave_achatada_original_str] = nome_chave_padrao_base
                    mapeamento_completo_arq = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG) or {"mapeamento_para_chaves_padronizadas": {}}
                    if "mapeamento_para_chaves_padronizadas" not in mapeamento_completo_arq: mapeamento_completo_arq["mapeamento_para_chaves_padronizadas"] = {}
                    mapeamento_completo_arq["mapeamento_para_chaves_padronizadas"][chave_achatada_original_str] = nome_chave_padrao_base
                    salvar_mapeamento_em_arquivo(mapeamento_completo_arq, ARQUIVO_MAPEAMENTO_CONFIG)
                    log_to_gui(f"MAPEAMENTO AUTOMÁTICO (Opção 3 Global): '{chave_achatada_original_str}' -> '{nome_chave_padrao_base}'", "INFO")
                elif manter_sinal_pular_todas_proximas:
                    nome_chave_padrao_base = chave_achatada_original_str
                elif chave_achatada_original_str not in mapeamento_local_atualizado :
                    nome_padronizado_usr, mapa_att_arq, pular_tudo_usr, mapear_todas_auto_nesta = \
                        gerenciar_chave_nao_mapeada_interativamente(
                            chave_achatada_original_str, ARQUIVO_MAPEAMENTO_CONFIG, mapeamento_local_atualizado)
                    if pular_tudo_usr: manter_sinal_pular_todas_proximas = True
                    if mapear_todas_auto_nesta: mapear_todas_automaticamente_global = True
                    if nome_padronizado_usr:
                        if mapa_att_arq or (nome_padronizado_usr != chave_achatada_original_str and not pular_tudo_usr):
                            mapeamento_local_atualizado[chave_achatada_original_str] = nome_padronizado_usr
                        nome_chave_padrao_base = nome_padronizado_usr
        if nome_chave_padrao_base is None: nome_chave_padrao_base = chave_achatada_original_str
        chave_final_placeholder = nome_chave_padrao_base.upper()
        for char_r, char_b in [(" ", "_"), ("/", "_"), ("-", "_"), ("(", ""), (")", ""), ("%", "PERC"), (".", "_"), (":", ""), ("*", "ITEM")]:
            chave_final_placeholder = chave_final_placeholder.replace(char_r, char_b)
        while "__" in chave_final_placeholder: chave_final_placeholder = chave_final_placeholder.replace("__", "_")
        chave_final_placeholder = re.sub(r'[^A-Z0-9_]', '', chave_final_placeholder)
        if chave_final_placeholder not in json_normalizado_para_placeholders:
            json_normalizado_para_placeholders[chave_final_placeholder] = valor_dado
        else:
            ct = 1; chave_conflito = f"{chave_final_placeholder}_DUPLICADO_{ct}"
            while chave_conflito in json_normalizado_para_placeholders: ct += 1; chave_conflito = f"{chave_final_placeholder}_DUPLICADO_{ct}"
            json_normalizado_para_placeholders[chave_conflito] = valor_dado
            log_to_gui(f"AVISO: Conflito placeholder '{chave_final_placeholder}', salvo como '{chave_conflito}'.", "WARNING")
    return json_normalizado_para_placeholders, mapeamento_local_atualizado, manter_sinal_pular_todas_proximas

def gerar_mapeamento_sugestao(json_achatado, nome_arquivo_origem="Desconhecido"):
    if not json_achatado or not isinstance(json_achatado, dict):
        log_to_gui("Mapeamento Sugestão: Nenhum JSON achatado.", "WARNING"); return {}
    log_to_gui(f"\n--- SUGESTÃO DE MAPEAMENTO para '{ARQUIVO_MAPEAMENTO_CONFIG}' (Base: {nome_arquivo_origem}) ---", "INFO")
    log_to_gui("{\n  \"mapeamento_para_chaves_padronizadas\": {", "INFO")
    sugestoes = {}; chaves_originais = list(json_achatado.keys())
    def nome_legivel(chave): return str(chave).replace("_", " ").title().replace("Cpf Cnpj", "CPF/CNPJ").replace("Src", "SRC")
    for chave_og in chaves_originais:
        match = re.match(r'^(.*?)_([0-9]+)(?:_(.*))?$', chave_og)
        if match:
            prefixo, idx_str, sufixo_parte = match.group(1), match.group(2), match.group(3) or ''
            chave_curinga = f"{prefixo}_*{'_' + sufixo_parte if sufixo_parte else ''}".strip('_')
            nome_amig_curinga = f"{nome_legivel(prefixo)}{f' - {nome_legivel(sufixo_parte)}' if sufixo_parte else ''} (Item *)"
            if chave_curinga not in sugestoes: sugestoes[chave_curinga] = nome_amig_curinga
    for chave_og in chaves_originais:
        coberta = any(re.fullmatch('^' + re.escape(sug_k).replace(r'\*', r'[0-9]+') + '$', chave_og)
                      for sug_k in sugestoes if '*' in sug_k)
        if not coberta and chave_og not in sugestoes: sugestoes[chave_og] = nome_legivel(chave_og)
    chaves_ord = sorted(sugestoes.keys(), key=lambda k: ('*' not in k, k))
    for i, chave_m in enumerate(chaves_ord):
        log_to_gui(f'    "{chave_m}": "{sugestoes[chave_m]}"{"," if i < len(chaves_ord) - 1 else ""}', "INFO")
    log_to_gui("  }\n}", "INFO")
    log_to_gui(f"--- Copie e edite em '{ARQUIVO_MAPEAMENTO_CONFIG}'. --- \n", "INFO")
    return sugestoes

def gerar_json_com_chaves_placeholder(json_dados_proc_norm, nome_arq_saida_path: Path):
    if not json_dados_proc_norm or not isinstance(json_dados_proc_norm, dict) :
        log_to_gui("JSON Placeholders: Nenhum dado normalizado.", "WARNING"); return False
    json_placeholders = {f"{{{{{chave_ph}}}}}" : val for chave_ph, val in json_dados_proc_norm.items()}
    try:
        with open(nome_arq_saida_path, "w", encoding="utf-8") as f: json.dump(json_placeholders, f, indent=2, ensure_ascii=False)
        log_to_gui(f"JSON com placeholders salvo em: '{nome_arq_saida_path.name}'", "INFO"); return True
    except Exception as e:
        log_to_gui(f"Erro ao salvar JSON com placeholders '{nome_arq_saida_path.name}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro Salvar JSON", f"Não foi possível salvar JSON placeholders.\n{e}")
        return False

# --- FUNÇÃO PRINCIPAL DE PROCESSAMENTO (COM LÓGICA DE PARTIÇÃO) ---
def processar_pdf_e_gerar_saidas(caminho_pdf_path_obj: Path):
    log_to_gui(f"--- Iniciando processamento para: {caminho_pdf_path_obj.name} ---", "INFO")
    iniciar_progresso()
    texto_completo_extraido = extrair_texto_do_pdf(caminho_pdf_path_obj)
    if not texto_completo_extraido:
        parar_progresso(f"Erro: Falha ao extrair texto de {caminho_pdf_path_obj.name}")
        if root.winfo_exists(): messagebox.showwarning("Aviso", "Nenhum texto pôde ser extraído do PDF.")
        return False
    log_to_gui(f"DEBUG: Primeiros 300 chars do texto para IA: {texto_completo_extraido[:300]}...", "DEBUG")
    log_to_gui(f"DEBUG: Últimos 300 chars do texto para IA: ...{texto_completo_extraido[-300:]}", "DEBUG")
    nome_arquivo_txt_completo = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_texto_completo_para_ia.txt"
    salvar_texto_em_arquivo(texto_completo_extraido, nome_arquivo_txt_completo)
    resultado_final_combinado_ia = {}
    sucesso_geral_api = True
    for i, nomes_blocos_nesta_parte in enumerate(LISTA_DE_NOMES_BLOCOS_PARTICIONADA):
        blocos_config_parcial = {nome_bloco: BLOCO_CONFIG[nome_bloco] for nome_bloco in nomes_blocos_nesta_parte if nome_bloco in BLOCO_CONFIG}
        if not blocos_config_parcial:
            log_to_gui(f"Parte {i+1} não tem blocos configurados para processar. Pulando.", "INFO")
            continue
        log_to_gui(f"Enviando Parte {i+1} ({len(blocos_config_parcial)} blocos) para a API Gemini...", "INFO")
        status_label.config(text=f"Processando Parte {i+1}/{len(LISTA_DE_NOMES_BLOCOS_PARTICIONADA)} com a IA...")
        if root.winfo_exists(): root.update_idletasks()
        resultado_parcial = enviar_texto_completo_para_gemini_todos_blocos(
            texto_completo_do_pdf=texto_completo_extraido,
            blocos_config_map=blocos_config_parcial,
            pdf_path_para_logs=caminho_pdf_path_obj
        )
        if resultado_parcial:
            log_to_gui(f"Parte {i+1} recebida e decodificada com sucesso.", "INFO")
            nome_arquivo_json_parcial_ia = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_resultado_parcial_{i+1}.json"
            salvar_json_em_arquivo(resultado_parcial, nome_arquivo_json_parcial_ia)
            resultado_final_combinado_ia.update(resultado_parcial)
        else:
            log_to_gui(f"ERRO: Falha ao processar Parte {i+1} dos blocos com a API Gemini.", "ERROR")
            sucesso_geral_api = False
            if root.winfo_exists():
                if not messagebox.askyesno("Erro API", f"Falha na Parte {i+1}. Continuar com as outras partes?"):
                    parar_progresso(f"Erro: Falha na API Gemini para Parte {i+1}")
                    return False 
            else: return False
    if not sucesso_geral_api and not resultado_final_combinado_ia:
        parar_progresso(f"Erro: Nenhuma parte do documento pôde ser processada pela API.")
        return False
    if not resultado_final_combinado_ia:
        parar_progresso(f"Erro: Nenhum dado foi retornado pela API após todas as tentativas parciais.")
        return False
    log_to_gui("Todas as partes processadas. Combinando resultados...", "INFO")
    nome_arquivo_json_bruto_ia = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_resultado_bruto_combinado.json"
    salvar_json_em_arquivo(resultado_final_combinado_ia, nome_arquivo_json_bruto_ia)
    status_label.config(text="Achatando JSON da IA...")
    if root.winfo_exists(): root.update_idletasks()
    json_achatado_da_ia = achatar_json(resultado_final_combinado_ia)
    if not json_achatado_da_ia:
        parar_progresso(f"Erro: JSON achatado da IA está vazio para {caminho_pdf_path_obj.name}")
        if root.winfo_exists(): messagebox.showerror("Erro de Processamento", "O JSON da IA achatado está vazio.")
        return False
    status_label.config(text="Normalizando chaves e aplicando mapeamento...")
    if root.winfo_exists(): root.update_idletasks()
    pular_todo_mapeamento_interativo_para_este_pdf = False
    mapeamento_carregado = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG)
    mapeamento_para_usar = None
    if mapeamento_carregado is None or \
       "mapeamento_para_chaves_padronizadas" not in mapeamento_carregado or \
       not isinstance(mapeamento_carregado.get("mapeamento_para_chaves_padronizadas"), dict):
        log_to_gui(f"Arquivo '{ARQUIVO_MAPEAMENTO_CONFIG}' não encontrado/inválido. Gerando sugestão...", "WARNING")
        mapeamento_sugerido = gerar_mapeamento_sugestao(json_achatado_da_ia, caminho_pdf_path_obj.name)
        if not mapeamento_sugerido: mapeamento_sugerido = {"EXEMPLO_CHAVE_IA": "Exemplo Placeholder"}
        mapeamento_para_salvar = {"mapeamento_para_chaves_padronizadas": mapeamento_sugerido}
        if salvar_mapeamento_em_arquivo(mapeamento_para_salvar, ARQUIVO_MAPEAMENTO_CONFIG):
            log_to_gui(f"Mapeamento padrão salvo em '{ARQUIVO_MAPEAMENTO_CONFIG}'.", "INFO")
            if root.winfo_exists(): messagebox.showinfo("Mapeamento Criado", f"'{ARQUIVO_MAPEAMENTO_CONFIG}' criado com sugestões. Edite-o para refinar.", parent=root)
        mapeamento_para_usar = mapeamento_sugerido
    else:
        mapeamento_para_usar = mapeamento_carregado["mapeamento_para_chaves_padronizadas"]
        log_to_gui(f"Usando mapeamento de '{ARQUIVO_MAPEAMENTO_CONFIG}'.", "INFO")
    json_final_normalizado_para_placeholders, _, _ = normalizar_chaves_json(
        json_achatado_da_ia,
        mapeamento_para_usar,
        pular_todo_mapeamento_interativo_para_este_pdf
    )
    if not json_final_normalizado_para_placeholders:
        parar_progresso("Erro na normalização do JSON para placeholders.")
        if root.winfo_exists(): messagebox.showerror("Erro de Normalização", "A normalização das chaves para placeholders falhou.")
        return False
    nome_arquivo_json_processado_debug = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_processado_final_normalizado_debug.json"
    salvar_json_em_arquivo(json_final_normalizado_para_placeholders, nome_arquivo_json_processado_debug)
    status_label.config(text="Gerando JSON de saída com chaves placeholder...")
    if root.winfo_exists(): root.update_idletasks()
    nome_arquivo_json_saida_placeholders = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_dados_com_placeholders.json"
    sucesso_ao_gerar_json_placeholder = gerar_json_com_chaves_placeholder(
        json_final_normalizado_para_placeholders,
        nome_arquivo_json_saida_placeholders
    )
    if sucesso_ao_gerar_json_placeholder:
        log_to_gui("Pronto para preencher Excel. Selecione o ARQUIVO EXCEL MODELO...", "INFO")
        caminho_template_excel_str = filedialog.askopenfilename(parent=root, title="Selecione ARQUIVO EXCEL MODELO", filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
        if caminho_template_excel_str:
            caminho_template_excel = Path(caminho_template_excel_str)
            nome_saida_sugerido = f"{caminho_pdf_path_obj.stem}_PREENCHIDO.xlsx"
            caminho_saida_excel_str = filedialog.asksaveasfilename(parent=root, title="Salvar NOVO Arquivo Excel Preenchido Como...", initialdir=str(caminho_pdf_path_obj.parent), initialfile=nome_saida_sugerido, defaultextension=".xlsx", filetypes=(("Arquivos Excel", "*.xlsx"),))
            if caminho_saida_excel_str:
                caminho_saida_excel = Path(caminho_saida_excel_str)
                nome_aba = simpledialog.askstring("Nome da Aba", "Nome da ABA no Excel (deixe em branco para ativa):", parent=root)
                nome_planilha_alvo = nome_aba.strip() if nome_aba and nome_aba.strip() else None
                preenchimento_ok = preencher_excel_novo_com_placeholders(nome_arquivo_json_saida_placeholders, caminho_template_excel, caminho_saida_excel, nome_planilha_alvo)
            else: parar_progresso("Salvamento do Excel cancelado.")
        else: parar_progresso("Seleção de modelo Excel cancelada.")
    else:
        parar_progresso(f"Falha ao gerar JSON com placeholders para {caminho_pdf_path_obj.name}")
    log_to_gui(f"--- Fim do processamento para: {caminho_pdf_path_obj.name} ---", "INFO")
    return True

# --- FUNÇÕES DA INTERFACE E MAINLOOP ---
def main_loop_wrapper():
    while True:
        log_to_gui("\n--- Novo Ciclo de Análise Iniciado ---", "INFO")
        if not genai_config_ok:
            if root.winfo_exists(): messagebox.showerror("Erro de API", "API Key do Google Gemini não configurada.")
            if not messagebox.askyesno("API Não Configurada", "API Key não configurada. Tentar selecionar PDF?"): break
        status_label.config(text="Aguardando seleção do PDF...")
        caminho_pdf_str = filedialog.askopenfilename(parent=root, title="Selecione o PDF de ENTRADA", filetypes=(("Arquivos PDF", "*.pdf"),))
        if not caminho_pdf_str:
            log_to_gui("Nenhum PDF selecionado.", "INFO")
            if not messagebox.askyesno("Nenhum PDF", "Nenhum PDF. Tentar novamente?"): break
            continue
        caminho_pdf = Path(caminho_pdf_str)
        sucesso = processar_pdf_e_gerar_saidas(caminho_pdf)
        if not messagebox.askyesno("Análise Concluída", f"Processamento de '{caminho_pdf.name}' {'concluído.' if sucesso else 'com falhas.'}\nOutra análise?"): break
    log_to_gui("Loop de análise encerrado.", "INFO")
    status_label.config(text="Pronto.")

def mostrar_sobre():
    if root.winfo_exists(): messagebox.showinfo("Sobre", f"Processador de Súmulas v2.8 (Final)\nConfig: {ARQUIVO_MAPEAMENTO_CONFIG}\nLog: {LOG_FILE_NAME}", parent=root)

def sair_aplicacao():
    if root.winfo_exists() and messagebox.askokcancel("Sair", "Fechar a aplicação?", parent=root):
        log_to_gui("Aplicação encerrando...", "INFO")
        root.quit()
        root.destroy()

def abrir_arquivo_mapeamento_para_edicao():
    caminho_config = Path(ARQUIVO_MAPEAMENTO_CONFIG).resolve()
    if not caminho_config.is_file():
        if root.winfo_exists(): messagebox.showinfo("Arquivo Não Existe", f"'{ARQUIVO_MAPEAMENTO_CONFIG}' será criado se necessário.", parent=root)
        return
    try:
        log_to_gui(f"Abrindo '{caminho_config}' para edição...", "INFO")
        if sys.platform == "win32": os.startfile(str(caminho_config))
        elif sys.platform == "darwin": subprocess.run(["open", str(caminho_config)], check=True)
        else: subprocess.run(["xdg-open", str(caminho_config)], check=True)
    except Exception as e:
        log_to_gui(f"ERRO ao abrir '{caminho_config}': {e}", "ERROR")
        if root.winfo_exists(): messagebox.showerror("Erro ao Abrir", f"Não foi possível abrir '{caminho_config.name}'.\n{caminho_config}")

# --- 8. CONFIGURAÇÃO FINAL DA GUI (Botões, Menu, mainloop) ---
texto_botao_analisar = "Analisar PDF e Gerar Saídas"
texto_botao_editar_mapeamento = "Editar Mapeamento de Chaves"
fator_largura_caractere = 1.0
largura_calculada_analisar = int(len(texto_botao_analisar) * fator_largura_caractere)
largura_calculada_editar = int(len(texto_botao_editar_mapeamento) * fator_largura_caractere)
largura_botao_final = max(largura_calculada_analisar, largura_calculada_editar, 28)

style.configure("Primary.TButton", font=FONTE_BOTAO_PRINCIPAL, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE))
style.map("Primary.TButton",
          background=[('active', COR_BOTAO_PRIMARIO_BG), ('!disabled', COR_BOTAO_PRIMARIO_BG)],
          foreground=[('!disabled', COR_BOTAO_PRIMARIO_FG)])
style.configure("Secondary.TButton", font=FONTE_BOTAO_SECUNDARIO, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE))
style.map("Secondary.TButton",
          background=[('active', '#FFD54F'), ('!disabled', COR_BOTAO_SECUNDARIO_BG)],
          foreground=[('!disabled', COR_BOTAO_SECUNDARIO_FG)])

botao_analisar = ttk.Button(
    frame_botoes_principais, text=texto_botao_analisar,
    command=main_loop_wrapper,
    style="Primary.TButton",
    width=largura_botao_final
)
botao_analisar.pack(side=tk.LEFT, padx=10, pady=5)

botao_editar_mapeamento = ttk.Button(
    frame_botoes_principais, text=texto_botao_editar_mapeamento,
    command=abrir_arquivo_mapeamento_para_edicao,
    style="Secondary.TButton",
    width=largura_botao_final
)
botao_editar_mapeamento.pack(side=tk.LEFT, padx=10, pady=5)

menu_bar = tk.Menu(root, font=FONTE_MENU)
menu_arquivo = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
try:
    script_dir = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).resolve().parent
    caminho_icone = script_dir / "assets" / "icone_pdf.png"
    if caminho_icone.is_file():
        icone_pdf_original = PhotoImage(file=caminho_icone)
        icone_pdf = icone_pdf_original
        menu_arquivo.add_command(label="Analisar PDF...", image=icone_pdf, compound="left", command=main_loop_wrapper)
    else:
        log_to_gui(f"AVISO: Ícone '{caminho_icone}' não encontrado.", "WARNING")
        menu_arquivo.add_command(label="Analisar PDF...", command=main_loop_wrapper)
except Exception as e_icon:
    log_to_gui(f"AVISO: Erro ao carregar ícone: {e_icon}.", "WARNING")
    menu_arquivo.add_command(label="Analisar PDF...", command=main_loop_wrapper)
menu_arquivo.add_command(label="Editar Mapeamento de Chaves", command=abrir_arquivo_mapeamento_para_edicao)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="Sair", command=sair_aplicacao)
menu_bar.add_cascade(label="Arquivo", menu=menu_arquivo)
menu_ajuda = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
menu_ajuda.add_command(label="Sobre", command=mostrar_sobre)
menu_bar.add_cascade(label="Ajuda", menu=menu_ajuda)
root.config(menu=menu_bar)
root.protocol("WM_DELETE_WINDOW", sair_aplicacao)

if __name__ == "__main__":
    log_to_gui("Interface Gráfica Pronta. Use o menu 'Arquivo' ou o botão 'Analisar PDF' para iniciar.", "INFO")
    root.mainloop()