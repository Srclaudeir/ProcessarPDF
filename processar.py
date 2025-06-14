# --- 0. IMPORTAÇÕES DE BIBLIOTECAS ---
import pdfplumber
import google.generativeai as genai
import json
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import re
import os
import sys
import subprocess
import openpyxl
from dotenv import load_dotenv
import time
import logging
from tenacity import retry, stop_after_attempt, wait_exponential, RetryError
from typing import Any, Dict, List, Optional, Tuple, Union
from PIL import Image, ImageTk

# --- Inicialização preventiva ---
progress: Optional[ttk.Progressbar] = None
status_label: Optional[ttk.Label] = None
_root_ref_for_log: Optional[tk.Tk] = None
_log_text_widget_ref: Optional[tk.Text] = None
# Referência global para o botão principal, para poder desabilitá-lo
botao_analisar: Optional[ttk.Button] = None


# --- HELPER PARA PYINSTALLER ---
def resource_path(relative_path: str) -> Path:
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller cria um diretório temporário e armazena o path em _MEIPASS
        base_path = Path(sys._MEIPASS) # type: ignore
    except AttributeError:
        # _MEIPASS não está definido, rodando em ambiente de desenvolvimento normal
        base_path = Path(__file__).resolve().parent
    return base_path / relative_path

# --- CONSTANTES ---
CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS = False # Mudar para True para salvar arquivos intermediários
ARQUIVO_MAPEAMENTO_CONFIG = "mapeamento_config.json"
ARQUIVO_SCHEMA_EXTRACAO = "extraction_schema.json"
LOG_FILE_PATH = resource_path("processamento_pdf.log")
MAX_TEXT_LENGTH_IA = 60 # Limite de caracteres para campos de texto longo na IA
MARCADOR_INICIO_TEXTO_PDF_PROMPT = "[INICIO_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
MARCADOR_FIM_TEXTO_PDF_PROMPT = "[FIM_TEXTO_DOCUMENTO_SICOOB_XYZ123]"
INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES = True # Se True, insere "N/A" no Excel para placeholders não encontrados nos dados
GEMINI_TEMPERATURE = 0.1 # Baixa para respostas mais determinísticas
GEMINI_MAX_OUTPUT_TOKENS = 8190 # Limite máximo de tokens para a resposta da IA
GEMINI_MODEL_NAME = 'gemini-1.5-flash-latest' # Modelo Gemini a ser utilizado

# --- CORES E FONTES PARA A GUI ---
COR_FUNDO_JANELA = "#F0F0F0"
COR_FUNDO_FRAMES_INTERNOS = "#F8F8F8"
COR_TEXTO_PADRAO = "#333333"
COR_TEXTO_TITULO_GUI = "#333333"
COR_TEXTO_LOG = "#1A1A1A"
COR_FUNDO_LOG = "#FFFFFF"
COR_BOTAO_PRIMARIO_BG = "#0078D4"
COR_BOTAO_PRIMARIO_FG = "#FFFFFF"
COR_BOTAO_SECUNDARIO_BG = "#FFC107"
COR_BOTAO_SECUNDARIO_FG = "#000000"
COR_BOTAO_HOVER_PRIMARIO = "#005FA3"
COR_BOTAO_HOVER_SECUNDARIO = "#FFD54F"
COR_STATUS_LABEL_FG = "#005A9E"
FONTE_TITULO_APP = ("Segoe UI", 18, "bold")
FONTE_SUBTITULO = ("Segoe UI", 10)
FONTE_BOTAO_PRINCIPAL = ("Segoe UI", 10, "bold")
FONTE_BOTAO_SECUNDARIO = ("Segoe UI", 10, "bold")
FONTE_STATUS = ("Segoe UI", 10, "italic")
FONTE_LOG = ("Consolas", 9)
FONTE_MENU = ("Segoe UI", 9)
FONTE_LABELFRAME_TITULO = ("Segoe UI", 10, "bold")
PADDING_X_BOTAO_STYLE = 15
PADDING_Y_BOTAO_STYLE = 5

# --- CONFIGURAÇÃO DO LOGGING ---
logging.basicConfig(
    level=logging.INFO, # Mudar para logging.DEBUG para mais detalhes durante desenvolvimento
    format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(lineno)d - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, encoding='utf-8', mode='w'), # 'w' para log limpo a cada execução
        logging.StreamHandler(sys.stdout) # Também loga para o console
    ]
)
logging.info("!!! CONFIGURAÇÃO DE LOGGING EXECUTADA NO TOPO DO SCRIPT !!!")


# --- 4. FUNÇÕES DE FEEDBACK DA INTERFACE GRÁFICA (log, progresso) ---

def is_gui_widget_available(widget: Optional[Union[tk.Tk, tk.Widget]]) -> bool:
    """Verifica se um widget da GUI não é None e ainda existe (não foi destruído)."""
    return widget is not None and widget.winfo_exists()

def setup_gui_logging_refs(root_window: tk.Tk, log_widget: tk.Text):
    """Configura referências globais para a janela raiz e o widget de log da GUI."""
    global _root_ref_for_log, _log_text_widget_ref
    _root_ref_for_log = root_window
    _log_text_widget_ref = log_widget

def log_to_gui(mensagem: str, level: str = "INFO") -> None:
    """Envia uma mensagem para o widget de log na GUI e para o logger backend."""
    formatted_message = f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {level} - {mensagem}"

    # Tenta logar na GUI se os widgets estiverem disponíveis
    if is_gui_widget_available(_root_ref_for_log) and is_gui_widget_available(_log_text_widget_ref):
        try:
            # Reafirma que os widgets são os tipos esperados antes de usar métodos específicos
            if isinstance(_log_text_widget_ref, tk.Text) and isinstance(_root_ref_for_log, tk.Tk):
                _log_text_widget_ref.configure(state=tk.NORMAL)
                _log_text_widget_ref.insert(tk.END, formatted_message + "\n")
                _log_text_widget_ref.see(tk.END) # Rola para a última mensagem
                _log_text_widget_ref.configure(state=tk.DISABLED)
                _root_ref_for_log.update_idletasks() # Atualiza a GUI para mostrar a mensagem
        except tk.TclError:
            # Comum se a GUI estiver no processo de ser fechada
            logging.warning(f"Falha ao tentar logar na GUI (TclError): {mensagem}")
        except Exception as e_gui_log:
            # Pega outros erros inesperados ao tentar logar na GUI
            logging.error(f"Erro inesperado ao logar na GUI: {e_gui_log} - Mensagem: {mensagem}")

    # Loga no backend (arquivo/console) independentemente da GUI
    # Usa a mensagem original, pois o logger do Python já adiciona timestamp/level
    if level == "INFO": logging.info(mensagem)
    elif level == "WARNING": logging.warning(mensagem)
    elif level == "ERROR": logging.error(mensagem)
    elif level == "DEBUG": logging.debug(mensagem)
    elif level == "CRITICAL": logging.critical(mensagem)
    else: logging.info(f"({level}) {mensagem}") # Nível de log desconhecido


def iniciar_progresso() -> None:
    """Inicia a barra de progresso indeterminada e atualiza o status na GUI."""
    if is_gui_widget_available(_root_ref_for_log) and \
       is_gui_widget_available(progress) and \
       is_gui_widget_available(status_label):
        if isinstance(progress, ttk.Progressbar) and isinstance(status_label, ttk.Label) and \
           isinstance(_root_ref_for_log, tk.Tk):
            progress.start(20) # Intervalo de atualização da animação
            status_label.config(text="Processando, por favor aguarde...")
            _root_ref_for_log.update_idletasks()

def parar_progresso(final_status: str = "") -> None:
    """Para a barra de progresso e define uma mensagem de status final na GUI."""
    if is_gui_widget_available(_root_ref_for_log) and \
       is_gui_widget_available(progress) and \
       is_gui_widget_available(status_label):
        if isinstance(progress, ttk.Progressbar) and isinstance(status_label, ttk.Label) and \
           isinstance(_root_ref_for_log, tk.Tk):
            progress.stop()
            status_label.config(text=final_status or "Pronto.") # "Pronto." se nenhum status final for fornecido
            _root_ref_for_log.update_idletasks()

def limpar_log_gui() -> None:
    """Limpa o conteúdo do widget de log na GUI após confirmação do usuário."""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if parent_dialog and is_gui_widget_available(_log_text_widget_ref) and isinstance(_log_text_widget_ref, tk.Text):
        confirmed = messagebox.askyesno(
            "Limpar Log da Tela",
            "Tem certeza que deseja limpar todo o log da tela?\n(O arquivo de log em disco não será afetado)",
            parent=parent_dialog
        )
        if confirmed:
            _log_text_widget_ref.configure(state=tk.NORMAL)
            _log_text_widget_ref.delete('1.0', tk.END)
            _log_text_widget_ref.configure(state=tk.DISABLED)
            log_to_gui("Log da tela limpo pelo usuário.", "INFO")
    else:
        logging.warning("Tentativa de limpar log da GUI, mas widget de log ou janela raiz não disponível.")

# --- 1. CARREGAR VARIÁVEIS DE AMBIENTE (.env) ---
dotenv_path_obj = resource_path(".env")
if dotenv_path_obj.exists():
    load_dotenv(dotenv_path=dotenv_path_obj)
    logging.info(f"Arquivo .env carregado de: {dotenv_path_obj}")
else:
    logging.warning(f"Arquivo .env não encontrado em: {dotenv_path_obj}. Usando variáveis de ambiente do sistema, se disponíveis.")

# --- 2. CONFIGURAÇÃO DA API KEY DO GOOGLE ---
GOOGLE_API_KEY: Optional[str] = os.getenv("GOOGLE_API_KEY")

# --- BLOCO DE CONFIGURAÇÃO (CARREGADO DE ARQUIVO EXTERNO) ---
BLOCO_CONFIG: Dict[str, Any] = {} # Dicionário para armazenar a configuração do schema de extração

def carregar_schema_extracao() -> bool:
    """Carrega e valida o schema de extração do arquivo JSON especificado."""
    global BLOCO_CONFIG
    caminho_schema_abs = resource_path(ARQUIVO_SCHEMA_EXTRACAO)
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    try:
        with open(caminho_schema_abs, "r", encoding="utf-8") as f:
            BLOCO_CONFIG_RAW = json.load(f) # Carrega o JSON bruto

        if not isinstance(BLOCO_CONFIG_RAW, dict): # O arquivo deve ser um dicionário na raiz
            log_to_gui(f"ERRO CRÍTICO: Conteúdo do schema '{ARQUIVO_SCHEMA_EXTRACAO}' não é um dicionário JSON no nível raiz.", "CRITICAL")
            if parent_dialog:
                messagebox.showerror("Erro de Configuração do Schema", f"O arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}' não contém um dicionário JSON válido no nível raiz.", parent=parent_dialog)
            return False

        BLOCO_CONFIG_VALIDADO: Dict[str, Any] = {} # Armazena apenas os blocos válidos
        schema_geral_valido = True

        for nome_bloco, config_bloco in BLOCO_CONFIG_RAW.items(): # Valida cada bloco no schema
            bloco_atual_valido = True
            if not isinstance(config_bloco, dict):
                log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}' em '{ARQUIVO_SCHEMA_EXTRACAO}' não é um dicionário. Bloco ignorado.", "ERROR")
                schema_geral_valido = False; continue

            # Validação de chaves obrigatórias e seus tipos
            chaves_obrigatorias = {"json_chave": str, "particao": int}
            for chave, tipo_esperado in chaves_obrigatorias.items():
                if chave not in config_bloco:
                    log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}' não possui a chave obrigatória '{chave}'. Bloco ignorado.", "ERROR")
                    bloco_atual_valido = False; break
                if not isinstance(config_bloco[chave], tipo_esperado):
                    log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}', chave '{chave}', esperava tipo '{tipo_esperado.__name__}' mas obteve '{type(config_bloco[chave]).__name__}'. Bloco ignorado.", "ERROR")
                    bloco_atual_valido = False; break
            if not bloco_atual_valido: schema_geral_valido = False; continue

            # Validação de 'particao'
            if not (isinstance(config_bloco.get("particao"), int) and config_bloco.get("particao", 0) > 0):
                log_to_gui(f"ERRO SCHEMA: Bloco '{nome_bloco}', chave 'particao' deve ser um inteiro maior que 0. Valor: {config_bloco.get('particao')}. Bloco ignorado.", "ERROR")
                schema_geral_valido = False; continue

            # Validação de tipos para chaves opcionais (se presentes e não None)
            chaves_opcionais_com_tipo = {
                "titulo_padrao": str, "campos_esperados": list, "nome_lista_json": str,
                "sub_campos_lista": list, "campos_texto_longo_limitar": list, "sub_lista_aninhada": dict
            }
            for chave, tipo_esperado in chaves_opcionais_com_tipo.items():
                if chave in config_bloco and config_bloco[chave] is not None and not isinstance(config_bloco[chave], tipo_esperado):
                    log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', chave opcional '{chave}', esperava tipo '{tipo_esperado.__name__}' mas obteve '{type(config_bloco[chave]).__name__}'. Chave será usada com cautela ou ignorada se causar erro.", "WARNING")

            # Validação de 'sub_lista_aninhada' (se presente)
            if "sub_lista_aninhada" in config_bloco and isinstance(config_bloco.get("sub_lista_aninhada"), dict):
                sub_lista_conf = config_bloco["sub_lista_aninhada"]
                sub_chaves_obrigatorias = {"nome_json": str, "campos": list}
                for sub_chave, sub_tipo_esperado in sub_chaves_obrigatorias.items():
                    if sub_chave not in sub_lista_conf:
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada' não possui a chave '{sub_chave}'. Sub-lista pode não funcionar como esperado.", "WARNING"); break
                    if not isinstance(sub_lista_conf[sub_chave], sub_tipo_esperado):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada', chave '{sub_chave}', esperava tipo '{sub_tipo_esperado.__name__}' mas obteve '{type(sub_lista_conf[sub_chave]).__name__}'.", "WARNING"); break
                    if sub_chave == "campos" and isinstance(sub_lista_conf[sub_chave], list) and \
                       not all(isinstance(item, str) for item in sub_lista_conf[sub_chave]):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', 'sub_lista_aninhada', chave 'campos' deve conter apenas strings.", "WARNING")

            # Validação de que listas de strings contêm apenas strings
            for chave_lista_str in ["campos_esperados", "sub_campos_lista", "campos_texto_longo_limitar"]:
                if chave_lista_str in config_bloco and config_bloco[chave_lista_str] is not None and isinstance(config_bloco.get(chave_lista_str), list):
                    if not all(isinstance(item, str) for item in config_bloco[chave_lista_str]):
                        log_to_gui(f"AVISO SCHEMA: Bloco '{nome_bloco}', chave '{chave_lista_str}' deve conter uma lista de strings. Encontrados outros tipos.", "WARNING")

            if bloco_atual_valido:
                BLOCO_CONFIG_VALIDADO[nome_bloco] = config_bloco
            else: # Se o bloco atual não passou na validação, marca o schema geral como não totalmente válido
                schema_geral_valido = False

        BLOCO_CONFIG = BLOCO_CONFIG_VALIDADO # Atualiza a configuração global com os blocos validados

        if not schema_geral_valido and BLOCO_CONFIG_RAW: # Se houve erros/avisos, mas o arquivo não estava vazio
            log_to_gui(f"AVISO: O schema '{ARQUIVO_SCHEMA_EXTRACAO}' contém erros ou avisos. Alguns blocos podem ter sido ignorados ou podem não funcionar como esperado. Verifique os logs.", "WARNING")
            if parent_dialog:
                 messagebox.showwarning("Aviso sobre o Schema de Extração", f"O arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}' contém erros ou avisos. Alguns blocos podem ter sido ignorados. Verifique os logs para detalhes.", parent=parent_dialog)

        if not BLOCO_CONFIG and BLOCO_CONFIG_RAW: # Se o arquivo não estava vazio, mas nenhum bloco foi validado
            log_to_gui(f"ERRO FATAL: Nenhum bloco válido foi encontrado no schema '{ARQUIVO_SCHEMA_EXTRACAO}' após a validação.", "CRITICAL")
            if parent_dialog:
                messagebox.showerror("Erro Fatal de Schema", f"Nenhum bloco válido foi encontrado no schema '{ARQUIVO_SCHEMA_EXTRACAO}'. A aplicação não pode processar PDFs sem um schema válido.", parent=parent_dialog)
            return False
        elif not BLOCO_CONFIG and not BLOCO_CONFIG_RAW: # Se o arquivo de schema estava completamente vazio
             log_to_gui(f"INFO: O arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}' está vazio. Nenhum bloco de extração definido.", "INFO")
        elif BLOCO_CONFIG: # Se há blocos válidos
            log_to_gui(f"Schema '{ARQUIVO_SCHEMA_EXTRACAO}' carregado e validado. Número de blocos válidos: {len(BLOCO_CONFIG)}", "INFO")
        return True

    except FileNotFoundError:
        log_to_gui(f"ERRO CRÍTICO: Arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}' não encontrado no caminho esperado: '{caminho_schema_abs}'.", "CRITICAL")
        if parent_dialog: messagebox.showerror("Erro de Configuração", f"Arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}' não encontrado.", parent=parent_dialog)
        return False
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO CRÍTICO: Falha ao decodificar JSON no arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}': {e.msg} na linha {e.lineno}, coluna {e.colno}", "CRITICAL")
        if parent_dialog: messagebox.showerror("Erro de Configuração", f"Erro ao ler o arquivo de schema '{ARQUIVO_SCHEMA_EXTRACAO}'. JSON inválido: {e.msg}", parent=parent_dialog)
        return False
    except Exception as e: # Pega outros erros inesperados
        log_to_gui(f"ERRO CRÍTICO e inesperado ao carregar ou validar o schema '{ARQUIVO_SCHEMA_EXTRACAO}': {e}", "CRITICAL")
        logging.exception("Erro detalhado ao carregar/validar schema:")
        if parent_dialog: messagebox.showerror("Erro de Configuração", f"Erro fatal e inesperado ao carregar o schema '{ARQUIVO_SCHEMA_EXTRACAO}': {e}", parent=parent_dialog)
        return False

# --- DEFINIÇÃO DAS PARTIÇÕES DO BLOCO_CONFIG (DINÂMICO) ---
# Usado para organizar o schema e logs, não para múltiplas chamadas de API.
LISTA_DE_NOMES_BLOCOS_PARTICIONADA: List[List[str]] = []
def gerar_particoes_dinamicamente() -> bool:
    """Organiza os nomes dos blocos do schema em 'partições' lógicas baseadas na chave 'particao'."""
    global LISTA_DE_NOMES_BLOCOS_PARTICIONADA
    if not BLOCO_CONFIG:
        log_to_gui("AVISO: Schema (BLOCO_CONFIG) está vazio. Nenhuma partição lógica para gerar.", "WARNING")
        LISTA_DE_NOMES_BLOCOS_PARTICIONADA = []
        return True # Considerado sucesso, pois não há o que particionar

    max_particao_num = 0
    for config_bloco in BLOCO_CONFIG.values(): # Encontra o maior número de partição definido
        num_part = config_bloco.get("particao")
        if isinstance(num_part, int) and num_part > 0:
            if num_part > max_particao_num:
                max_particao_num = num_part

    if max_particao_num == 0: # Se nenhum bloco tem 'particao' > 0
        all_block_names = list(BLOCO_CONFIG.keys())
        if all_block_names:
            log_to_gui(f"AVISO: Nenhum bloco no schema possui uma 'particao' numérica válida (>0). Todos os {len(all_block_names)} blocos serão agrupados na partição lógica 1.", "WARNING")
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [all_block_names]
        else: # Nenhum bloco no schema
            log_to_gui("AVISO: Nenhum bloco encontrado no schema para particionar logicamente.", "WARNING")
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA = []
        return True

    # Inicializa a lista de partições (listas de nomes de blocos)
    LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [[] for _ in range(max_particao_num)]
    blocos_com_particao_invalida = []

    for nome_bloco, config_bloco in BLOCO_CONFIG.items():
        num_part = config_bloco.get("particao")
        if isinstance(num_part, int) and 1 <= num_part <= max_particao_num:
            LISTA_DE_NOMES_BLOCOS_PARTICIONADA[num_part - 1].append(nome_bloco)
        else:
            blocos_com_particao_invalida.append(nome_bloco)

    if blocos_com_particao_invalida:
        log_to_gui(f"AVISO: Os seguintes blocos possuem valores de 'particao' inválidos e foram ignorados no agrupamento lógico: {', '.join(blocos_com_particao_invalida)}", "WARNING")

    # Remove partições que ficaram vazias (se houver)
    LISTA_DE_NOMES_BLOCOS_PARTICIONADA = [p_list for p_list in LISTA_DE_NOMES_BLOCOS_PARTICIONADA if p_list]

    if not LISTA_DE_NOMES_BLOCOS_PARTICIONADA and BLOCO_CONFIG:
         log_to_gui("ERRO: Nenhuma partição lógica foi gerada, embora o schema contenha blocos. Verifique os valores da chave 'particao' em todos os blocos do schema.", "ERROR")
         return False # Indica falha se havia blocos mas nenhuma partição foi formada

    log_to_gui(f"Partições lógicas (para organização do schema/log) geradas: {len(LISTA_DE_NOMES_BLOCOS_PARTICIONADA)} ativas.", "INFO")
    for i, part_list in enumerate(LISTA_DE_NOMES_BLOCOS_PARTICIONADA):
        log_to_gui(f"  Partição lógica {i+1} contém {len(part_list)} blocos.", "DEBUG")
    return True


# --- 3. CONFIGURAÇÃO INICIAL DA INTERFACE GRÁFICA (Widgets Globais) ---
root = tk.Tk()
_root_ref_for_log = root

root.title(f"Processador de Súmulas de Crédito - Gemini v{GEMINI_MODEL_NAME.split('-')[-2] if GEMINI_MODEL_NAME and '-' in GEMINI_MODEL_NAME else '?.?'}")
root.geometry("900x750")
root.configure(bg=COR_FUNDO_JANELA)

# Configuração de Estilos ttk
style = ttk.Style()
style.theme_use('clam')
style.configure("TButton", padding=6, relief="flat", borderwidth=0)
style.configure("TLabel", font=FONTE_SUBTITULO, background=COR_FUNDO_JANELA, foreground=COR_TEXTO_PADRAO)
style.configure("TProgressbar", thickness=15, background=COR_BOTAO_PRIMARIO_BG)
style.configure("Title.TLabel", font=FONTE_TITULO_APP, foreground=COR_TEXTO_TITULO_GUI, background=COR_FUNDO_JANELA)
style.configure("Header.TFrame", background=COR_FUNDO_JANELA)
style.configure("Controls.TFrame", background=COR_FUNDO_JANELA)
style.configure("Status.TFrame", background=COR_FUNDO_JANELA)
style.configure("TLabelframe", background=COR_FUNDO_FRAMES_INTERNOS, relief=tk.GROOVE, borderwidth=1)
style.configure("TLabelframe.Label", font=FONTE_LABELFRAME_TITULO, background=COR_FUNDO_FRAMES_INTERNOS, foreground=COR_TEXTO_PADRAO, padding=(5,2))

# Frame do Topo (Título, Logo, Instrução, Botões Principais)
frame_topo = ttk.Frame(root, padding=(20, 10), style="Header.TFrame")
frame_topo.pack(pady=(10,0), fill=tk.X)

# Seção Logo e Título
logging.info("--- INICIANDO BLOCO DE CARREGAMENTO DO LOGO PRINCIPAL (GUI) ---")
frame_titulo_com_logo = ttk.Frame(frame_topo, style="Header.TFrame")
frame_titulo_com_logo.pack(pady=(0, 5), anchor=tk.CENTER)

logo_app_image_tk: Optional[ImageTk.PhotoImage] = None
caminho_logo_relativo = "assets/logo_sicoob.png"
NOVO_LARGURA_LOGO = 45
NOVO_ALTURA_LOGO = 45

try:
    caminho_logo_abs = resource_path(caminho_logo_relativo)
    if caminho_logo_abs.is_file():
        img_pil = Image.open(caminho_logo_abs)
        img_redimensionada_pil = img_pil.resize((NOVO_LARGURA_LOGO, NOVO_ALTURA_LOGO), Image.Resampling.LANCZOS)
        logo_app_image_tk = ImageTk.PhotoImage(img_redimensionada_pil)

        label_logo = ttk.Label(frame_titulo_com_logo, image=logo_app_image_tk)
        label_logo.image = logo_app_image_tk
        label_logo.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        logging.info(f"Logo '{caminho_logo_relativo}' carregada.")
    else:
        logging.warning(f"Arquivo de logo '{caminho_logo_abs}' NÃO ENCONTRADO.")
        log_to_gui(f"AVISO: Arquivo de logo '{caminho_logo_abs}' NÃO ENCONTRADO.", "WARNING")
except Exception as e_logo:
    logging.error(f"Erro ao carregar/redimensionar logo '{caminho_logo_relativo}': {e_logo}", exc_info=True)
    log_to_gui(f"AVISO: Erro ao carregar/redimensionar logo '{caminho_logo_relativo}': {e_logo}", "WARNING")

label_titulo_app = ttk.Label(frame_titulo_com_logo, text="Processador de Súmulas de Crédito", style="Title.TLabel")
label_titulo_app.pack(side=tk.LEFT, pady=5)
logging.info("--- FIM DO BLOCO DE CARREGAMENTO DO LOGO PRINCIPAL (GUI) ---")

label_instrucao_app = ttk.Label(frame_topo, text="Selecione um PDF para análise. O log do processamento aparecerá abaixo.", justify=tk.CENTER)
label_instrucao_app.pack(pady=(5,10))

# Botões Principais
frame_botoes_principais = ttk.Frame(frame_topo, style="Controls.TFrame")
frame_botoes_principais.pack(pady=10, fill=tk.X, padx=20)

# Status e Barra de Progresso
frame_status_progresso = ttk.Frame(root, padding=(20,5), style="Status.TFrame")
frame_status_progresso.pack(fill=tk.X, padx=10)
progress = ttk.Progressbar(frame_status_progresso, orient="horizontal", length=500, mode="indeterminate")
progress.pack(pady=5, fill=tk.X, expand=True)
status_label = ttk.Label(frame_status_progresso, text="Pronto para iniciar.", foreground=COR_STATUS_LABEL_FG, font=FONTE_STATUS, anchor=tk.CENTER)
status_label.pack(pady=5, fill=tk.X)

# Área de Log
log_labelframe = ttk.LabelFrame(root, text=" Log de Processamento ", padding=(10,5))
log_labelframe.pack(pady=(5,10), padx=10, fill=tk.BOTH, expand=True)
log_text_widget_instance = tk.Text(log_labelframe, height=15, width=90, wrap=tk.WORD, font=FONTE_LOG, bg=COR_FUNDO_LOG, fg=COR_TEXTO_LOG, relief=tk.SOLID, bd=1, state=tk.DISABLED)
log_scrollbar_y = ttk.Scrollbar(log_labelframe, orient="vertical", command=log_text_widget_instance.yview)
log_text_widget_instance.config(yscrollcommand=log_scrollbar_y.set)
log_text_widget_instance.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=5, padx=(0,5))
log_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

setup_gui_logging_refs(root, log_text_widget_instance)

# --- 5. CONFIGURAÇÃO DA API GOOGLE GEMINI ---
genai_config_ok = False
if not GOOGLE_API_KEY:
    log_to_gui("ERRO CRÍTICO: Variável de ambiente 'GOOGLE_API_KEY' não foi encontrada ou está vazia.", "CRITICAL")
elif GOOGLE_API_KEY == "SUA_CHAVE_DE_API_AQUI":
    log_to_gui("ERRO CRÍTICO: A GOOGLE_API_KEY no arquivo .env ainda é o valor placeholder 'SUA_CHAVE_DE_API_AQUI'. Por favor, configure-a com sua chave real.", "CRITICAL")
else:
    try:
        genai.configure(api_key=GOOGLE_API_KEY)
        log_to_gui("API Key do Google Gemini configurada com sucesso.", "INFO")
        genai_config_ok = True
    except Exception as e:
        log_to_gui(f"ERRO ao tentar configurar a API Key do Google Gemini: {e}. Verifique se a chave é válida.", "ERROR")
        genai_config_ok = False

# --- FUNÇÕES PARA CARREGAR E SALVAR MAPEAMENTO DE CHAVES ---
def carregar_mapeamento_de_arquivo(caminho_arquivo_str: str) -> Optional[Dict[str, Any]]:
    """Carrega um arquivo de mapeamento JSON."""
    caminho_arquivo_abs = resource_path(caminho_arquivo_str)
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if not caminho_arquivo_abs.is_file():
        log_to_gui(f"Arquivo de mapeamento '{caminho_arquivo_abs.name}' não encontrado em '{caminho_arquivo_abs}'.", "WARNING")
        return None
    try:
        with open(caminho_arquivo_abs, "r", encoding="utf-8") as f:
            mapeamento = json.load(f)
        log_to_gui(f"Arquivo de mapeamento '{caminho_arquivo_abs.name}' carregado com sucesso.", "INFO")
        if not isinstance(mapeamento, dict):
            log_to_gui(f"ERRO: Conteúdo do arquivo de mapeamento '{caminho_arquivo_abs.name}' não é um dicionário JSON válido.", "ERROR")
            if parent_dialog: messagebox.showerror("Erro de Mapeamento", f"O arquivo '{caminho_arquivo_abs.name}' não contém um dicionário JSON válido.", parent=parent_dialog)
            return None
        return mapeamento
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO ao decodificar JSON no arquivo de mapeamento '{caminho_arquivo_abs.name}': {e.msg} na linha {e.lineno} col {e.colno}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro de Mapeamento", f"Erro de JSON no arquivo '{caminho_arquivo_abs.name}': {e.msg}", parent=parent_dialog)
        return None
    except Exception as e:
        log_to_gui(f"ERRO inesperado ao carregar o arquivo de mapeamento '{caminho_arquivo_abs.name}': {e}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro de Mapeamento", f"Erro inesperado ao carregar '{caminho_arquivo_abs.name}': {e}", parent=parent_dialog)
        return None

def salvar_mapeamento_em_arquivo(mapeamento: Dict[str, Any], caminho_arquivo_str: str) -> bool:
    """Salva um dicionário de mapeamento em um arquivo JSON."""
    caminho_arquivo_abs = resource_path(caminho_arquivo_str)
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    try:
        caminho_arquivo_abs.parent.mkdir(parents=True, exist_ok=True)
        with open(caminho_arquivo_abs, "w", encoding="utf-8") as f:
            json.dump(mapeamento, f, indent=2, ensure_ascii=False)
        log_to_gui(f"Arquivo de mapeamento salvo com sucesso em '{caminho_arquivo_abs.name}'.", "INFO")
        return True
    except Exception as e:
        log_to_gui(f"ERRO ao salvar o arquivo de mapeamento '{caminho_arquivo_abs.name}': {e}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro ao Salvar Mapeamento", f"Não foi possível salvar o arquivo de mapeamento '{caminho_arquivo_abs.name}': {e}", parent=parent_dialog)
        return False

# --- 6. FUNÇÕES AUXILIARES DE PROCESSAMENTO ---
def extrair_texto_do_pdf(caminho_pdf: Path) -> Optional[str]:
    """Extrai texto de todas as páginas de um arquivo PDF."""
    texto_completo = ""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            num_paginas = len(pdf.pages)
            log_to_gui(f"Iniciando extração de texto do PDF '{caminho_pdf.name}' ({num_paginas} páginas)...", "INFO")
            for i, pagina in enumerate(pdf.pages):
                if is_gui_widget_available(parent_dialog) and \
                   is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and \
                   isinstance(parent_dialog, tk.Tk):
                    status_label.config(text=f"Extraindo texto da página {i+1}/{num_paginas}...")
                    parent_dialog.update_idletasks()

                texto_pagina_stream = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=False)
                texto_pagina_layout = None

                if not texto_pagina_stream or len(texto_pagina_stream.split()) < 5:
                    texto_pagina_layout = pagina.extract_text(x_tolerance=1, y_tolerance=3, layout=True)

                texto_pagina_final = texto_pagina_stream
                if texto_pagina_layout and (not texto_pagina_stream or len(texto_pagina_layout.split()) > len(texto_pagina_stream.split())):
                    log_to_gui(f"Página {i+1}: Usando extração com layout=True pois produziu mais texto ou stream falhou.", "DEBUG")
                    texto_pagina_final = texto_pagina_layout

                if texto_pagina_final:
                    texto_completo += texto_pagina_final if texto_pagina_final.endswith("\n") else texto_pagina_final + "\n"
                else:
                    log_to_gui(f"Página {i+1} do PDF '{caminho_pdf.name}' não retornou texto (nem com layout=False, nem com layout=True).", "DEBUG")

            log_to_gui("Extração de texto do PDF concluída.", "INFO")
            texto_limpo_para_ia = texto_completo
            texto_limpo_para_ia = re.sub(r"^\s*Súmula de Crédito\s*$", "", texto_limpo_para_ia, flags=re.MULTILINE | re.IGNORECASE)
            texto_limpo_para_ia = re.sub(r"^\s*SICOOB\s*\n(?!Data:)", "", texto_limpo_para_ia, count=1, flags=re.IGNORECASE | re.MULTILINE)
            texto_limpo_para_ia = re.sub(r'Página:\s*\d+\s*\/\s*\d+\s*$', '', texto_limpo_para_ia, flags=re.MULTILINE | re.IGNORECASE)
            texto_limpo_para_ia = texto_limpo_para_ia.strip()
            return texto_limpo_para_ia
    except pdfplumber.exceptions.PDFSyntaxError as e_syntax:
        log_to_gui(f"Erro de sintaxe no arquivo PDF '{caminho_pdf.name}': {e_syntax}. O arquivo pode estar corrompido ou mal formatado.", "ERROR")
        if parent_dialog: messagebox.showerror("Erro de PDF", f"Erro de sintaxe no PDF '{caminho_pdf.name}':\n{e_syntax}\nO arquivo pode estar corrompido.", parent=parent_dialog)
        return None
    except Exception as e:
        log_to_gui(f"Erro inesperado ao extrair texto do PDF '{caminho_pdf.name}': {e}", "ERROR")
        logging.error(f"Erro detalhado ao extrair PDF {caminho_pdf.name}:", exc_info=True)
        if parent_dialog: messagebox.showerror("Erro na Extração de Texto do PDF", f"Ocorreu um erro inesperado ao tentar extrair o texto do PDF '{caminho_pdf.name}':\n{e}", parent=parent_dialog)
        return None

def preencher_excel_novo_com_placeholders(
    caminho_json_dados: Path,
    caminho_excel_modelo: Path,
    caminho_excel_saida: Path,
    nome_planilha_alvo: Optional[str] = None
) -> bool:
    """Preenche um arquivo Excel modelo com dados de um JSON, substituindo placeholders."""
    log_to_gui(f"Iniciando preenchimento do Excel: Modelo='{caminho_excel_modelo.name}', Dados='{caminho_json_dados.name}', Saída='{caminho_excel_saida.name}'", "INFO")
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if is_gui_widget_available(parent_dialog): iniciar_progresso()

    try:
        with open(caminho_json_dados, 'r', encoding='utf-8') as f:
            dados_para_preencher: Dict[str, Any] = json.load(f)

        if not dados_para_preencher:
            log_to_gui("AVISO: Arquivo JSON de dados para preenchimento do Excel está vazio. O arquivo Excel de saída será uma cópia do modelo.", "WARNING")
            if parent_dialog:
                messagebox.showwarning("Dados Vazios para Excel", "O arquivo JSON de dados está vazio. O Excel de saída será uma cópia do modelo, sem preenchimentos.", parent=parent_dialog)
            workbook = openpyxl.load_workbook(caminho_excel_modelo)
            workbook.save(caminho_excel_saida)
            if is_gui_widget_available(parent_dialog): parar_progresso(f"Excel copiado (JSON de dados vazio): {caminho_excel_saida.name}")
            return True

        workbook = openpyxl.load_workbook(caminho_excel_modelo)
        sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

        if nome_planilha_alvo:
            if nome_planilha_alvo in workbook.sheetnames:
                sheet = workbook[nome_planilha_alvo]
            else:
                log_to_gui(f"AVISO: Aba '{nome_planilha_alvo}' não encontrada no arquivo Excel modelo. Tentando usar a aba ativa: '{workbook.active.title if workbook.active else 'NENHUMA ABA ATIVA'}'.", "WARNING")
                if parent_dialog:
                     messagebox.showwarning("Aba Não Encontrada", f"A aba especificada '{nome_planilha_alvo}' não foi encontrada no modelo Excel. Usando a aba ativa.", parent=parent_dialog)
                sheet = workbook.active
        else:
            sheet = workbook.active

        if sheet is None:
            log_to_gui("ERRO CRÍTICO: Nenhuma planilha (aba) pôde ser selecionada no arquivo Excel modelo. Verifique se o arquivo contém abas ou se a aba ativa é válida.", "ERROR")
            if parent_dialog:
                messagebox.showerror("Erro de Planilha no Excel", "Nenhuma planilha ativa ou válida foi encontrada no arquivo Excel modelo.", parent=parent_dialog)
                parar_progresso("Erro no Excel: Planilha não encontrada")
            return False

        log_to_gui(f"Iniciando varredura da planilha '{sheet.title}' para substituição de placeholders...", "INFO")
        substituicoes_feitas = 0
        placeholder_regex = re.compile(r'\{\{([A-Z0-9_]+?)\}\}')

        for r_idx, row in enumerate(sheet.iter_rows()):
            if is_gui_widget_available(parent_dialog) and \
               is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and \
               isinstance(parent_dialog, tk.Tk) and \
               r_idx > 0 and r_idx % 50 == 0:
                status_label.config(text=f"Processando linha {r_idx+1} do Excel...")
                parent_dialog.update_idletasks()

            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_cell_value = str(cell.value)
                    current_cell_string = original_cell_value
                    modified_in_loop = False
                    matches_in_cell = list(placeholder_regex.finditer(original_cell_value))

                    if not matches_in_cell: continue

                    is_single_full_match = len(matches_in_cell) == 1 and matches_in_cell[0].group(0) == original_cell_value

                    if is_single_full_match:
                        placeholder_com_chaves = matches_in_cell[0].group(0)
                        if placeholder_com_chaves in dados_para_preencher:
                            cell.value = dados_para_preencher[placeholder_com_chaves]
                            substituicoes_feitas += 1
                        elif INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES:
                            cell.value = "N/A"
                            substituicoes_feitas += 1
                    else:
                        for match_obj in reversed(matches_in_cell):
                            placeholder_com_chaves = match_obj.group(0)
                            span_inicio, span_fim = match_obj.span(0)

                            if placeholder_com_chaves in dados_para_preencher:
                                valor_substituto = dados_para_preencher[placeholder_com_chaves]
                                valor_substituto_str = str(valor_substituto) if valor_substituto is not None else ""
                                current_cell_string = current_cell_string[:span_inicio] + valor_substituto_str + current_cell_string[span_fim:]
                                modified_in_loop = True; substituicoes_feitas += 1
                            elif INSERIR_NA_PARA_PLACEHOLDERS_AUSENTES:
                                current_cell_string = current_cell_string[:span_inicio] + "N/A" + current_cell_string[span_fim:]
                                modified_in_loop = True; substituicoes_feitas += 1
                        if modified_in_loop: cell.value = current_cell_string

        if substituicoes_feitas == 0:
            log_to_gui(f"AVISO: Nenhuma substituição de placeholder foi realizada na planilha '{sheet.title}'. Verifique se os placeholders no Excel (formato {{CHAVE}}) correspondem às chaves no JSON de dados.", "WARNING")
            if parent_dialog: messagebox.showwarning("Nenhuma Substituição no Excel", "Nenhum placeholder foi substituído na planilha. Verifique o modelo Excel e os dados JSON gerados.", parent=parent_dialog)
        else:
            log_to_gui(f"INFO: {substituicoes_feitas} substituições de placeholders realizadas com sucesso na planilha '{sheet.title}'.", "INFO")

        workbook.save(caminho_excel_saida)
        log_to_gui(f"Novo arquivo Excel preenchido salvo em '{caminho_excel_saida.name}'.", "INFO")
        if is_gui_widget_available(parent_dialog): parar_progresso(f"Excel Gerado: {caminho_excel_saida.name}")
        return True

    except FileNotFoundError as e:
        log_to_gui(f"ERRO: Arquivo não encontrado ao tentar preencher o Excel: {e.filename}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro de Arquivo no Excel", f"Arquivo não encontrado durante o processo do Excel: {e.filename}", parent=parent_dialog)
    except json.JSONDecodeError as e:
        log_to_gui(f"ERRO: JSON inválido no arquivo de dados '{caminho_json_dados.name}' usado para preencher o Excel: {e.msg} L{e.lineno}C{e.colno}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro de JSON para Excel", f"Erro de JSON no arquivo de dados para Excel ('{caminho_json_dados.name}'): {e.msg}", parent=parent_dialog)
    except Exception as e:
        log_to_gui(f"ERRO geral e inesperado ao preencher o arquivo Excel: {e}", "ERROR")
        logging.error("Erro geral ao preencher Excel:", exc_info=True)
        if parent_dialog: messagebox.showerror("Erro Inesperado no Excel", f"Ocorreu um erro inesperado durante o preenchimento do Excel: {e}", parent=parent_dialog)

    if is_gui_widget_available(parent_dialog): parar_progresso("Erro no preenchimento do Excel")
    return False

def salvar_texto_em_arquivo(texto: str, nome_arquivo_path: Path) -> None:
    """Salva uma string de texto em um arquivo."""
    try:
        nome_arquivo_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arquivo_path, "w", encoding="utf-8") as f:
            f.write(texto)
        log_to_gui(f"Arquivo de texto '{nome_arquivo_path.name}' salvo com sucesso.", "INFO")
    except Exception as e:
        log_to_gui(f"Erro ao salvar arquivo de texto '{nome_arquivo_path.name}': {e}", "ERROR")

def salvar_json_em_arquivo(dados: Dict[str, Any], nome_arquivo_path: Path) -> None:
    """Salva um dicionário Python como um arquivo JSON formatado."""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    try:
        nome_arquivo_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arquivo_path, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
        log_to_gui(f"Arquivo JSON '{nome_arquivo_path.name}' salvo com sucesso.", "INFO")
    except Exception as e:
        log_to_gui(f"Erro ao salvar arquivo JSON '{nome_arquivo_path.name}': {e}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro ao Salvar JSON", f"Erro ao tentar salvar o arquivo JSON '{nome_arquivo_path.name}': {e}", parent=parent_dialog)

@retry(wait=wait_exponential(multiplier=1, min=2, max=30), stop=stop_after_attempt(3), reraise=True)
def gerar_conteudo_gemini_com_retry(model: genai.GenerativeModel, prompt_usuario: str, generation_config: genai.types.GenerationConfig) -> str:
    """Envia um prompt para a API Gemini com política de retry e tratamento de feedback."""
    log_to_gui("Enviando requisição para API Gemini (com retry)...", "DEBUG")
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if is_gui_widget_available(parent_dialog) and isinstance(parent_dialog, tk.Tk):
        parent_dialog.update_idletasks()

    response = model.generate_content(prompt_usuario, generation_config=generation_config)
    log_to_gui("Resposta recebida da API Gemini.", "DEBUG")

    if hasattr(response, 'prompt_feedback') and response.prompt_feedback:
        log_to_gui(f"API Gemini - Prompt Feedback: Block Reason: {response.prompt_feedback.block_reason}, Safety Ratings: {response.prompt_feedback.safety_ratings}", "DEBUG")

    if not response.parts:
        log_to_gui("AVISO: Resposta da API Gemini não contém 'parts'. Verificando feedback do prompt para bloqueio...", "WARNING")
        if hasattr(response, 'prompt_feedback') and response.prompt_feedback.block_reason:
            block_reason_value = response.prompt_feedback.block_reason
            block_reason_str = block_reason_value.name if hasattr(block_reason_value, 'name') else str(block_reason_value)

            block_reason_msg = f"Prompt foi bloqueado pela API Gemini. Razão do bloqueio: {block_reason_str}."
            safety_ratings_info = ""
            if hasattr(response.prompt_feedback, 'safety_ratings') and response.prompt_feedback.safety_ratings:
                ratings_str_list = [
                    f"  - Categoria: {getattr(getattr(r, 'category', None), 'name', 'N/A')}, Probabilidade: {getattr(getattr(r, 'probability', None), 'name', 'N/A')}"
                    for r in response.prompt_feedback.safety_ratings
                    if getattr(getattr(r, 'category', None), 'name', 'N/A') != 'N/A'
                ]
                if ratings_str_list:
                    safety_ratings_info = "\n\nDetalhes das Avaliações de Segurança do Prompt:\n" + "\n".join(ratings_str_list)

            log_to_gui(f"ERRO API: {block_reason_msg}{safety_ratings_info}", "ERROR")

            if parent_dialog:
                messagebox.showwarning(
                    "Bloqueio de Conteúdo pela API Gemini",
                    f"A solicitação para a API Gemini foi bloqueada devido a políticas de segurança.\n\n"
                    f"Razão do bloqueio: {block_reason_str}"
                    f"{safety_ratings_info}\n\n"
                    f"O processamento desta parte do PDF pode ter falhado. Verifique os logs para mais detalhes e ajuste o conteúdo ou as configurações se necessário.",
                    parent=parent_dialog
                )
            raise Exception(block_reason_msg)
        log_to_gui("AVISO: Resposta da API Gemini não contém 'parts' e não parece ser um bloqueio. Retornando string vazia.", "WARNING")
        return ""
    return response.text

def enviar_texto_completo_para_gemini_todos_blocos(
    texto_completo_do_pdf: str,
    blocos_config_map: Dict[str, Any],
    pdf_path_para_logs: Path
) -> Optional[Dict[str, Any]]:
    """Envia o texto completo do PDF e o schema de extração para a API Gemini em uma única chamada."""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if not texto_completo_do_pdf:
        log_to_gui("Nenhum texto foi extraído do PDF para enviar à API. Abortando.", "WARNING"); return None
    if not genai_config_ok:
        log_to_gui("Configuração da API Key do Google Gemini falhou ou não foi realizada. Abortando chamada à API.", "ERROR")
        if parent_dialog:
            messagebox.showerror("Erro de API", "A API Key do Google Gemini não está configurada corretamente. Verifique o arquivo .env e os logs.", parent=parent_dialog)
        return None

    resposta_texto_bruto_api = ""
    try:
        log_to_gui("Construindo prompt detalhado para a API Gemini com todos os blocos do schema...", "DEBUG")
        model = genai.GenerativeModel(GEMINI_MODEL_NAME)
        prompt_instrucoes_blocos = []

        for nome_bloco, config_bloco in blocos_config_map.items():
            json_chave = config_bloco["json_chave"]
            titulo_bloco_regex_hint = config_bloco.get("titulo_padrao", nome_bloco)
            campos_esperados = config_bloco.get("campos_esperados", [])
            nome_lista_json = config_bloco.get("nome_lista_json")
            sub_campos_lista = config_bloco.get("sub_campos_lista", [])
            sub_lista_aninhada_config = config_bloco.get("sub_lista_aninhada")
            campos_texto_longo_limitar = config_bloco.get("campos_texto_longo_limitar", [])

            instrucao_especifica = (
                f"Para o bloco '{nome_bloco}' (geralmente identificado por títulos como '{titulo_bloco_regex_hint}'), "
                f"mapeie todas as informações extraídas para a chave JSON principal '{json_chave}':\n")

            if nome_bloco == "0 Informacoes do Documento":
                instrucao_especifica += (
                    "  - Para este bloco específico ('0 Informacoes do Documento'), extraia os seguintes dados geralmente encontrados no cabeçalho ou topo do documento:\n"
                    "    - O valor após 'Cooperativa:' para o campo 'cooperativa'.\n"
                    "    - O valor após 'PA:' (Posto de Atendimento) para o campo 'pa'.\n"
                    "    - O valor após 'Data:' (data de emissão do documento) para o campo 'data'.\n"
                    "    - O valor após 'Hora:' (hora de emissão do documento) para o campo 'hora'.\n"
                    "    - O valor após 'Data/hora referência:' para o campo 'data_ref_doc'.\n"
                    "    - Ignore termos isolados como 'SICOOB' ou 'Súmula de Crédito' se aparecerem como títulos soltos; foque nos valores rotulados.\n"
                )
                campos_ja_tratados_info_doc = ["cooperativa", "pa", "data", "hora", "data_ref_doc"]
                campos_esperados_restantes = [c for c in (campos_esperados or []) if c not in campos_ja_tratados_info_doc]
            else:
                campos_esperados_restantes = campos_esperados or []

            if campos_esperados_restantes:
                instrucao_especifica += f"  - Extraia os seguintes campos diretos (chave-valor): {', '.join(campos_esperados_restantes)}.\n"
                for campo_limitar in (campos_texto_longo_limitar or []):
                    if campo_limitar in campos_esperados_restantes:
                        instrucao_especifica += (
                            f"    - Para o campo '{campo_limitar}', limite o texto extraído aos primeiros {MAX_TEXT_LENGTH_IA} caracteres OU forneça um resumo conciso DENTRO desse limite. "
                            f"Se o texto original for mais longo, TRUNQUE-O. Garanta JSON válido. Se impossível serializar ou se o resumo exceder o limite, retorne 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL'. NÃO exceda {MAX_TEXT_LENGTH_IA} caracteres.\n"
                        )

            if nome_lista_json and (sub_campos_lista or []):
                instrucao_especifica += (
                    f"  - Extraia uma LISTA DE OBJETOS JSON sob a chave '{nome_lista_json}'. Cada objeto na lista deve conter os campos: {', '.join(sub_campos_lista)}.\n")
                for campo_limitar in (campos_texto_longo_limitar or []):
                    if campo_limitar in (sub_campos_lista or []):
                         instrucao_especifica += (
                            f"    - Dentro de cada objeto da lista '{nome_lista_json}', para o campo '{campo_limitar}', limite o texto a {MAX_TEXT_LENGTH_IA} chars OU um resumo. "
                            f"Se mais longo, TRUNQUE. Use 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL' se necessário. NÃO exceda {MAX_TEXT_LENGTH_IA} chars.\n"
                        )
                if sub_lista_aninhada_config and isinstance(sub_lista_aninhada_config, dict):
                    nome_sub_lista = sub_lista_aninhada_config.get("nome_json")
                    campos_sub_lista_aninhada = sub_lista_aninhada_config.get("campos", [])
                    if nome_sub_lista and (campos_sub_lista_aninhada or []):
                        instrucao_especifica += (
                            f"    - Dentro de CADA objeto da lista '{nome_lista_json}', se aplicável, extraia uma SUB-LISTA DE OBJETOS JSON sob a chave '{nome_sub_lista}', com campos: {', '.join(campos_sub_lista_aninhada)}.\n")
                        for campo_limitar_sub in (campos_texto_longo_limitar or []):
                            if campo_limitar_sub in (campos_sub_lista_aninhada or []):
                                instrucao_especifica += (
                                    f"      - Em '{nome_sub_lista}', para o campo '{campo_limitar_sub}', limite a {MAX_TEXT_LENGTH_IA} chars OU resumo. "
                                    f"Se mais longo, TRUNQUE. Use 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL' se necessário. NÃO exceda {MAX_TEXT_LENGTH_IA} chars.\n"
                                )
                instrucao_especifica += f"  - Se não houver itens para a lista '{nome_lista_json}', retorne uma lista vazia ([]) para ela.\n"
            prompt_instrucoes_blocos.append(instrucao_especifica)

        prompt_usuario = (
            "Você é um especialista em análise de documentos Súmula de Crédito do SICOOB. "
            "Sua tarefa é analisar o texto do documento fornecido, que estará entre os marcadores "
            f"'{MARCADOR_INICIO_TEXTO_PDF_PROMPT}' e '{MARCADOR_FIM_TEXTO_PDF_PROMPT}'.\n"
            "Você deve EXTRAIR INFORMAÇÕES **EXCLUSIVAMENTE PARA OS BLOCOS E CAMPOS ESPECIFICADOS ABAIXO**. "
            "Ignore completamente quaisquer outros blocos ou seções do documento não listados nas instruções.\n"
            "A saída DEVE SER UM ÚNICO OBJETO JSON VÁLIDO.\n\n"
            "REGRAS IMPORTANTES PARA O JSON DE SAÍDA:\n"
            "1.  O JSON deve ser estritamente válido (chaves e strings com aspas duplas, escapes corretos como \\\" para aspas e \\n para novas linhas dentro de strings, etc.).\n"
            "2.  Se um valor para um campo solicitado não for encontrado no texto, omita a chave correspondente do objeto JSON, OU use `null` para campos simples e `[]` (lista vazia) para campos de lista, conforme especificado para cada bloco.\n"
            "2.1. NÃO INVENTE dados. Se a informação não estiver explicitamente no texto fornecido para um campo, siga a regra de omissão ou uso de null/[] (Regra 2).\n"
            "3.  Use EXATAMENTE as 'json_chave' fornecidas nas instruções para cada bloco como chaves de primeiro nível no objeto JSON de saída.\n"
            "4.  Extraia os valores o mais literalmente possível, mas limpe espaços extras desnecessários no início/fim.\n"
            "5.  Preserve o formato original de datas, números, códigos e CPFs/CNPJs.\n"
            f"6.  Para CAMPOS DE TEXTO LONGO (como pareceres, observações, etc.) que estão listados em 'campos_texto_longo_limitar': o texto extraído DEVE ser limitado a um máximo de {MAX_TEXT_LENGTH_IA} caracteres OU um resumo muito conciso que caiba nesse limite. Se o texto original for maior, ele DEVE ser TRUNCADO para respeitar o limite. Certifique-se de que a string resultante (truncada ou resumida) seja válida em JSON. Se, mesmo após a tentativa de resumo ou truncamento, o conteúdo for muito complexo para serializar corretamente (ex: caracteres de controle estranhos que não podem ser escapados) ou ainda exceder o limite de forma significativa, você DEVE retornar a string literal 'TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL' para esse campo específico. É CRUCIAL não exceder os {MAX_TEXT_LENGTH_IA} caracteres para esses campos.\n\n"
            f"INSTRUÇÕES ESPECÍFICAS PARA CADA BLOCO A SER EXTRAÍDO (IGNORE TODOS OS OUTROS BLOCOS DO DOCUMENTO):\n"
            + "\n\n".join(prompt_instrucoes_blocos) + "\n\n"
            f"TEXTO COMPLETO DO DOCUMENTO (lembre-se de focar apenas nos blocos e campos listados acima):\n{MARCADOR_INICIO_TEXTO_PDF_PROMPT}\n"
            f"{texto_completo_do_pdf}\n"
            f"{MARCADOR_FIM_TEXTO_PDF_PROMPT}\n\n"
            "REFORÇO CRÍTICO: Sua resposta DEVE ser um ÚNICO objeto JSON. O nível raiz deste objeto JSON DEVE conter APENAS as seguintes chaves de primeiro nível (correspondentes às 'json_chave' dos blocos solicitados): " +
            ", ".join([f'"{blocos_config_map[bn]["json_chave"]}"' for bn in blocos_config_map.keys() if bn in blocos_config_map]) + ". " +
            "Cada uma dessas chaves de primeiro nível conterá os dados extraídos para o respectivo bloco. Verifique DUAS VEZES a sintaxe JSON e os limites de caracteres para textos longos antes de finalizar a resposta."
        )

        log_to_gui(f"Enviando {len(texto_completo_do_pdf)} caracteres (aprox.) para API Gemini (Modelo: {GEMINI_MODEL_NAME})...", "INFO")
        if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
            prompt_debug_path = pdf_path_para_logs.parent / f"{pdf_path_para_logs.stem}_prompt_gemini_completo.txt"
            salvar_texto_em_arquivo(prompt_usuario, prompt_debug_path)

        generation_config = genai.types.GenerationConfig(
            temperature=GEMINI_TEMPERATURE,
            max_output_tokens=GEMINI_MAX_OUTPUT_TOKENS,
            response_mime_type="application/json"
        )

        resposta_texto_bruto_api = gerar_conteudo_gemini_com_retry(model, prompt_usuario, generation_config).strip()

        if resposta_texto_bruto_api.startswith("```json"): resposta_texto_bruto_api = resposta_texto_bruto_api[7:]
        if resposta_texto_bruto_api.endswith("```"): resposta_texto_bruto_api = resposta_texto_bruto_api[:-3]
        resposta_texto_bruto_api = resposta_texto_bruto_api.strip()

        if not (resposta_texto_bruto_api.startswith("{") and resposta_texto_bruto_api.endswith("}")):
            log_to_gui(f"AVISO: Resposta da API Gemini não parece ser um objeto JSON completo (não começa com '{{' e termina com '}}'). Início: '{resposta_texto_bruto_api[:100]}...', Fim: '...{resposta_texto_bruto_api[-100:]}'. Tentando decodificar...", "WARNING")

        dados_json_combinados: Dict[str, Any] = json.loads(resposta_texto_bruto_api)
        log_to_gui("JSON da API Gemini decodificado com sucesso.", "INFO")
        return dados_json_combinados

    except RetryError as e_retry:
        log_to_gui(f"ERRO FATAL API: Falha na conexão com Gemini após múltiplas tentativas: {e_retry}. Verifique sua conexão e as configurações da API.", "CRITICAL")
        if parent_dialog: messagebox.showerror("Erro de API", f"Falha na conexão com a API Gemini após várias tentativas: {e_retry}", parent=parent_dialog)
    except json.JSONDecodeError as e_json:
        log_to_gui(f"ERRO JSONDecodeError API: Não foi possível decodificar a resposta JSON da API Gemini. Erro: {e_json.msg} na posição {e_json.pos}. Resposta bruta: '{resposta_texto_bruto_api[:500]}...'", "ERROR")
        if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
            nome_arq_erro = pdf_path_para_logs.parent / f"gemini_resposta_erro_json_{pdf_path_para_logs.stem}.txt"
            salvar_texto_em_arquivo(resposta_texto_bruto_api or "Nenhuma resposta da API recebida.", nome_arq_erro)
            log_to_gui(f"Resposta bruta da API com erro JSON salva em '{nome_arq_erro.name}'.", "INFO")
        if parent_dialog: messagebox.showerror("Erro de API", f"A resposta da API Gemini não foi um JSON válido: {e_json.msg}", parent=parent_dialog)
    except Exception as e_api_general:
        log_to_gui(f"ERRO GERAL DURANTE CHAMADA À API GEMINI: {e_api_general}", "ERROR")
        logging.exception("Erro geral durante chamada à API Gemini:")
        if parent_dialog: messagebox.showerror("Erro de API", f"Ocorreu um erro inesperado ao comunicar com a API Gemini: {e_api_general}", parent=parent_dialog)
    return None

def achatar_json(objeto_json: Union[Dict[str, Any], List[Any]], prefixo_pai: str = '', separador: str = '_') -> Dict[str, Any]:
    """Converte um JSON aninhado (dicionários e listas) em um dicionário achatado."""
    items_achatados: Dict[str, Any] = {}
    if isinstance(objeto_json, dict):
        for chave, valor in objeto_json.items():
            chave_str = str(chave)
            nova_chave_prefixada = f"{prefixo_pai}{separador}{chave_str}" if prefixo_pai else chave_str
            items_achatados.update(achatar_json(valor, nova_chave_prefixada, separador=separador))
    elif isinstance(objeto_json, list):
        if not objeto_json:
            pass
        elif not any(isinstance(item, (dict, list)) for item in objeto_json):
            try:
                items_achatados[prefixo_pai if prefixo_pai else "lista_simples_na_raiz"] = ', '.join(map(str, objeto_json))
            except TypeError:
                 items_achatados[prefixo_pai if prefixo_pai else "erro_lista_simples_na_raiz"] = str(objeto_json)
        else:
            for i, item_lista in enumerate(objeto_json):
                chave_item_lista_indexada = f"{prefixo_pai}{separador}{i+1}"
                items_achatados.update(achatar_json(item_lista, chave_item_lista_indexada, separador=separador))
    else:
        if prefixo_pai:
            items_achatados[prefixo_pai] = objeto_json
    return items_achatados

def gerenciar_chave_nao_mapeada_interativamente(
    chave_original_nao_mapeada: str,
    caminho_arquivo_config_str: str,
    mapa_atual_em_memoria: Dict[str, str]
) -> Tuple[str, bool, bool, bool]:
    """Gerencia interativamente uma chave da IA que não possui mapeamento definido."""
    log_to_gui(f"INFO: Chave da IA não mapeada encontrada: '{chave_original_nao_mapeada}'. Solicitando interação do usuário.", "INFO")
    caminho_arquivo_config_abs = resource_path(caminho_arquivo_config_str)
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None

    nomes_padronizados_existentes: List[str] = sorted(list(set(mapa_atual_em_memoria.values()))) if mapa_atual_em_memoria else []
    nome_original_formatado = chave_original_nao_mapeada.replace("_", " ").title()

    msg_prompt = (
        f"A chave extraída pela IA '{chave_original_nao_mapeada}' não possui um mapeamento definido em '{caminho_arquivo_config_abs.name}'.\n\n"
        "Escolha uma opção para esta chave:\n"
        f"1. Usar um Nome Padronizado Existente (você digitará o nome; será salvo no arquivo de mapeamento).\n"
        f"2. Criar um Novo Nome Padronizado (você digitará o nome; será salvo no arquivo de mapeamento).\n"
        f"3. Usar o nome sugerido '{nome_original_formatado}' (derivado da chave da IA; será salvo e usado automaticamente para próximas chaves não mapeadas *neste PDF*).\n"
        f"4. PULAR TODAS as interações de mapeamento para as chaves restantes *deste PDF* (serão usados os nomes originais da IA, sem salvar no arquivo de mapeamento).\n\n"
        "Digite o número da opção (1-4). Se cancelar (ESC), esta chave específica será pulada (usando o nome original da IA) e você será perguntado novamente na próxima chave não mapeada."
    )
    if nomes_padronizados_existentes:
        msg_prompt += "\n\nExemplos de Nomes Padronizados Existentes (para opção 1):\n - " + "\n - ".join(nomes_padronizados_existentes[:min(5, len(nomes_padronizados_existentes))])
        if len(nomes_padronizados_existentes) > 5: msg_prompt += "\n - ..."

    escolha_opcao_str: Optional[str] = simpledialog.askstring("Mapeamento de Nova Chave da IA", msg_prompt, parent=parent_dialog) if parent_dialog else None

    nome_para_salvar_no_mapa: Optional[str] = None
    nome_padronizado_escolhido: str = chave_original_nao_mapeada
    mapa_foi_atualizado_no_arquivo: bool = False
    pular_todas_proximas_neste_pdf: bool = False
    mapear_todas_auto_neste_pdf: bool = False

    if escolha_opcao_str is None:
        log_to_gui(f"MAPEAMENTO PULADO (esta ocorrência): Para a chave '{chave_original_nao_mapeada}'. Usando o nome original da IA para esta vez.", "INFO")
        return nome_padronizado_escolhido, False, False, False

    escolha_opcao_str = escolha_opcao_str.strip()

    if escolha_opcao_str == "1":
        nome_existente_input: Optional[str] = simpledialog.askstring("Usar Nome Existente", "Digite o Nome Padronizado Existente que deseja usar para esta chave:", parent=parent_dialog) if parent_dialog else None
        if nome_existente_input and nome_existente_input.strip():
            nome_para_salvar_no_mapa = nome_existente_input.strip()
            nome_padronizado_escolhido = nome_para_salvar_no_mapa
        else: log_to_gui(f"Nenhum nome existente fornecido para '{chave_original_nao_mapeada}'. Usando original da IA para esta vez.", "INFO")
    elif escolha_opcao_str == "2":
        novo_nome_input: Optional[str] = simpledialog.askstring("Criar Novo Nome Padronizado", f"Digite o novo nome padronizado para a chave IA '{chave_original_nao_mapeada}':", parent=parent_dialog) if parent_dialog else None
        if novo_nome_input and novo_nome_input.strip():
            nome_para_salvar_no_mapa = novo_nome_input.strip()
            nome_padronizado_escolhido = nome_para_salvar_no_mapa
        else: log_to_gui(f"Nenhum novo nome fornecido para '{chave_original_nao_mapeada}'. Usando original da IA para esta vez.", "INFO")
    elif escolha_opcao_str == "3":
        nome_para_salvar_no_mapa = nome_original_formatado
        nome_padronizado_escolhido = nome_para_salvar_no_mapa
        mapear_todas_auto_neste_pdf = True
        log_to_gui(f"Opção 3: Mapeando '{chave_original_nao_mapeada}' para '{nome_para_salvar_no_mapa}'. Próximas não mapeadas neste PDF usarão nomes formatados automaticamente e serão salvas.", "INFO")
    elif escolha_opcao_str == "4":
        pular_todas_proximas_neste_pdf = True
        log_to_gui(f"Opção 4: PULAR TODAS as interações de mapeamento para chaves restantes *deste PDF*. Nomes originais da IA serão usados e não salvos.", "INFO")
    else:
        log_to_gui(f"Opção de mapeamento inválida ('{escolha_opcao_str}') para '{chave_original_nao_mapeada}'. Usando nome original da IA para esta vez.", "INFO")
        return chave_original_nao_mapeada, False, False, False

    if nome_para_salvar_no_mapa and not pular_todas_proximas_neste_pdf:
        config_completo_map = carregar_mapeamento_de_arquivo(caminho_arquivo_config_str) or {"mapeamento_para_chaves_padronizadas": {}}
        if "mapeamento_para_chaves_padronizadas" not in config_completo_map or \
           not isinstance(config_completo_map.get("mapeamento_para_chaves_padronizadas"), dict):
            config_completo_map["mapeamento_para_chaves_padronizadas"] = {}
            log_to_gui(f"AVISO: Estrutura 'mapeamento_para_chaves_padronizadas' recriada no arquivo '{caminho_arquivo_config_abs.name}'.", "WARNING")

        current_mapping_on_file = config_completo_map["mapeamento_para_chaves_padronizadas"]

        if current_mapping_on_file.get(chave_original_nao_mapeada) != nome_para_salvar_no_mapa:
            current_mapping_on_file[chave_original_nao_mapeada] = nome_para_salvar_no_mapa
            if salvar_mapeamento_em_arquivo(config_completo_map, caminho_arquivo_config_str):
                mapa_foi_atualizado_no_arquivo = True
                mapa_atual_em_memoria[chave_original_nao_mapeada] = nome_para_salvar_no_mapa
                log_to_gui(f"MAPEAMENTO SALVO: Chave IA '{chave_original_nao_mapeada}' mapeada para '{nome_para_salvar_no_mapa}' e salva em '{caminho_arquivo_config_abs.name}'.", "INFO")
            else:
                log_to_gui(f"ERRO: Falha ao salvar o novo mapeamento para '{chave_original_nao_mapeada}' no arquivo '{caminho_arquivo_config_abs.name}'.", "ERROR")
        else:
            mapa_foi_atualizado_no_arquivo = True
            log_to_gui(f"Mapeamento '{chave_original_nao_mapeada}' -> '{nome_para_salvar_no_mapa}' já existe no arquivo de configuração com o mesmo valor.", "DEBUG")

    return nome_padronizado_escolhido, mapa_foi_atualizado_no_arquivo, pular_todas_proximas_neste_pdf, mapear_todas_auto_neste_pdf

def normalizar_chaves_json(
    json_achatado_da_ia: Dict[str, Any],
    mapeamento_chaves_padronizadas: Optional[Dict[str, str]],
    pular_mapeamento_interativo_pdf_inicial: bool
) -> Tuple[Dict[str, Any], Dict[str, str], bool]:
    """Normaliza chaves de um JSON achatado usando um mapa e interage com o usuário para novas chaves."""
    if not json_achatado_da_ia:
        log_to_gui("Normalizador: JSON achatado da IA está vazio. Nenhuma chave para normalizar.", "WARNING")
        return {}, dict(mapeamento_chaves_padronizadas or {}), pular_mapeamento_interativo_pdf_inicial

    mapa_em_memoria_atual = dict(mapeamento_chaves_padronizadas or {})
    json_final_com_placeholders: Dict[str, Any] = {}

    _pular_interacao_nesta_sessao = pular_mapeamento_interativo_pdf_inicial
    _mapear_automatico_nesta_sessao = False

    mapa_regex_wildcard: Dict[str, re.Pattern] = {}
    if mapa_em_memoria_atual:
        chaves_wildcard_ordenadas = sorted(
            [k for k in mapa_em_memoria_atual.keys() if '*' in k],
            key=lambda k: (k.count('_'), len(k)), reverse=True
        )
        for chave_wc_mapa in chaves_wildcard_ordenadas:
            regex_str = '^' + re.escape(chave_wc_mapa).replace(r'\*', r'([0-9]+)') + '$'
            try: mapa_regex_wildcard[chave_wc_mapa] = re.compile(regex_str)
            except re.error as e_re:
                log_to_gui(f"ERRO ao compilar regex para wildcard '{chave_wc_mapa}': {e_re}", "ERROR")

    for chave_ia, valor in json_achatado_da_ia.items():
        nome_padronizado_final: Optional[str] = None

        if _pular_interacao_nesta_sessao:
            nome_padronizado_final = chave_ia
        elif chave_ia in mapa_em_memoria_atual:
            nome_padronizado_final = str(mapa_em_memoria_atual[chave_ia])
        else:
            for chave_wc_mapa_original, regex_compilado in mapa_regex_wildcard.items():
                match = regex_compilado.match(chave_ia)
                if match:
                    nome_base_wc_padronizado = str(mapa_em_memoria_atual[chave_wc_mapa_original])
                    try:
                        if '*' in nome_base_wc_padronizado and match.groups():
                            nome_padronizado_final = nome_base_wc_padronizado.replace("*", match.group(1))
                        else: nome_padronizado_final = nome_base_wc_padronizado
                    except IndexError: nome_padronizado_final = nome_base_wc_padronizado.replace("*", "ERRO_IDX_WC")
                    log_to_gui(f"MAPEAMENTO WILDCARD: Chave IA '{chave_ia}' -> '{nome_padronizado_final}' usando padrão '{chave_wc_mapa_original}'.", "DEBUG")
                    break

            if not nome_padronizado_final:
                if _mapear_automatico_nesta_sessao:
                    nome_padronizado_final = chave_ia.replace("_", " ").title()
                    mapa_em_memoria_atual[chave_ia] = nome_padronizado_final
                    config_mapa_atual = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG) or {"mapeamento_para_chaves_padronizadas": {}}
                    if "mapeamento_para_chaves_padronizadas" not in config_mapa_atual or not isinstance(config_mapa_atual.get("mapeamento_para_chaves_padronizadas"), dict):
                        config_mapa_atual["mapeamento_para_chaves_padronizadas"] = {}
                    config_mapa_atual["mapeamento_para_chaves_padronizadas"][chave_ia] = nome_padronizado_final
                    salvar_mapeamento_em_arquivo(config_mapa_atual, ARQUIVO_MAPEAMENTO_CONFIG)
                    log_to_gui(f"MAPEAMENTO AUTOMÁTICO (Opção 3): Chave IA '{chave_ia}' -> '{nome_padronizado_final}' (salvo).", "INFO")
                elif not _pular_interacao_nesta_sessao:
                    nome_escolhido_usr, _, _pular_agora_usr, _mapear_auto_agora_usr = \
                        gerenciar_chave_nao_mapeada_interativamente(
                            chave_ia, ARQUIVO_MAPEAMENTO_CONFIG, mapa_em_memoria_atual)
                    if _pular_agora_usr: _pular_interacao_nesta_sessao = True
                    if _mapear_auto_agora_usr: _mapear_automatico_nesta_sessao = True
                    nome_padronizado_final = nome_escolhido_usr

        if nome_padronizado_final is None:
            nome_padronizado_final = chave_ia
            log_to_gui(f"DEBUG: Nome padronizado para '{chave_ia}' voltou ao original da IA após lógica de mapeamento.", "DEBUG")

        placeholder_excel = str(nome_padronizado_final).strip().upper()
        placeholder_excel = re.sub(r'[^A-Z0-9_]+', '_', placeholder_excel)
        while "__" in placeholder_excel: placeholder_excel = placeholder_excel.replace("__", "_")
        placeholder_excel = placeholder_excel.strip('_')

        if not placeholder_excel:
            placeholder_excel = f"CHAVE_INVALIDA_GERADA_{len(json_final_com_placeholders)}"
            log_to_gui(f"AVISO: Normalização de '{nome_padronizado_final}' (origem IA: '{chave_ia}') resultou em placeholder vazio. Usando '{placeholder_excel}'.", "WARNING")

        if isinstance(valor, str) and valor == "TEXTO_LONGO_COMPLEXO_VERIFICAR_ORIGINAL":
            log_to_gui(f"ALERTA DA IA: O campo '{chave_ia}' (mapeado para placeholder '{placeholder_excel}') contém o valor '{valor}'. RECOMENDA-SE verificar este campo no documento PDF original.", "WARNING")

        if placeholder_excel not in json_final_com_placeholders:
            json_final_com_placeholders[placeholder_excel] = valor
        else:
            ct_colisao = 1
            chave_colisao = f"{placeholder_excel}_DUPLICADO_{ct_colisao}"
            while chave_colisao in json_final_com_placeholders:
                ct_colisao += 1; chave_colisao = f"{placeholder_excel}_DUPLICADO_{ct_colisao}"
            json_final_com_placeholders[chave_colisao] = valor
            log_to_gui(f"AVISO: Conflito de placeholder para '{placeholder_excel}' (de '{nome_padronizado_final}'). Salvo como '{chave_colisao}'. Verifique o mapeamento.", "WARNING")

    return json_final_com_placeholders, mapa_em_memoria_atual, _pular_interacao_nesta_sessao

def gerar_mapeamento_sugestao(json_achatado_ia: Dict[str, Any], nome_arquivo_origem: str = "Desconhecido") -> Dict[str, str]:
    """Gera um JSON de sugestão de mapeamento com base nas chaves extraídas pela IA."""
    if not json_achatado_ia:
        log_to_gui("Mapeamento Sugestão: JSON da IA está vazio, nenhuma sugestão a gerar.", "WARNING"); return {}

    caminho_map_config_abs = resource_path(ARQUIVO_MAPEAMENTO_CONFIG)
    log_to_gui(f"\n--- SUGESTÃO DE MAPEAMENTO para o arquivo '{caminho_map_config_abs.name}' (Baseado no PDF: {nome_arquivo_origem}) ---", "INFO")
    log_to_gui("{\n  \"mapeamento_para_chaves_padronizadas\": {", "INFO")

    sugestoes: Dict[str, str] = {}

    def formatar_nome_sugerido(chave_ia: str) -> str:
        s = str(chave_ia).replace("_", " ").title()
        s = s.replace("Cpf Cnpj", "CPF/CNPJ").replace("Src", "SCR").replace("Id ", "ID ")
        return s.strip()

    list_item_pattern = re.compile(r"^(.*?)_(\d+)(_.*)?$")
    potential_wildcards: Dict[str, List[str]] = {}

    for chave_ia_original in json_achatado_ia.keys():
        match = list_item_pattern.match(chave_ia_original)
        if match:
            prefixo, _, sufixo_opcional_com_underscore = match.groups()
            sufixo = sufixo_opcional_com_underscore if sufixo_opcional_com_underscore else ""
            wildcard_pattern_base = f"{prefixo}_*{sufixo}"
            if wildcard_pattern_base not in potential_wildcards:
                potential_wildcards[wildcard_pattern_base] = []
            potential_wildcards[wildcard_pattern_base].append(chave_ia_original)

    for wc_pattern, ia_keys_list_match in potential_wildcards.items():
        if len(ia_keys_list_match) > 1:
            match_exemplo = list_item_pattern.match(ia_keys_list_match[0])
            if match_exemplo:
                prefixo_ex, _, sufixo_ex_op_underscore = match_exemplo.groups()
                sufixo_ex = sufixo_ex_op_underscore.lstrip('_') if sufixo_ex_op_underscore else ""
                nome_sug_wc = formatar_nome_sugerido(prefixo_ex)
                if sufixo_ex: nome_sug_wc += f" - {formatar_nome_sugerido(sufixo_ex)}"
                nome_sug_wc += " (Item *)"
                sugestoes[wc_pattern] = nome_sug_wc

    for chave_ia_original in json_achatado_ia.keys():
        coberta_por_wc = False
        for wc_sug_key in sugestoes.keys():
            if '*' in wc_sug_key:
                regex_wc_str = '^' + re.escape(wc_sug_key).replace(r'\*', r'[0-9]+') + '$'
                if re.fullmatch(regex_wc_str, chave_ia_original):
                    coberta_por_wc = True; break
        if not coberta_por_wc and chave_ia_original not in sugestoes:
            sugestoes[chave_ia_original] = formatar_nome_sugerido(chave_ia_original)

    chaves_sugestoes_ordenadas = sorted(sugestoes.keys(), key=lambda k: ('*' not in k, k.lower()))
    for i, chave_ia_sug in enumerate(chaves_sugestoes_ordenadas):
        nome_pad_sug = sugestoes[chave_ia_sug]
        log_to_gui(f'    "{chave_ia_sug}": "{nome_pad_sug}"{"," if i < len(chaves_sugestoes_ordenadas) - 1 else ""}', "INFO")

    log_to_gui("  }\n}", "INFO")
    log_to_gui(f"--- Copie o bloco acima e cole/edite no arquivo '{caminho_map_config_abs.name}' para refinar os mapeamentos. --- \n", "INFO")
    return sugestoes

def gerar_json_com_chaves_placeholder(json_dados_normalizados: Dict[str, Any], nome_arq_saida_path: Path) -> bool:
    """Gera um JSON onde as chaves são formatadas como placeholders para o Excel."""
    if not json_dados_normalizados:
        log_to_gui("JSON para Placeholders: Dados normalizados estão vazios. Nenhum arquivo JSON para Excel será gerado.", "WARNING"); return False
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None

    json_para_excel: Dict[str, Any] = { f"{{{{{key}}}}}" : val for key, val in json_dados_normalizados.items() }

    try:
        nome_arq_saida_path.parent.mkdir(parents=True, exist_ok=True)
        with open(nome_arq_saida_path, "w", encoding="utf-8") as f:
            json.dump(json_para_excel, f, indent=2, ensure_ascii=False)
        log_to_gui(f"JSON formatado com placeholders para Excel salvo em: '{nome_arq_saida_path.name}'", "INFO"); return True
    except Exception as e:
        log_to_gui(f"Erro ao salvar JSON formatado com placeholders para Excel ('{nome_arq_saida_path.name}'): {e}", "ERROR")
        if parent_dialog:
            messagebox.showerror("Erro ao Salvar JSON para Excel", f"Erro ao tentar salvar o JSON para Excel: {e}", parent=parent_dialog)
        return False

# --- FUNÇÃO PRINCIPAL DE PROCESSAMENTO ---
def processar_pdf_e_gerar_saidas(caminho_pdf_path_obj: Path) -> bool:
    """Orquestra todo o processo de análise de um PDF e geração de saídas."""
    log_to_gui(f"--- Iniciando processamento do PDF: {caminho_pdf_path_obj.name} ---", "INFO")
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if parent_dialog: iniciar_progresso()

    if not BLOCO_CONFIG:
        log_to_gui("ERRO CRÍTICO: Schema de extração (BLOCO_CONFIG) está vazio ou não foi carregado corretamente. Processamento interrompido.", "CRITICAL")
        if parent_dialog:
            parar_progresso("Erro de configuração do schema.")
            messagebox.showerror("Erro Configuração", "O schema de extração está vazio ou inválido. Verifique os logs e o arquivo de schema.", parent=parent_dialog)
        return False

    texto_completo_extraido = extrair_texto_do_pdf(caminho_pdf_path_obj)
    if texto_completo_extraido is None:
        if parent_dialog: parar_progresso(f"Erro ao extrair texto de {caminho_pdf_path_obj.name}")
        return False

    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
        salvar_texto_em_arquivo(texto_completo_extraido, caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_texto_completo_para_ia.txt")

    log_to_gui(f"Enviando todos os {len(BLOCO_CONFIG)} blocos do schema em uma única requisição para a API Gemini...", "INFO")
    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
        status_label.config(text=f"Processando PDF com IA ({len(BLOCO_CONFIG)} blocos)...")
        parent_dialog.update_idletasks()

    resultado_api_gemini = enviar_texto_completo_para_gemini_todos_blocos(
        texto_completo_do_pdf=texto_completo_extraido,
        blocos_config_map=BLOCO_CONFIG,
        pdf_path_para_logs=caminho_pdf_path_obj
    )

    if not resultado_api_gemini:
        log_to_gui("ERRO FATAL: Falha ao processar o PDF com a API Gemini ou a resposta foi inválida/vazia.", "CRITICAL")
        if parent_dialog: parar_progresso("Erro no processamento da API.")
        return False

    log_to_gui("API Gemini: Processamento do PDF concluído com sucesso.", "INFO")
    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS:
        salvar_json_em_arquivo(resultado_api_gemini, caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_bruto_combinado.json")

    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
        status_label.config(text="Achatando JSON retornado pela IA...")
        parent_dialog.update_idletasks()
    json_achatado_ia = achatar_json(resultado_api_gemini)

    if not json_achatado_ia and resultado_api_gemini:
        log_to_gui("ERRO: Falha ao achatar o JSON da API. O resultado do achatamento está vazio, embora a API tenha retornado dados.", "ERROR")
        if parent_dialog:
            parar_progresso(f"Erro: JSON achatado da IA resultou vazio para {caminho_pdf_path_obj.name}")
            messagebox.showerror("Erro Processamento", "O JSON da IA achatado resultou vazio. Verifique os logs.", parent=parent_dialog)
        return False
    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS and json_achatado_ia:
        salvar_json_em_arquivo(json_achatado_ia, caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_ia_achatado_debug.json")

    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
        status_label.config(text="Normalizando chaves do JSON...")
        parent_dialog.update_idletasks()

    _pular_interacao_para_este_pdf = False
    map_config = carregar_mapeamento_de_arquivo(ARQUIVO_MAPEAMENTO_CONFIG)
    map_chaves_atuais: Optional[Dict[str, str]] = None

    if map_config and "mapeamento_para_chaves_padronizadas" in map_config and \
       isinstance(map_config.get("mapeamento_para_chaves_padronizadas"), dict):
        map_chaves_atuais = map_config["mapeamento_para_chaves_padronizadas"]
        log_to_gui(f"Usando mapeamento de chaves do arquivo '{resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name}'. Chaves mapeadas: {len(map_chaves_atuais)}", "INFO")
    elif json_achatado_ia:
        nome_arq_map = resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name
        log_to_gui(f"Arquivo de mapeamento '{nome_arq_map}' não encontrado ou inválido. Gerando sugestões...", "WARNING")
        sugestoes_map = gerar_mapeamento_sugestao(json_achatado_ia, caminho_pdf_path_obj.name)
        map_chaves_atuais = sugestoes_map
        if salvar_mapeamento_em_arquivo({"mapeamento_para_chaves_padronizadas": sugestoes_map}, ARQUIVO_MAPEAMENTO_CONFIG):
            log_to_gui(f"Arquivo de mapeamento com sugestões salvo em '{nome_arq_map}'. Recomenda-se editá-lo.", "INFO")
            if parent_dialog:
                messagebox.showinfo("Mapeamento Criado", f"'{nome_arq_map}' foi criado com sugestões. Edite-o para refinar os nomes das chaves para o Excel.", parent=parent_dialog)
        else: log_to_gui(f"ERRO: Não foi possível salvar o mapeamento sugerido em '{nome_arq_map}'.", "ERROR")
    else:
        log_to_gui("INFO: Nenhum dado da IA para normalizar e nenhum arquivo de mapeamento. Chaves originais da IA (se houver) serão usadas.", "INFO")

    json_final_excel: Dict[str, Any] = {}
    if json_achatado_ia:
        json_final_excel, _, _pular_interacao_para_este_pdf = normalizar_chaves_json(
            json_achatado_ia, map_chaves_atuais, _pular_interacao_para_este_pdf)
        if not json_final_excel and json_achatado_ia:
            log_to_gui("ERRO: Falha ao normalizar as chaves do JSON. Resultado normalizado vazio.", "ERROR")
            if parent_dialog:
                parar_progresso("Erro na normalização do JSON.")
                messagebox.showerror("Erro Normalização", "A normalização das chaves do JSON falhou. Verifique os logs.", parent=parent_dialog)
            return False
    elif resultado_api_gemini:
        log_to_gui("AVISO: JSON achatado estava vazio, pulando normalização de chaves.", "WARNING")

    if CRIAR_ARQUIVOS_DEBUG_INTERMEDIARIOS and json_final_excel:
        salvar_json_em_arquivo(json_final_excel, caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_final_normalizado_debug.json")

    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
        status_label.config(text="Gerando JSON final para preenchimento do Excel...")
        parent_dialog.update_idletasks()

    arq_json_para_excel = caminho_pdf_path_obj.parent / f"{caminho_pdf_path_obj.stem}_dados_para_excel.json"
    if gerar_json_com_chaves_placeholder(json_final_excel, arq_json_para_excel):
        log_to_gui("JSON para Excel gerado. Por favor, selecione o ARQUIVO EXCEL MODELO.", "INFO")
        if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
            status_label.config(text="Aguardando seleção do Excel modelo...")
            parent_dialog.update_idletasks()

        caminho_template_excel_str: Optional[str] = filedialog.askopenfilename(
            parent=parent_dialog, title="Selecione o ARQUIVO EXCEL MODELO (.xlsx)",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))

        if caminho_template_excel_str:
            caminho_template_excel = Path(caminho_template_excel_str)
            nome_arq_saida_sugerido = f"{caminho_pdf_path_obj.stem}_PREENCHIDO.xlsx"
            caminho_excel_saida_str: Optional[str] = filedialog.asksaveasfilename(
                parent=parent_dialog, title="Salvar Excel Preenchido Como...",
                initialdir=str(caminho_pdf_path_obj.parent), initialfile=nome_arq_saida_sugerido,
                defaultextension=".xlsx", filetypes=(("Arquivos Excel", "*.xlsx"),))

            if caminho_excel_saida_str:
                caminho_excel_saida_final = Path(caminho_excel_saida_str)
                nome_aba_excel: Optional[str] = simpledialog.askstring(
                    "Nome da Aba no Excel",
                    "Digite o nome da ABA no arquivo Excel modelo que deve ser preenchida (deixe em branco para usar a aba ativa):",
                    parent=parent_dialog)
                nome_planilha_final = nome_aba_excel.strip() if nome_aba_excel and nome_aba_excel.strip() else None

                if not preencher_excel_novo_com_placeholders(arq_json_para_excel, caminho_template_excel, caminho_excel_saida_final, nome_planilha_final):
                    log_to_gui(f"Falha ao preencher o arquivo Excel '{caminho_excel_saida_final.name}'.", "ERROR")
            else:
                if parent_dialog: parar_progresso("Salvamento do arquivo Excel cancelado.")
                log_to_gui("Processo de salvamento do Excel preenchido cancelado.", "INFO")
        else:
            if parent_dialog: parar_progresso("Seleção do modelo Excel cancelada.")
            log_to_gui("Seleção do arquivo Excel modelo cancelada.", "INFO")
    else:
        if parent_dialog: parar_progresso(f"Falha gerar JSON para Excel do PDF {caminho_pdf_path_obj.name}")
        log_to_gui(f"ERRO: Não foi possível gerar o arquivo JSON com placeholders para o Excel ('{arq_json_para_excel.name}').", "ERROR")

    log_to_gui(f"--- Fim do processamento para: {caminho_pdf_path_obj.name} ---", "INFO")
    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label):
        if not status_label.cget("text").startswith(("Pronto", "Excel Gerado", "Erro", "cancelado")):
            parar_progresso(f"Processamento de {caminho_pdf_path_obj.name} finalizado.")
    return True

# --- FUNÇÕES DA INTERFACE E MAINLOOP ---
def iniciar_fluxo_analise_pdf() -> None:
    """Inicia o fluxo principal de análise de um PDF selecionado pelo usuário."""
    log_to_gui("\n--- Novo Ciclo de Análise de PDF Iniciado ---", "INFO")
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None

    if not genai_config_ok:
        log_to_gui("AVISO: API Key do Google Gemini não configurada/inválida. Extração de dados falhará.", "WARNING")
        if parent_dialog:
            messagebox.showerror("Erro API", "API Key do Google Gemini não configurada. Verifique .env e logs.", parent=parent_dialog)
        return

    if not BLOCO_CONFIG:
        log_to_gui("ERRO CRÍTICO: Schema de extração vazio ou inválido. Processamento não pode continuar.", "CRITICAL")
        if parent_dialog:
            messagebox.showerror("Erro Configuração", f"Schema '{resource_path(ARQUIVO_SCHEMA_EXTRACAO).name}' vazio ou inválido. Verifique os logs.", parent=parent_dialog)
            if is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label): status_label.config(text="Erro de configuração do schema.")
        return

    if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and isinstance(parent_dialog, tk.Tk):
        status_label.config(text="Aguardando seleção do arquivo PDF...")
        parent_dialog.update_idletasks()

    caminho_pdf_str: Optional[str] = filedialog.askopenfilename(
        parent=parent_dialog, title="Selecione o arquivo PDF de Súmula de Crédito",
        filetypes=(("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")))

    if not caminho_pdf_str:
        log_to_gui("Nenhum arquivo PDF foi selecionado.", "INFO")
        if parent_dialog and is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label):
             status_label.config(text="Nenhum PDF selecionado.")
        return

    caminho_pdf_selecionado = Path(caminho_pdf_str)
    try:
        sucesso_processamento = processar_pdf_e_gerar_saidas(caminho_pdf_selecionado)
        msg_final = f"Processamento do arquivo '{caminho_pdf_selecionado.name}' "
        msg_final += "concluído com sucesso." if sucesso_processamento else "encontrou falhas (verifique os logs)."
        log_to_gui(msg_final, "INFO" if sucesso_processamento else "ERROR")

        if parent_dialog:
            if sucesso_processamento: messagebox.showinfo("Análise Finalizada", msg_final, parent=parent_dialog)
            else: messagebox.showwarning("Análise com Falhas", msg_final, parent=parent_dialog)
            if is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label) and not status_label.cget("text").startswith("Excel Gerado"):
                status_label.config(text="Pronto.")

    except Exception as e_fluxo_principal:
        log_to_gui(f"ERRO INESPERADO no fluxo principal para '{caminho_pdf_selecionado.name}': {e_fluxo_principal}", "CRITICAL")
        logging.exception(f"Exceção não tratada no fluxo principal para {caminho_pdf_selecionado.name}:")
        if parent_dialog:
            parar_progresso(f"Erro Crítico ao Processar {caminho_pdf_selecionado.name}")
            messagebox.showerror("Erro Crítico Inesperado", f"Ocorreu um erro crítico inesperado ao processar '{caminho_pdf_selecionado.name}'.\nConsulte o arquivo de log: '{LOG_FILE_PATH.name}'.\nDetalhes: {e_fluxo_principal}", parent=parent_dialog)
            if is_gui_widget_available(status_label) and isinstance(status_label, ttk.Label):
                status_label.config(text="Erro crítico. Verifique os logs.")

def mostrar_sobre() -> None:
    """Mostra a janela 'Sobre' com informações da aplicação."""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if parent_dialog:
        messagebox.showinfo("Sobre o Processador de Súmulas",
                            f"Processador de Súmulas de Crédito v1.2 (Gemini)\n\n"
                            f"Desenvolvido por: Claudeir de Souza Alves\n"
                            f"Analista de TI\n\n"
                            f"Arquivos de Configuração Utilizados:\n"
                            f"  - Schema de Extração: {resource_path(ARQUIVO_SCHEMA_EXTRACAO).name}\n"
                            f"  - Mapeamento de Chaves: {resource_path(ARQUIVO_MAPEAMENTO_CONFIG).name}\n\n"
                            f"Arquivo de Log Principal: {LOG_FILE_PATH.resolve()}\n"
                            f"Modelo IA: {GEMINI_MODEL_NAME}",
                            parent=parent_dialog)

def sair_aplicacao() -> None:
    """Fecha a aplicação após confirmação do usuário."""
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None
    if parent_dialog and isinstance(parent_dialog, tk.Tk):
        if messagebox.askokcancel("Sair da Aplicação", "Tem certeza que deseja sair do processador de súmulas?", parent=parent_dialog):
            log_to_gui("Aplicação encerrando por solicitação do usuário...", "INFO")
            parent_dialog.quit()
            parent_dialog.destroy()
    else:
        sys.exit()

def abrir_arquivo_para_edicao(nome_arq_relativo: str, desc_arq: str) -> None:
    """Tenta abrir um arquivo no editor padrão do sistema."""
    caminho_abs_arq = resource_path(nome_arq_relativo)
    parent_dialog = _root_ref_for_log if is_gui_widget_available(_root_ref_for_log) else None

    if not caminho_abs_arq.is_file():
        log_to_gui(f"Arquivo de {desc_arq} '{caminho_abs_arq.name}' não encontrado em: '{caminho_abs_arq}'.", "WARNING")
        if parent_dialog:
            messagebox.showinfo("Arquivo Não Encontrado", f"O arquivo de {desc_arq} '{caminho_abs_arq.name}' não foi encontrado no local esperado:\n'{caminho_abs_arq}'.", parent=parent_dialog)
        return
    try:
        log_to_gui(f"Tentando abrir o arquivo de {desc_arq} '{caminho_abs_arq}' para edição...", "INFO")
        if sys.platform == "win32": os.startfile(str(caminho_abs_arq))
        elif sys.platform == "darwin": subprocess.run(["open", str(caminho_abs_arq)], check=True)
        else: subprocess.run(["xdg-open", str(caminho_abs_arq)], check=True)
    except FileNotFoundError:
        log_to_gui(f"ERRO: Comando para abrir arquivos não encontrado no sistema ('{sys.platform}'). Não foi possível abrir '{caminho_abs_arq}'.", "ERROR")
        if parent_dialog: messagebox.showerror("Erro ao Abrir Arquivo", f"Não foi possível encontrar o comando do sistema para abrir arquivos. Arquivo: {caminho_abs_arq.name}", parent=parent_dialog)
    except Exception as e:
        log_to_gui(f"ERRO ao tentar abrir o arquivo de {desc_arq} '{caminho_abs_arq}': {e}", "ERROR")
        if parent_dialog: messagebox.showerror("Erro ao Abrir Arquivo", f"Não foi possível abrir o arquivo '{caminho_abs_arq.name}':\n{e}", parent=parent_dialog)

def abrir_mapeamento_para_edicao(): abrir_arquivo_para_edicao(ARQUIVO_MAPEAMENTO_CONFIG, "mapeamento de chaves")
def abrir_schema_para_edicao(): abrir_arquivo_para_edicao(ARQUIVO_SCHEMA_EXTRACAO, "schema de extração da IA")

# --- 8. CONFIGURAÇÃO FINAL DA GUI (Botões, Menu, mainloop) ---
style.configure("Primary.TButton", font=FONTE_BOTAO_PRINCIPAL, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE), foreground=COR_BOTAO_PRIMARIO_FG, background=COR_BOTAO_PRIMARIO_BG)
style.configure("Secondary.TButton", font=FONTE_BOTAO_SECUNDARIO, padding=(PADDING_X_BOTAO_STYLE, PADDING_Y_BOTAO_STYLE), foreground=COR_BOTAO_SECUNDARIO_FG, background=COR_BOTAO_SECUNDARIO_BG)
style.map("Primary.TButton", background=[('active', COR_BOTAO_HOVER_PRIMARIO), ('pressed', COR_BOTAO_HOVER_PRIMARIO), ('!disabled', COR_BOTAO_PRIMARIO_BG)], foreground=[('!disabled', COR_BOTAO_PRIMARIO_FG)])
style.map("Secondary.TButton", background=[('active', COR_BOTAO_HOVER_SECUNDARIO), ('pressed', COR_BOTAO_HOVER_SECUNDARIO), ('!disabled', COR_BOTAO_SECUNDARIO_BG)], foreground=[('!disabled', COR_BOTAO_SECUNDARIO_FG)])

frame_botoes_principais.columnconfigure(0, weight=1)
frame_botoes_principais.columnconfigure(1, weight=1)
frame_botoes_principais.columnconfigure(2, weight=1)

botao_analisar = ttk.Button(frame_botoes_principais, text="Analisar PDF e Gerar Saídas", command=iniciar_fluxo_analise_pdf, style="Primary.TButton")
botao_analisar.grid(row=0, column=0, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")

botao_editar_mapeamento = ttk.Button(frame_botoes_principais, text="Editar Mapeamento Chaves", command=abrir_mapeamento_para_edicao, style="Secondary.TButton")
botao_editar_mapeamento.grid(row=0, column=1, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")

botao_editar_schema = ttk.Button(frame_botoes_principais, text="Editar Schema Extração IA", command=abrir_schema_para_edicao, style="Secondary.TButton")
botao_editar_schema.grid(row=0, column=2, padx=5, pady=10, ipadx=5, ipady=5, sticky="ew")

menu_bar = tk.Menu(root, font=FONTE_MENU)
menu_arquivo = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
icone_menu_pdf_tk: Optional[ImageTk.PhotoImage] = None
caminho_icone_menu_relativo = "assets/icone_pdf.ico"

try:
    caminho_icone_menu_abs = resource_path(caminho_icone_menu_relativo)
    if caminho_icone_menu_abs.is_file():
        img_pil_menu = Image.open(caminho_icone_menu_abs)
        img_pil_menu = img_pil_menu.resize((16, 16), Image.Resampling.LANCZOS)
        icone_menu_pdf_tk = ImageTk.PhotoImage(img_pil_menu)
        log_to_gui(f"Ícone de menu '{caminho_icone_menu_relativo}' carregado.", "DEBUG")
    else:
        log_to_gui(f"AVISO: Ícone de menu '{caminho_icone_menu_abs}' NÃO ENCONTRADO.", "WARNING")
except Exception as e_icon_menu:
    log_to_gui(f"ERRO ao carregar ícone de menu '{caminho_icone_menu_relativo}': {e_icon_menu}", "ERROR")
    logging.exception(f"Erro detalhado ao carregar ícone de menu '{caminho_icone_menu_relativo}':")

if icone_menu_pdf_tk:
    menu_arquivo.add_command(label="Analisar PDF e Gerar Saídas...", image=icone_menu_pdf_tk, compound="left", command=iniciar_fluxo_analise_pdf)
else:
    menu_arquivo.add_command(label="Analisar PDF e Gerar Saídas...", command=iniciar_fluxo_analise_pdf)

menu_arquivo.add_command(label=f"Editar Mapeamento ({Path(ARQUIVO_MAPEAMENTO_CONFIG).name})", command=abrir_mapeamento_para_edicao)
menu_arquivo.add_command(label=f"Editar Schema IA ({Path(ARQUIVO_SCHEMA_EXTRACAO).name})", command=abrir_schema_para_edicao)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="Limpar Log da Tela", command=limpar_log_gui)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="Sair", command=sair_aplicacao)
menu_bar.add_cascade(label="Arquivo", menu=menu_arquivo)
menu_ajuda = tk.Menu(menu_bar, tearoff=0, font=FONTE_MENU)
menu_ajuda.add_command(label="Sobre", command=mostrar_sobre)
menu_bar.add_cascade(label="Ajuda", menu=menu_ajuda)
root.config(menu=menu_bar)
root.protocol("WM_DELETE_WINDOW", sair_aplicacao)

if __name__ == "__main__":
    log_to_gui(f"Aplicação Processador de Súmulas iniciada. PID: {os.getpid()}", "INFO")
    log_to_gui(f"O arquivo de log principal está sendo salvo em: {LOG_FILE_PATH.resolve()}", "INFO")
    log_to_gui("Verifique o arquivo de log para mensagens de erro ou avisos da inicialização.", "INFO")

    schema_carregado_ok = carregar_schema_extracao()
    particoes_logicas_ok = False

    if schema_carregado_ok:
        if BLOCO_CONFIG:
            particoes_logicas_ok = gerar_particoes_dinamicamente()
            if not particoes_logicas_ok:
                 log_to_gui("ERRO CRÍTICO: Falha ao gerar as partições lógicas do schema. Verifique a chave 'particao' nos blocos.", "CRITICAL")
                 if is_gui_widget_available(_root_ref_for_log) and isinstance(_root_ref_for_log, tk.Tk):
                     messagebox.showerror("Erro de Partição do Schema", "Falha ao organizar os blocos do schema. Verifique logs e o arquivo de schema.", parent=_root_ref_for_log)
        else:
            particoes_logicas_ok = True
            log_to_gui("INFO: Schema de extração carregado, mas está vazio (sem blocos).", "INFO")

    configuracao_critica_falhou = False
    msg_erro_critico_str = ""
    if not schema_carregado_ok:
        configuracao_critica_falhou = True
        msg_erro_critico_str = "ERRO FATAL: Falha ao carregar o schema de extração. A aplicação não pode funcionar."
    elif BLOCO_CONFIG and not particoes_logicas_ok:
        configuracao_critica_falhou = True
        msg_erro_critico_str = "ERRO FATAL: Falha na organização lógica do schema. A aplicação pode não funcionar corretamente."
    
    if configuracao_critica_falhou:
        log_to_gui(msg_erro_critico_str, "CRITICAL")
        if is_gui_widget_available(_root_ref_for_log) and isinstance(_root_ref_for_log, tk.Tk):
            messagebox.showerror("Erro na Configuração Inicial", f"{msg_erro_critico_str} Verifique os logs e os arquivos de configuração.", parent=_root_ref_for_log)
            if is_gui_widget_available(botao_analisar) and isinstance(botao_analisar, ttk.Button):
                botao_analisar.config(state=tk.DISABLED)
                log_to_gui("Botão 'Analisar PDF' desabilitado devido a erro crítico na configuração.", "WARNING")

    log_to_gui("Interface Gráfica Pronta e Aguardando Ações.", "INFO")

    if not genai_config_ok:
        log_to_gui("AVISO IMPORTANTE: API Key do Google Gemini não configurada ou inválida. A extração de dados dos PDFs FALHARÁ.", "CRITICAL")
        if is_gui_widget_available(_root_ref_for_log) and isinstance(_root_ref_for_log, tk.Tk):
            messagebox.showwarning("Configuração da API Gemini Pendente",
                                 "A API Key do Google Gemini não foi configurada corretamente ou é inválida.\n\n"
                                 "A funcionalidade de extração de dados de PDFs estará DESABILITADA.\n\n"
                                 "Por favor, verifique o arquivo '.env' e os logs para corrigir o problema.",
                                 parent=_root_ref_for_log)
        if is_gui_widget_available(botao_analisar) and isinstance(botao_analisar, ttk.Button):
            botao_analisar.config(state=tk.DISABLED)
            log_to_gui("Botão 'Analisar PDF' desabilitado devido à falta de configuração da API Gemini.", "WARNING")

    root.mainloop()